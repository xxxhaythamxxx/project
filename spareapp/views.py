from typing import List
from django.forms import NullBooleanField
from django.shortcuts import get_object_or_404, redirect, render, HttpResponse, HttpResponseRedirect
# import sqlalchemy
# from sqlalchemy import null
from .models import *
from django.views import View
from .cart import *
import json
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from .forms import UserRegisterForm
from random import seed
from random import random
import random
import datetime
from datetime import date
from datetime import datetime, timezone
from datetime import timedelta
from django.contrib.auth.models import User, Permission
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
# import numpy as np
# from flask_sqlalchemy import SQLAlchemy

# Create your views here.

# Código para saber si usa el input o el filtro

import locale
# locale.setlocale(locale.LC_ALL, 'es_CR.UTF-8')
from django.db.models import Q

def contLogin(request):

    if request.method == "POST":

        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("contDay")
        else:
            return redirect("contLogin")

    return render(request,"spareapp/contLogin.html")

def contLogout(request):
    logout(request)
    return redirect("contDay")

def selectf(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all().order_by("spare_name","spare_code")
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}


    if request.method=="POST":
        list = request.POST.getlist('toAdd')
        delist = request.POST.getlist("toDel")
        if delist:
            carrito = Cart(request)
            for a in delist:                
                spare_part = get_object_or_404(spare, id = a)
                carrito.remove(spare_part)
        if list:
            carrito = Cart(request)
            for a in list:                
                spare_part = get_object_or_404(spare, id = a)
                carrito.add(spare_part)
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    if request.method=="GET":
        search=request.GET.get("engine_id")
        # Si usaron el filter
        if search:
            # Valor del modelo del motor
            engModel=request.GET.get("engine_id")  
            # Valor del carro enviado     
            carManu=request.GET.get("car_id")    
            #Valor del modelo del carro enviado       
            carModel=request.GET.get("car_model_id")    
            comp=spare.objects.filter(engine_info__engine_ide__icontains=engModel,car_info__car_manufacturer__icontains=carManu,car_info__car_model__icontains=carModel).order_by("id")  # Creo un Query de spare que tenga el valor del motor pasado
            engcomp=engine.objects.filter(engine_ide__icontains=engModel,car_engine_info__car_manufacturer__icontains=carManu)
            dic.update({"carManu":carManu,"carModel":carModel,"spare":comp,"mig":engModel,"engcomp":engcomp})
            return render(request,"spareapp/findfil.html",dic)
        # Si se usa el buscador por código de repuesto
        else:   
            valor=request.GET.get("search")
            bol = False
            if valor:
                if valor=="":
                    return render(request,"spareapp/home.html",dic)
                else:
                    # Compara el codigoRepuesto con valor
                    b=[]
                    vec = valor.split(" ")
                    # comp=spare.objects.filter(spare_code__icontains=vec).order_by("spare_code","spare_brand","spare_name").distinct() 
                    todos=spare.objects.all()
                    ref=reference.objects.all()
                    cars=car.objects.all()
                    engines=engine.objects.all()
                    contVar = 0
                    for t in todos:
                        s=t.spare_code
                        # br=t.spare_brand
                        n=t.spare_name
                        ch=t.car_info
                        for v in vec:
                            if s:
                                out = s.translate(str.maketrans('', '', '.''-'))
                                if valor.upper() in out.upper():
                                    bol = True
                                if v.upper() in out.upper():
                                    bol = True
                                if v.upper() in s.upper():
                                    bol = True
                            if n:
                                out = n.split(" ")
                                for o in out:
                                    if o.upper().startswith(v.upper()):
                                        bol = True
                            for r in ref:
                                sr=r.referenceCode
                                if sr:
                                    out = sr.translate(str.maketrans('', '', '.''-'))
                                    if valor.upper() in out.upper():
                                        if s == r.referenceSpare.spare_code:
                                            bol = True
                                    if v.upper() in out.upper():
                                        if s == r.referenceSpare.spare_code:
                                            bol = True
                            for r in cars:
                                sr=r.transmission
                                man=r.car_manufacturer
                                if sr:
                                    out = sr.translate(str.maketrans('', '', '.''-'))
                                    if valor.upper() in out.upper():
                                        for to in t.car_info.all():
                                            if to.transmission in r.transmission:
                                                bol = True
                                    if v.upper() in out.upper():
                                        for to in t.car_info.all():
                                            if to.transmission in r.transmission:
                                                bol = True
                                if man:
                                    out = man.translate(str.maketrans('', '', '.''-'))
                                    if out.upper().startswith(valor.upper()):
                                        for to in t.car_info.all():
                                            if to.car_manufacturer in r.car_manufacturer:
                                                bol = True
                                    if out.upper().startswith(v.upper()):
                                        for to in t.car_info.all():
                                            if to.car_manufacturer in r.car_manufacturer:
                                                bol = True
                            
                            if bol == True:
                                contVar=contVar+1
                            bol = False
                            
                            # -----------------------------------------------------

                            for r in engines:
                                # transmission
                                sr=r.engine_ide
                                # car_manufacturer
                                # man=r.car_manufacturer
                                if sr:
                                    out = sr.translate(str.maketrans('', '', '.''-'))
                                    if valor.upper() in out.upper():
                                        for to in t.engine_info.all():
                                            if to.engine_ide in r.engine_ide:
                                                bol = True
                                    if v.upper() in out.upper():
                                        for to in t.engine_info.all():
                                            if to.engine_ide in r.engine_ide:
                                                bol = True
                                # if man:
                                #     out = man.translate(str.maketrans('', '', '.''-'))
                                #     if out.upper().startswith(valor.upper()):
                                #         for to in t.car_info.all():
                                #             if to.car_manufacturer in r.car_manufacturer:
                                #                 bol = True
                                #     if out.upper().startswith(v.upper()):
                                #         for to in t.car_info.all():
                                #             if to.car_manufacturer in r.car_manufacturer:
                                #                 bol = True
                            
                            if bol == True:
                                contVar=contVar+1
                            bol = False

                            # --------------------------------------------------------------
                        
                        if contVar == len(vec):
                            b.append(t)
                        contVar = 0
                            
                    b = (set(b))
                    dic.update({"spare":b,"mig":valor,"parameter":"Parameters"})
                    return render(request,"spareapp/find.html",dic)
            else:
                return False


def home(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    
    spCart = spareCart.objects.filter(nameUser=request.user.get_username()).values("spareId","nameUser").distinct().order_by("spareId")
    if request.user.get_username() == "":
        spCart = spareCart.objects.filter(nameUser="AnonymousUser").values("spareId","nameUser").distinct()
    dic={"spCart":spCart,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}


    if request.method=="POST":
        list = request.POST.getlist('toAdd')
        delist = request.POST.getlist("toDel")
        # Si es una lista que se envía para guardarla
        if list:
            # cart = spareCart()
            # Si se va a crear un nuevo carrito
            if request.POST.get("cartNumber") == "":
                randomNum = datetime.now().strftime("%Y%m%d%H%M%S")
                # cart = spareCart()
                # cart.save()
                # form = UserRegisterForm(request.POST)
                # print("Voy a entrar a list")
                for a in list:
                    cart = spareCart()
                    cart.spareId = randomNum
                    # cart1 = spareCart.objects.get(id=cart.id)
                    # print(cart)
                    cart.spareCode = a
                    # cart1.spareCode.add(a)
                    # Si hay iniciada una sesion
                    if request.user.get_username() != "":
                        username = request.user.get_username()
                        cart.nameUser = username
                        # cart1.nameUser.add(username)
                    # Si no hay iniciada una sesion
                    else:
                        cart.nameUser = "AnonymousUser"
                        # cart1.nameUser.add("User")
                    cart.save()
                    # cart1.add(cart1)
            else:
                for a in list:
                    cart = spareCart()
                    if request.user.get_username() == "":
                        cartPrueba=spareCart.objects.filter(spareId=request.POST.get("cartNumber"),spareCode=a,nameUser="AnonymousUser")
                    else:
                        cartPrueba=spareCart.objects.filter(spareId=request.POST.get("cartNumber"),spareCode=a,nameUser=request.user.get_username())
                    if cartPrueba:
                        print("Ya existe")
                    else:
                        cart.spareId = request.POST.get("cartNumber")
                        cart.spareCode = a
                        # Si hay iniciada una sesion
                        if request.user.get_username() != "":
                            username = request.user.get_username()
                            cart.nameUser = username
                        # Si no hay iniciada una sesion
                        else:
                            cart.nameUser = "AnonymousUser"
                        cart.save()
        

        # # Carrito antes
        # if delist:
        #     carrito = Cart(request)
        #     for a in delist:                
        #         spare_part = get_object_or_404(spare, id = a)
        #         carrito.remove(spare_part)
        # if list:
        #     carrito = Cart(request)
        #     for a in list:                
        #         spare_part = get_object_or_404(spare, id = a)
        #         carrito.add(spare_part)
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    if request.method=="GET":

        if selectf(request)==False:
            # alls = spare.objects.all().order_by("spare_name","spare_code","spare_brand").distinct() 
            # from typing import List
            # alls = list(spare.objects.all().order_by("spare_name","spare_code","spare_brand").distinct())
            alls = [x for x in spare.objects.all().order_by("spare_name","spare_code").distinct()]
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            return selectf(request)

def find(request):
    return render(request,"spareapp/find.html")

def sparedetails(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    prev = ""
    nex = ""
    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}


    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            pr1=spare.objects.filter(id=val).order_by("spare_name","spare_code")
            pr=spare.objects.all().order_by("spare_name","spare_code")
            pr2=spare.objects.get(id=val)
            c=0
            for a in allSparesall:
                c=c+1
                if a.id == pr2.id:
                    break
            # previo -------------------------------------------
            if c>1:
                prev = pr[c-2:c-1] 
            if c==1:
                prev = pr[:c] 
            # siguiente -----------------------------------------
            if c == allSparesall.count():
                nex = pr[c-1:c] 
            else:
                nex = pr[c:c+1] 

            dbactual = c
            dbtotal = allSparesall.count()
            dic.update({"spare1":pr1,"spare":pr1,"prev":prev,"next":nex,"dbActual":dbactual,"dbTotal":dbtotal})
            # dic.update({"spare1":pr1,"vector":vector,"dbTotal":dbTotal,"dbActual":dbActual,"spareAux":spareaux,"spare":pr,"spareReference":ar})
            return render(request,"spareapp/sparedetails.html",dic)
    else:
        return selectf(request)
    

# ARREGLAR
def brand(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            # pr=spare.objects.values("id","spare_photo","spare_code","spare_brand","spare_name","car_info__car_manufacturer").filter(spare_brand__icontains=val).distinct()
            # pr=spare.objects.filter(spare_brand__icontains=val).order_by("spare_name","spare_code","spare_brand").distinct()
            # dic.update({"spare":pr,"mig":val,"parameter":"Spare brand"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def name(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            # pr=spare.objects.values("id","spare_photo","spare_code","spare_brand","spare_name","car_info__car_manufacturer").filter(spare_name__icontains=val).distinct()
            pr=spare.objects.filter(spare_name__icontains=val).order_by("spare_name","spare_code").distinct()
            dic.update({"spare":pr,"mig":val,"parameter":"Description"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def trunc(val):
    allen = car.objects.filter(car_manufacturer__icontains=val).values("car_model").order_by("car_model")
    b=[]
    for a in allen:
        a = a["car_model"].split(" ")[0]
        b.append(a+" .")
    b = list(set(b)) 
    return b

def manuf(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            tr = trunc(val)
            pr=car.objects.filter(car_manufacturer__icontains=val)
            dic.update({"car":pr,"mig":val,"tr":tr})
            return render(request,"spareapp/manuf.html",dic)
    else:
        return selectf(request)

def allmodel(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            rep = spare.objects.filter(car_info__car_model__icontains=val).order_by("spare_name","spare_code").distinct()
            dic.update({"spare":rep,"mig":val,"parameter":"Car model"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def allmanu(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    
    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            rep = spare.objects.filter(car_info__car_manufacturer__icontains=val).order_by("spare_name","spare_code").distinct()
            dic.update({"spare":rep,"mig":val,"parameter":"Car manufacturer"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def model(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            pr=engine.objects.filter(car_engine_info__car_model__icontains=val)
            dic.update({"engine":pr,"mig":val,"parameter":"Car model"})
            return render(request,"spareapp/model.html",dic)
    else:
        return selectf(request)

def enginel(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            pr=spare.objects.filter(engine_info__engine_ide__icontains=val).order_by("spare_name","spare_code")
            dic.update({"spare":pr,"parameter":"Engine code","mig":val})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def detail(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        spares = spare.objects.all().order_by("spare_name","spare_code")
        carrito = Cart(request)
        dic.update({'carrito': carrito,'spare':spares})
        return render(request, 'spareapp/detail.html', dic)
    else:
        return selectf(request)

def shape(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            spares = spare.objects.all().filter(shape__icontains=val).order_by("spare_name","spare_code")
            dic.update({'spare':spares,'parameter':'Shape','mig':val})
            return render(request, 'spareapp/find.html',dic)
    else:
        return selectf(request)

def longi(request,val1,val2):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        spares = spare.objects.all().filter(
            dimension__atributeName__icontains=val1,
            dimension__atributeVal=val2).order_by("spare_name","spare_code")
        dic.update({'spare':spares,'parameter':val1,'mig':val2})
        return render(request, 'spareapp/find.html',dic)
    else:
        return selectf(request)

def atributes(request,val1,val2):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        spares = spare.objects.all().filter(
            atribute__atributeName__icontains=val1,
            atribute__atributeVal=val2).order_by("spare_name","spare_code")
        dic.update({'spare':spares,'parameter':val1,'mig':val2})
        return render(request, 'spareapp/find.html',dic)
    else:
        return selectf(request)

def carBrands(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        return render(request,"spareapp/carBrands.html",dic)
    else:
        return selectf(request)

def categoryi(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            # pr=spare.objects.filter(spare_category__category__icontains=val).order_by("spare_name","spare_code","spare_brand").distinct()
            pr=spare.objects.filter(spare_category__category__icontains=val).order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":pr,"mig":val,"parameter":"Category"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def chasis(request, val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    dic={"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            pr=spare.objects.filter(car_info__transmission__icontains=val).order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":pr,"mig":val,"parameter":"Chasis"})
            return render(request,"spareapp/find.html",dic)
    else:
        return selectf(request)

def prev(request,val,val2):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}


    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:
            pr1=spare.objects.filter(id=val2).order_by("spare_name","spare_code")
            pr=spare.objects.all().order_by("spare_name","spare_code")
            pr2=spare.objects.get(id=val2)
            c=0
            for a in allSparesall:
                c=c+1
                if a.id == pr2.id:
                    break
            # previo -------------------------------------------
            if c>1:
                prev = pr[c-2:c-1] 
            if c==1:
                prev = pr[:c] 
            # siguiente -----------------------------------------
            if c == allSparesall.count():
                nex = pr[c-1:c] 
            else:
                nex = pr[c:c+1] 

            dbactual = c
            dbtotal = allSparesall.count()
            dic.update({"spare1":pr1,"spare":pr1,"prev":prev,"next":nex,"dbActual":dbactual,"dbTotal":dbtotal})           
            return render(request,"spareapp/sparedetails.html",dic)
    else:
        return selectf(request)

def next(request,val,val2):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}


    if selectf(request)==False:
        valsearch=request.GET.get("search")
        if valsearch=="":
            alls = spare.objects.all().order_by("spare_name","spare_code").distinct() 
            dic.update({"spare":alls})
            return render(request,"spareapp/home.html",dic)
        else:

            # ar=spare.objects.values("spare_code","car_info__car_manufacturer").filter(spare_code=val).distinct()
            # spareaux = spare.objects.none()
            # spareaux = val2
            # valor=spareaux.lstrip("<QuerySet [spare:")

            # valor2=valor.rstrip("]>")
            # characters = "<>"
            # valor = ''.join( x for x in valor2 if x not in characters)

            # line= valor.replace(" spare: ", "")
            # line= line.replace("{spare: ", "")
            # line=line.rstrip("}")
            
            
            # vector=line.split(",")
            # i=0
            # valAux = 0
            
            # for v in vector:
            #     if val == v.split(" ")[0]:
            #         if i>0:
            #             valAux=(i-1)
            #         else:
            #             valAux=0
            #     i=i+1

            # i=0
            # for v in vector:
            #     if i == valAux:
            #         codeAux=v.split(" ")[0]
            #     i=i+1
            pr1=spare.objects.filter(id=val2).order_by("spare_name","spare_code")
            pr=spare.objects.all().order_by("spare_name","spare_code")
            pr2=spare.objects.get(id=val2)
            c=0
            for a in allSparesall:
                c=c+1
                if a.id == pr2.id:
                    break
            # previo -------------------------------------------
            if c>1:
                prev = pr[c-2:c-1] 
            if c==1:
                prev = pr[:c] 
            # siguiente -----------------------------------------
            if c == allSparesall.count():
                nex = pr[c-1:c] 
            else:
                nex = pr[c:c+1] 

            dbactual = c
            dbtotal = allSparesall.count()
            dic.update({"spare1":pr1,"spare":pr1,"prev":prev,"next":nex,"dbActual":dbactual,"dbTotal":dbtotal}) 
            return render(request,"spareapp/sparedetails.html",dic)
    else:
        return selectf(request)

# def getCar(request):
#     id = request.GET.get('id', '')
#     print(id+"---------------------------------------------------------------------")
#     result = list(car.objects.filter(
#     spare_id=int(id)).values('id', 'car_manufacturer'))
#     return HttpResponse(json.dumps(result), content_type="application/json")

def filldb(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/filldb.html",dic)

def filltrans(request):

    aux = ""
    exist = False
    auxTrans = ""

    if request.method == "POST":

        exist = False

        trans1 = transmission()

        if request.POST.get("transmission"):

            trans1.trans=request.POST.get("transmission")
            auxTrans=transmission.objects.filter(trans=request.POST.get("transmission"))
            if auxTrans:
                exist = True
            aux=request.POST.get("transmission").replace(" ","")
            if len(aux)>0 and exist==False:
                trans1.save()
            else:
                print("Cadena vacía")

        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # return render(request,"spareapp/filltrans.html")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def fillcar(request):

    carAux = []
    allTrans=transmission.objects.all()
    allManu=manufacturer.objects.all().order_by("manufacturer")
    allCars=car.objects.all()
    allCarsDistinct=car.objects.values("car_manufacturer").distinct()
    allModelsDistinct=car.objects.values("car_model").distinct()
    auxiliarCar = None
    dic = {"allManu":allManu,"allTrans":allTrans,"allCars":allCars,"allCarsDistinct":allCarsDistinct,"allModelsDistinct":allModelsDistinct}

    if request.method == "POST":

        manuAux = request.POST.get("manufactur")
        manuAux2 = request.POST.get("manufactur")
        modelAux = request.POST.get("model")
        yearFromAux = request.POST.get("yearfro")
        yeartToAux = request.POST.get("yeart")
        chasiAux = request.POST.get("chasi")
        cartoreg = request.POST.getlist("engcartoReg")
        cartopass = request.POST.getlist("engcartoPass")
        carAux = []
        idAux = ""
        targetCar = ""

        auxiliarCar = car.objects.filter(car_manufacturer__manufacturer=manuAux,car_model=modelAux)

        if yearFromAux:
            auxiliarCar = auxiliarCar.filter(carfrom=yearFromAux)
        
        if chasiAux:
            auxiliarCar = auxiliarCar.filter(chasis=chasiAux)

        if auxiliarCar:

            pass

        else:

            if manufacturer.objects.filter(manufacturer__icontains=manuAux2):
                manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)
            else:
                manuNew = manufacturer()
                manuNew.manufacturer = manuAux2
                manuNew.save()

            manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)

            car1 = car()
            car1.car_manufacturer = manuAux
            if modelAux:
                car1.car_model = modelAux
            if yearFromAux:
                car1.carfrom = yearFromAux
            if chasiAux:
                car1.chasis = chasiAux

            car1.save()

        if request.POST.get("id") == "secondForm":
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        else:
            return render(request,"spareapp/fillcar.html",dic)
        
    else:
        return render(request,"spareapp/fillcar.html",dic)

def editcar(request,val):
    
    allCars=car.objects.all().order_by("car_manufacturer")
    allTrans=transmission.objects.all().order_by("trans")
    allManu=manufacturer.objects.all().order_by("manufacturer")
    allModelsDistinct=car.objects.values("car_model").distinct()
    sparefind=car.objects.filter(id=val)
    auxTrans=transmission.objects.filter(car__id=val)
    car1 = car.objects.get(id=val)
    allCarsAll=car.objects.all().values("id","transmission__id").order_by("car_manufacturer")

    dic={"allModelsDistinct":allModelsDistinct,"allManu":allManu,"allCarsAll":allCarsAll,"auxTrans":auxTrans,"allTrans":allTrans,"val":val,"sparefind":sparefind,"allCars":allCars}
    if request.method == "POST":

        manuAux = request.POST.get("manufactur")
        manuAux2 = request.POST.get("manufactur")
        modelAux = request.POST.get("model")
        yearFromAux = request.POST.get("yearfro")
        yeartToAux = request.POST.get("yeart")
        chasiAux = request.POST.get("chasi")
        transAux = request.POST.get("trans")
        cartoreg = request.POST.getlist("engcartoReg")
        cartopass = request.POST.getlist("engcartoPass")
        carAux = []
        idAux = ""
        targetCar = ""

        if manufacturer.objects.filter(manufacturer__icontains=manuAux2):
            manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)
        else:
            manuNew = manufacturer()
            manuNew.manufacturer = manuAux2
            manuNew.save()

        manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)

        car1.car_manufacturer = manuAux
        car1.car_model=modelAux

        if yearFromAux:
            print("Guardaría")
            car1.carfrom=yearFromAux
        else:
            print("No guardaría")
            car1.carfrom=None

        # if yeartToAux:
        #     car1.carto=yeartToAux
        # else:
        #     car1.carto=None

        # if transAux.isspace():
        #     car1.transmission=""
        # else:
        #     if transAux:
        #         car1.transmission=transAux
        #     else:
        #         car1.transmission=""

        if chasiAux.isspace():
            car1.chasis=""
        else:
            if chasiAux:
                car1.chasis=chasiAux
            else:
                car1.chasis=""
        
        car1.save()

        # spAux = car.objects.filter(id=val)

        # for sp in spAux:
        #     for a in sp.transmission.all():
        #         car1.transmission.remove(a.id)

        # if cartoreg == []:
        #     for c in allTrans:
        #         bandt = False
        #         for ca in cartopass:
        #             if str(c.trans) == str(ca):
        #                 bandt = True
        #         if bandt == False:
        #             carAux.append(c.trans)

        # if cartoreg == []:
        #     for sp in carAux:
        #         for c in allTrans:
        #             if str(c.trans) == str(sp):
        #                 idAux = c.id
        #         targetCar = transmission.objects.get(id=idAux)
        #         car1.transmission.add(targetCar)
        # else:
        #     for sp in cartoreg:
        #         for c in allTrans:
        #             if str(c.trans) == str(sp):
        #                 idAux = c.id
        #         targetCar = transmission.objects.get(id=idAux)
        #         car1.transmission.add(targetCar)

        return render(request,"spareapp/listcar.html",dic)
    else:
        return render(request,"spareapp/editcar.html",dic)


def fillengine(request):

    allManu=manufacturer.objects.all().order_by("manufacturer")
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    allEnginesMake=engine.objects.all().exclude(engine_manufacturer=None).values("id","engine_manufacturer").distinct()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    dic={"allManu":allManu,"allEnginesMake":allEnginesMake,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if request.method == "POST":
        # print("Entra en el POST de fillengine")
        engine1 = engine()

        manuAux2 = request.POST.get("manufactur")

        if manufacturer.objects.filter(manufacturer__icontains=manuAux2):
            manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)
        else:
            manuNew = manufacturer()
            manuNew.manufacturer = manuAux2
            manuNew.save()

        manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)

        engine1.engine_manufacturer = manuAux
        # print(request.POST.get("manufactur"))
        engine1.engine_l = request.POST.get("litresfill")
        engine1.engine_ide = request.POST.get("codefill")
        engine1.engine_type = request.POST.get("typefill")
        if request.POST.get("valvefill")=="":
            pass
        else:
            engine1.engine_cylinder = request.POST.get("valvefill")
        if request.POST.get("pistonsfill")=="":
            pass
        else:
            engine1.engine_pistons = request.POST.get("pistonsfill")
        engine1.save()
        # print(request.POST.getlist("engcartoReg"))
        cartoreg = request.POST.getlist("engcartoReg")
        carAux = []
        cartopass = request.POST.getlist("engcartoPass")
        # print(cartopass)
        # print(request.POST)

        if cartoreg == []:
            for c in allCars:
                bandt = False
                for ca in cartopass:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    carAux.append(c)

        if cartoreg == []:
            for sp in carAux:
                for c in allCars:
                    if str(c).replace(" ","") == str(sp).replace(" ",""):
                        idAux = c.id
                targetCar = car.objects.get(id=idAux)
                engine1.car_engine_info.add(targetCar)
        else:
            for sp in cartoreg:
                for c in allCars:
                    if str(c).replace(" ","") == str(sp).replace(" ",""):
                        idAux = c.id
                targetCar = car.objects.get(id=idAux)
                engine1.car_engine_info.add(targetCar)


        # for sp in cartoreg:
        #     for c in allCars:
        #         if str(c) == str(sp):
        #             idAux = c.id
        #     targetCar = car.objects.get(id=idAux)
        #     engine1.car_engine_info.add(targetCar)

        if request.POST.get("id") == "secondForm":
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
            # return render(request,"spareapp/fillengine.html",dic)
        else:
            return render(request,"spareapp/fillengine.html",dic)

        # return render(request,"spareapp/fillengine.html",dic)
    else:
        return render(request,"spareapp/fillengine.html",dic)

def editengine(request,val):

    allManu=manufacturer.objects.all().order_by("manufacturer")
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    allEnginesMake=engine.objects.all().exclude(engine_manufacturer=None).values("engine_manufacturer").distinct()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all().order_by("car_manufacturer__manufacturer")
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    engine1 = engine.objects.get(id=val)
    auxCar = car.objects.filter(engine__id = val)

    sparefind = engine.objects.filter(id=val)
    dic={"allManu":allManu,"allEnginesMake":allEnginesMake,"auxCar":auxCar,"sparefind":sparefind,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if request.method == "POST":
        engine1.engine_l = request.POST.get("litresfill")
        engine1.engine_ide = request.POST.get("codefill")
        engine1.engine_type = request.POST.get("typefill")

        manuAux2=request.POST.get("manufactur")

        if manufacturer.objects.filter(manufacturer__icontains=manuAux2):
            manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)
        else:
            manuNew = manufacturer()
            manuNew.manufacturer = manuAux2
            manuNew.save()

        manuAux = manufacturer.objects.get(manufacturer__icontains=manuAux2)

        engine1.engine_manufacturer=manuAux

        # if request.POST.get("manufactur"):
        #     engine1.engine_manufacturer=request.POST.get("manufactur")
        if request.POST.get("valvefill")=="":
            engine1.engine_cylinder = None
        else:
            engine1.engine_cylinder = request.POST.get("valvefill")
        if request.POST.get("pistonsfill")=="":
            engine1.engine_pistons = None
        else:
            engine1.engine_pistons = request.POST.get("pistonsfill")
        engine1.save()


# Car info ----------------------------------------------------

        # cartoreg = request.POST.getlist("cartoReg")
        # cartopass = request.POST.getlist("cartoPass")
        # carAux = []

        # if cartoreg:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartoreg:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == True:
        #             carAux.append(c)
        # else:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartopass:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == False:
        #             carAux.append(c)

        # for sp in carAux:
        #     for c in allCars:
        #         if str(c).replace("  ","") == str(sp).replace("  ",""):
        #             idAux = c.id
        #     targetCar = car.objects.get(id=idAux)
        #     spare1.car_info.add(targetCar)
        # print(request.POST)

        cartoreg = request.POST.getlist("engcartoReg")
        cartopass = request.POST.getlist("engcartoPass")
        carAux = []

        spAux = engine.objects.filter(id=val)

        for sp in spAux:
            for a in sp.car_engine_info.all():
                engine1.car_engine_info.remove(a.id)
        
        # if cartoreg == []:
        if cartoreg:
            for c in allCars:
                bandt = False
                for ca in cartoreg:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == True:
                    carAux.append(c)
        else:
            for c in allCars:
                # print("Carro")
                # print(c)
                bandt = False
                for ca in cartopass:
                    # print("cartopass")
                    # print(ca)
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    carAux.append(c)

        # if cartoreg == []:
        #     for sp in carAux:
        #         for c in allCars:
        #             if str(c) == str(sp):
        #                 idAux = c.id
        #         targetCar = car.objects.get(id=idAux)
        #         engine1.car_engine_info.add(targetCar)
        # else:
        for sp in carAux:
            for c in allCars:
                if str(c).replace(" ","") == str(sp).replace(" ",""):
                    idAux = c.id
            targetCar = car.objects.get(id=idAux)
            engine1.car_engine_info.add(targetCar)

        return render(request,"spareapp/listengine.html",dic)
    else:
        return render(request,"spareapp/editengine.html",dic)

def fillspare(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allSubCategories=subcategory.objects.all()
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    allBrands=branddb.objects.all().order_by("brand")
    cartoreg = ""
    cartopass = ""

    spare1 = spare()
    dic={"allBrands":allBrands,"allSubCategories":allSubCategories,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if request.method == "POST":
        # print("POST")
        # print(request.POST)
# Spare ------------------------------------------------------
        cartoreg = ""
        cartopass = ""
        spareAux = spare.objects.filter(spare_code=request.POST.get("cod"))
        if spareAux:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        else:
            pass
        sparrr = request.POST.getlist("toReg")
        spare1.spare_code = request.POST.get("cod")
        spare1.spare_name = request.POST.get("descriptio")
        # spare1.spare_brand = request.POST.get("brand")
        # manuAux2 = request.POST.get("brand")
        # if branddb.objects.filter(brand__icontains=manuAux2):
        #     manuAux = branddb.objects.get(brand__icontains=manuAux2)
        # else:
        #     manuNew = branddb()
        #     manuNew.brand = manuAux2
        #     manuNew.save()

        # manuAux = branddb.objects.get(brand__icontains=manuAux2)

        # spare1.spare_brand = manuAux

        if "phot" in request.FILES:
            spare1.spare_photo = request.FILES['phot']
        else:
            pass

        # if request.POST.get("pricem")=="":
        #     pass
        # else:
        #     spare1.price_m = request.POST.get("pricem")

        # if request.POST.get("priced")=="":
        #     pass
        # else:
        #     spare1.price_d = request.POST.get("priced")

        manuAux2 = request.POST.get("catSelect")

        if manuAux2:
            if category.objects.filter(category=manuAux2):
                manuAux = category.objects.get(category=manuAux2)
            else:
                manuNew = category()
                manuNew.category = manuAux2
                manuNew.save()

            manuAux = category.objects.get(category=manuAux2)
            spare1.spare_category = manuAux
        else:
            spare1.spare_category = None

        

        # if request.POST.get("catSelect") == "":
        #     pass
        # else:
        #     # print("catSelect")
        #     # print(request.POST.get("catSelect"))
        #     category1 = category.objects.get(category=request.POST.get("catSelect"))
        #     spare1.spare_category = category1    

        spare1.save()

# Car info ----------------------------------------------------

        cartoreg = request.POST.getlist("cartoReg")
        cartopass = request.POST.getlist("cartoPass")
        carAux = []

        if cartoreg:
            for c in allCars:
                bandt = False
                for ca in cartoreg:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == True:
                    carAux.append(c)
        else:
            for c in allCars:
                bandt = False
                for ca in cartopass:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    carAux.append(c)

        for sp in carAux:
            for c in allCars:
                if str(c).replace(" ","") == str(sp).replace(" ",""):
                    idAux = c.id
            targetCar = car.objects.get(id=idAux)
            spare1.car_info.add(targetCar)

        # cartoreg = request.POST.getlist("cartoReg")
        # cartopass = request.POST.getlist("cartoPass")
        # carAux = []

        # if cartoreg == []:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartopass:
        #             if str(c).replace(" ","") == str(ca).replace(" ",""):
        #                 bandt = True
        #         if bandt == False:
        #             carAux.append(c)

        # if cartoreg == []:
        #     for sp in carAux:
        #         for c in allCars:
        #             if str(c).replace("  ","") == str(sp).replace("  ",""):
        #                 idAux = c.id
        #         targetCar = car.objects.get(id=idAux)
        #         spare1.car_info.add(targetCar)
        # else:
        #     for sp in cartoreg:
        #         for c in allCars:
        #             if str(c).replace("  ","") == str(sp).replace("  ",""):
        #                 idAux = c.id
        #         targetCar = car.objects.get(id=idAux)
        #         spare1.car_info.add(targetCar)

        

        # if cartoreg:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartoreg:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == True:
        #             carAux.append(c)
        # else:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartopass:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == False:
        #             carAux.append(c)

        # for sp in carAux:
        #     for c in allCars:
        #         if str(c).replace("  ","") == str(sp).replace("  ",""):
        #             idAux = c.id
        #     targetCar = car.objects.get(id=idAux)
        #     spare1.car_info.add(targetCar)
        
        # cartoreg = request.POST.getlist("cartoReg")
        # for sp in cartoreg:
        #     for c in allCars:
        #         if str(c) == str(sp):
        #             # print(c.id)
        #             idAux = c.id
        #     targetCar = car.objects.get(id=idAux)
        #     spare1.car_info.add(targetCar)

# Engine info ----------------------------------------------------

        enginetoReg = request.POST.getlist("enginetoReg")
        enginetopass = request.POST.getlist("enginetoPass")
        # enginetoRegGet = request.GET.getlist("enginetoReg")
        # enginetopassGet = request.GET.getlist("enginetoPass")
        idAux = None
        engineAux = []

        # spAux = spare.objects.filter(spare_code = val)

        # for sp in spAux:
        #     for a in sp.engine_info.all():
        #         spare1.engine_info.remove(a.id)

        if enginetoReg:
            for c in allEngines:
                bandt = False
                for ca in enginetoReg:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == True:
                    engineAux.append(c)
        else:
            for c in allEngines:
                bandt = False
                for ca in enginetopass:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    engineAux.append(c)

        for sp in engineAux:
            for c in allEngines:
                if str(c).replace(" ","") == str(sp).replace(" ",""):
                    idAux = c.id
            targetEngine = engine.objects.get(id=idAux)
            spare1.engine_info.add(targetEngine)
        
        # enginetoreg = request.POST.getlist("enginetoReg")
        # for sp in enginetoreg:
        #     for c in allEngines:
        #         if str(c) == str(sp):
        #             idAux = c.id
        #     targetCar = engine.objects.get(id=idAux)
        #     spare1.engine_info.add(targetCar)

# Vendor -----------------------------------------------

        # vendortoReg = request.POST.getlist("vendortoReg")
        # for sp in vendortoReg:
        #     for c in allVendors:
        #         if str(c) == str(sp):
        #             idAux = c.id
        #     targetVendor = vendor.objects.get(id=idAux)
        #     spare1.spare_vendor.add(targetVendor)

# Spare targets -----------------------------------------------
        for sp in sparrr:
            varId = 0
            aux = sp.split(" ")
            code = (aux[0])

            auxSp = spare.objects.filter(spare_code=code)
            for sp in auxSp:
                varId = sp.id
            targetCode = spare.objects.get(id=varId)
            spare1.spare_spare.add(targetCode)
        
# Reference ------------------------------------------------------
        codesList = request.POST.getlist("refcodes")
        notesList = request.POST.getlist("refcodesnote")
        # quantityList = request.POST.getlist("refquantity")
        i = 0
        for ref in codesList:

            if (ref != "") and (notesList[i] != ""):
                reference1 = reference()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                reference1.referenceSpare = targetSpare
                reference1.referenceCode = ref
                reference1.referenceNote = notesList[i]
                # reference1.cantidad = quantityList[i]
                reference1.save()
            i = i + 1

# Atributes ------------------------------------------------------

        atrtName = request.POST.getlist("atributName")
        atrtVal = request.POST.getlist("atributVal")
        i = 0
        for ref in atrtName:

            if (ref != "") and (atrtVal[i] != ""):
                atribute1 = atribute()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                atribute1.atributeSpare = targetSpare
                atribute1.atributeName = ref
                atribute1.atributeVal = atrtVal[i]
                atribute1.save()
            i = i + 1

# Dimensions ------------------------------------------------------

        dimName = request.POST.getlist("dimensName")
        dimVal = request.POST.getlist("dimensVal")
        # print(dimName)
        # print(dimVal)
        i = 0
        for ref in dimName:

            if (ref != "") and (dimVal[i] != ""):
                dimension1 = dimension()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                dimension1.dimensionSpare = targetSpare
                # print(ref)
                # print(dimVal[i])
                dimension1.atributeName = ref
                dimension1.atributeVal = dimVal[i]
                dimension1.save()
            i = i + 1

        return render(request,"spareapp/fillspare.html",dic)
    else:
        return render(request,"spareapp/fillspare.html",dic)

def subCategoria(request):

    # cobrarPagar = None
    # cate = None
    bandera = False

    val = request.GET.get("val")
    catAux = subcategory.objects.filter(category__category=val)

    if catAux:
        bandera = True

    catAux = list(catAux.values())

    # if request.GET.get("fecha") == "change":

    #     creado = request.GET.get("creado")
    #     creadoAux = datetime.strptime(creado,"%Y-%m-%d")
    #     deadlineDefault=(creadoAux+timedelta(days=30)).date()
    #     actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)

    #     actual = actualAux

    return JsonResponse({'bandera':bandera,'cat':val,'subcat':catAux})

def spareCat(request):

    atributesSpare = None

    val = request.GET.get("val")
    # print(val)
    atAux = atribute.objects.values("atributeName").filter(atributeSpare__spare_category__category=val).distinct()
    print(atAux)
    # print(atAux)
    # catAux = subcategory.objects.filter(category__category=val)
    if atAux:
        print("Hay atributos")
        atributesSpare = list(atAux)
        # atributesSpare = list(atAux.values())
    else:
        print("No hay nada")

    print(atributesSpare)

    # if request.GET.get("fecha") == "change":

    #     creado = request.GET.get("creado")
    #     creadoAux = datetime.strptime(creado,"%Y-%m-%d")
    #     deadlineDefault=(creadoAux+timedelta(days=30)).date()
    #     actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)

    #     actual = actualAux

    return JsonResponse({'cat':val,'atributesSpare':atributesSpare})

def fillcategoryadmin(request):

    allCategories = category.objects.all().order_by("category")
    deleteAux = {}

    if request.method == "POST":

        catAux = request.POST.get("category")

        if category.objects.filter(category=catAux):

            pass

        else:

            cat = category()
            cat.category = catAux
            cat.save()

        for ty in allCategories:

            catAux = spare.objects.filter(spare_category=ty)
            if catAux:
                deleteAux[ty.id] = "on"
            else:
                deleteAux[ty.id] = "off"

        dic = {"deleteAux":deleteAux,"allCategories":allCategories}

        if request.POST.get("otro"):

            return render(request,"spareapp/fillcategoryadmin.html")

        else:

            return render(request,"spareapp/listcat.html",dic)
    
    return render(request,"spareapp/fillcategoryadmin.html")

def listcat(request):

    allCategories = category.objects.all().order_by("category")

    deleteAux = {}

    for ty in allCategories:

        catAux = spare.objects.filter(spare_category=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allCategories":allCategories}

    return render(request,"spareapp/listcat.html",dic)

# REVISAR
def listbrand(request):

    allBrands = branddb.objects.all().order_by("brand")

    deleteAux = {}

    for ty in allBrands:

        catAux = spare.objects.filter(spare_brand=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allBrands":allBrands}

    return render(request,"spareapp/listbrand.html",dic)

def listmake(request):

    allMake = manufacturer.objects.all().order_by("manufacturer")

    deleteAux = {}

    for ty in allMake:

        catAux = car.objects.filter(car_manufacturer=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allMake":allMake}

    return render(request,"spareapp/listmake.html",dic)

def editecategory(request,val):

    catAux = category.objects.get(id=val)
    allCategories = category.objects.all().order_by("category")
    deleteAux = {}
    dic = {"category":catAux}

    if request.method == "POST":

        cat = request.POST.get("category")
        catAux.category = cat
        catAux.save()

        for ty in allCategories:

            catAux = spare.objects.filter(spare_category=ty)
            if catAux:
                deleteAux[ty.id] = "on"
            else:
                deleteAux[ty.id] = "off"

        dic = {"deleteAux":deleteAux,"allCategories":allCategories}

        return render(request,"spareapp/listcat.html",dic)

    return render(request,"spareapp/editecategory.html",dic)

# REVISAR
def editebrand(request,val):

    catAux = branddb.objects.get(id=val)
    allBrands = branddb.objects.all().order_by("brand")
    deleteAux = {}
    dic = {"brand":catAux}

    if request.method == "POST":

        cat = request.POST.get("brand")
        catAux.brand = cat
        catAux.save()

        for ty in allBrands:

            catAux = spare.objects.filter(spare_brand=ty)
            if catAux:
                deleteAux[ty.id] = "on"
            else:
                deleteAux[ty.id] = "off"

        dic = {"deleteAux":deleteAux,"allBrands":allBrands}

        return render(request,"spareapp/listbrand.html",dic)

    return render(request,"spareapp/editebrand.html",dic)

def editemake(request,val):

    catAux = manufacturer.objects.get(id=val)
    allMake = manufacturer.objects.all().order_by("manufacturer")
    deleteAux = {}
    dic = {"make":catAux}

    if request.method == "POST":

        cat = request.POST.get("make")
        catAux.manufacturer = cat
        catAux.save()

        for ty in allMake:

            catAux = car.objects.filter(car_manufacturer=ty)
            if catAux:
                deleteAux[ty.id] = "on"
            else:
                deleteAux[ty.id] = "off"

        dic = {"deleteAux":deleteAux,"allMake":allMake}

        return render(request,"spareapp/listmake.html",dic)

    return render(request,"spareapp/editemake.html",dic)

def deletecategory(request,val):

    delete = category.objects.get(id=val)
    delete.delete()
    allCategories = category.objects.all().order_by("category")
    deleteAux = {}

    for ty in allCategories:

        catAux = spare.objects.filter(spare_category=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allCategories":allCategories}

    return render(request,"spareapp/listcat.html",dic)

def deletemake(request,val):

    delete = manufacturer.objects.get(id=val)
    delete.delete()
    allMakes = manufacturer.objects.all().order_by("manufacturer")
    deleteAux = {}

    for ty in allMakes:

        catAux = car.objects.filter(car_manufacturer=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allMakes":allMakes}

    return render(request,"spareapp/listmake.html",dic)

# REVISAR
def deletebrand(request,val):

    delete = branddb.objects.get(id=val)
    delete.delete()
    allBrands = branddb.objects.all().order_by("brand")
    deleteAux = {}

    for ty in allBrands:

        catAux = spare.objects.filter(spare_brand=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allBrands":allBrands}

    return render(request,"spareapp/listbrand.html",dic)

def listsubcat(request):

    allCategories = category.objects.all().order_by("category")
    allSubCategories = subcategory.objects.all().order_by("subcategory")
    deleteAux = {}

    for ty in allSubCategories:

        catAux = spare.objects.filter(spare_subcategory=ty)
        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"deleteAux":deleteAux,"allCategories":allCategories,"allSubCategories":allSubCategories}

    return render(request,"spareapp/listsubcat.html",dic)

def fillsubcategory(request):

    allCategories = category.objects.all()
    dic = {"allCategories":allCategories}

    if request.method == "POST":

        catAux = request.POST.get("category")
        subAux = request.POST.get("subcategory")
        categoryAux = category.objects.get(category=catAux)

        subcat = subcategory()
        subcat.category = categoryAux
        subcat.subcategory = subAux
        subcat.save()

        if request.POST.get("otro"):

            return render(request,"spareapp/fillsubcategory.html",dic)

        else:

            allCategories = category.objects.all().order_by("category")
            allSubCategories = subcategory.objects.all().order_by("subcategory")
            deleteAux = {}

            for ty in allSubCategories:

                catAux = spare.objects.filter(spare_subcategory=ty)
                if catAux:
                    deleteAux[ty.id] = "on"
                else:
                    deleteAux[ty.id] = "off"

            dic = {"deleteAux":deleteAux,"allCategories":allCategories,"allSubCategories":allSubCategories}

            return render(request,"spareapp/listsubcat.html",dic)
    
    return render(request,"spareapp/fillsubcategory.html",dic)

def deletesubcategory(request,val):

    subAux = subcategory.objects.get(id=val)

    subAux.delete()

    allCategories = category.objects.all().order_by("category")
    allSubCategories = subcategory.objects.all().order_by("subcategory")
    deleteAux = {}

    for ty in allSubCategories:

        catAux = spare.objects.filter(spare_subcategory=ty)

        if catAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    dic = {"subcategory":subAux,"deleteAux":deleteAux,"allCategories":allCategories,"allSubCategories":allSubCategories}

    return render(request,"spareapp/listsubcat.html",dic)

def editesubcategory(request,val):

    allCategories = category.objects.all().order_by("category")
    allSubCategories = subcategory.objects.all().order_by("subcategory")
    subAux = subcategory.objects.get(id=val)

    if request.method == "POST":

        catAux1 = request.POST.get("category")
        catAux = category.objects.get(category=catAux1)
        subAux2 = request.POST.get("subcategory")

        subAux.category = catAux
        subAux.subcategory = subAux2
        subAux.save()

        deleteAux = {}
        allCategories = category.objects.all().order_by("category")
        allSubCategories = subcategory.objects.all().order_by("subcategory")

        for ty in allSubCategories:

            catAux = spare.objects.filter(spare_subcategory=ty)
            if catAux:
                deleteAux[ty.id] = "on"
            else:
                deleteAux[ty.id] = "off"

        dic = {"deleteAux":deleteAux,"allCategories":allCategories,"allSubCategories":allSubCategories}

        return render(request,"spareapp/listsubcat.html",dic)

    dic = {"allCategories":allCategories,"subcategory":subAux}

    return render(request,"spareapp/editesubcategory.html",dic)

def editspare(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allSubCategories=subcategory.objects.all()
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    refAux = reference.objects.filter(referenceSpare__id=val)

    spare1 = spare.objects.get(id=val)
    sparefind=spare.objects.filter(id=val)
    # print(sparefind[0].spare_category)
    # allSubCategories=subcategory.objects.filter(category=sparefind[0].spare_category)
    # Con sparefind puedo enviar el mismo valor al final para seguir editando

    auxCar = car.objects.filter(spare__id = val)
    auxEnegine = engine.objects.filter(spare__id = val)
    auxVendor = vendor.objects.filter(spare__id = val)
    auxAtributes = atribute.objects.filter(atributeSpare__id = val)
    auxDimensions = dimension.objects.filter(dimensionSpare__id = val)

    dic={"allSubCategories":allSubCategories,"auxDimensions":auxDimensions,"auxAtributes":auxAtributes,"auxVendor":auxVendor,"auxEnegine":auxEnegine,"auxCar":auxCar,"refAux":refAux,"val":val,"sparefind":sparefind,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if request.method == "POST":
        
        # Spare ------------------------------------------------------
        sparrr = request.POST.getlist("toReg")
        # print("sparrr")
        # print(sparrr)
        sparrrPasar = request.POST.getlist("toPass")
        print("To Reg")
        print(request.POST.getlist("toReg"))
        print("To Pass")
        print(request.POST.getlist("toPass"))
        sparrrAux = []
        varcont = 0
        if sparrr == []:
            for sp in allSparesall:
                bandt = False
                for sp2 in sparrrPasar:
                    if str(sp.spare_code).replace(" ","") == str(sp2).replace(" ",""):
                        varcont = varcont + 1
                        bandt = True
                if bandt == False:
                    sparrrAux.append(sp.spare_code)
        # print("sparrrAux")
        # print(sparrrAux)
        spare1.spare_code = request.POST.get("cod")
        if request.POST.get("descriptio") == "":
            spare1.spare_name = None
        else:
            spare1.spare_name = request.POST.get("descriptio")

        # if request.POST.get("brand") == "":
        #     spare1.spare_brand = None
        # else:
        #     spare1.spare_brand = request.POST.get("brand")
        
        if "phot" in request.FILES:
            spare1.spare_photo = request.FILES['phot']
        else:
            pass

        # Cagetory ----------------------------------------------------

        manuAux2 = request.POST.get("catSelect")

        if manuAux2:
            if category.objects.filter(category=manuAux2):
                # print("QUERY")
                # print(category.objects.filter(category=manuAux2))
                manuAux = category.objects.get(category=manuAux2)
            else:
                manuNew = category()
                manuNew.category = manuAux2
                manuNew.save()

            manuAux = category.objects.get(category=manuAux2)
            spare1.spare_category = manuAux
        else:
            spare1.spare_category = None

        # if request.POST.get("catSelect") == "":
        #     spare1.spare_category = None
        # else:
        #     category1 = category.objects.get(category=request.POST.get("catSelect"))
        #     spare1.spare_category = category1

        # if request.POST.get("subcatSelect") == "" or request.POST.get("subcatSelect") == None:           
        #     spare1.spare_subcategory = None
        # else:
        #     category1 = subcategory.objects.get(subcategory=request.POST.get("subcatSelect"))            
        #     spare1.spare_subcategory = category1

        # Prices ------------------------------------------------------

        # if request.POST.get("pricem")=="":
        #     spare1.price_m = None
        # else:
        #     spare1.price_m = request.POST.get("pricem")

        # if request.POST.get("priced")=="":
        #     spare1.price_d = None
        # else:
        #     spare1.price_d = request.POST.get("priced")
        spare1.save()

        # Car info ----------------------------------------------------

        cartoreg = request.POST.getlist("cartoReg")
        cartopass = request.POST.getlist("cartoPass")
        carAux = []

        spAux = spare.objects.filter(id=val)

        # for sp in spAux:
        #     for a in sp.car_engine_info.all():
        #         engine1.car_engine_info.remove(a.id)

        for sp in spAux:
            for a in sp.car_info.all():
                spare1.car_info.remove(a.id)
        
        # if cartoreg:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartoreg:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == True:
        #             carAux.append(c)
        # else:
        #     for c in allCars:
        #         bandt = False
        #         for ca in cartopass:
        #             # print("cartopass")
        #             # print(ca)
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == False:
        #             carAux.append(c)

        if cartoreg:
            for c in allCars:
                bandt = False
                for ca in cartoreg:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == True:
                    carAux.append(c)
        else:
            for c in allCars:
                bandt = False
                for ca in cartopass:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    carAux.append(c)

        # for sp in carAux:
        #     for c in allCars:
        #         if str(c).replace("  ","") == str(sp).replace("  ",""):
        #             idAux = c.id
        #     targetCar = car.objects.get(id=idAux)
        #     engine1.car_engine_info.add(targetCar)

        for sp in carAux:
            for c in allCars:
                if str(c).replace(" ","") == str(sp).replace(" ",""):
                    idAux = c.id
            targetCar = car.objects.get(id=idAux)
            spare1.car_info.add(targetCar)

        # Engine info ----------------------------------------------------

        # enginetoReg = None
        enginetoReg = request.POST.getlist("enginetoReg")
        enginetopass = request.POST.getlist("enginetoPass")
        enginetoRegGet = request.GET.getlist("enginetoReg")
        enginetopassGet = request.GET.getlist("enginetoPass")
        idAux = None
        engineAux = []

        spAux = spare.objects.filter(id = val)

        for sp in spAux:
            for a in sp.engine_info.all():
                spare1.engine_info.remove(a.id)

        # if enginetoReg == []:
        if enginetoReg:
            for c in allEngines:
                bandt = False
                for ca in enginetoReg:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == True:
                    engineAux.append(c)
        else:
            for c in allEngines:
                bandt = False
                for ca in enginetopass:
                    if str(c).replace(" ","") == str(ca).replace(" ",""):
                        bandt = True
                if bandt == False:
                    engineAux.append(c)

        # if enginetoReg == []:
        #     pass
        #     # for sp in engineAux:
        #     #     for c in allEngines:
        #     #         if str(c).replace("  ","") == str(sp).replace("  ",""):
        #     #             idAux = c.id
        #     #     targetEngine = engine.objects.get(id=idAux)
        #     #     spare1.engine_info.add(targetEngine)
        # else:
        #     for sp in enginetoReg:
        #         for c in allEngines:
        #             print("Compara")
        #             print(str(c).replace("  ",""))
        #             print(str(sp).replace("  ",""))
        #             if str(c).replace("  ","") == str(sp).replace("  ",""):
        #                 idAux = c.id
        #         targetEngine = engine.objects.get(id=idAux)
        #         spare1.engine_info.add(targetEngine)

        for sp in engineAux:
            for c in allEngines:
                if str(c).replace(" ","") == str(sp).replace(" ",""):
                    idAux = c.id
            targetEngine = engine.objects.get(id=idAux)
            spare1.engine_info.add(targetEngine)

        # Vendors ----------------------------------------------------

        # vendortoreg = request.POST.getlist("vendortoReg")
        # vendortopass = request.POST.getlist("vendortoPass")
        # vendorAux = []

        # spAux = spare.objects.filter(spare_code = val)

        # for sp in spAux:
        #     for a in sp.spare_vendor.all():
        #         spare1.spare_vendor.remove(a.id)

        # if vendortoreg:
        #     for c in allVendors:
        #         bandt = False
        #         for ca in vendortoreg:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == True:
        #             vendorAux.append(c)
        # else:
        #     for c in allVendors:
        #         bandt = False
        #         for ca in vendortopass:
        #             if str(c).replace("  ","") == str(ca).replace("  ",""):
        #                 bandt = True
        #         if bandt == False:
        #             vendorAux.append(c)

        # for sp in vendorAux:
        #     for c in allVendors:
        #         if str(c).replace("  ","") == str(sp).replace("  ",""):
        #             idAux = c.id
        #     targetVendor = vendor.objects.get(id=idAux)
        #     spare1.spare_vendor.add(targetVendor)
        
        # if vendortoreg == []:
        #     for c in allVendors:
        #         bandt = False
        #         for ca in vendortopass:
        #             if str(c) == str(ca):
        #                 bandt = True
        #         if bandt == False:
        #             vendorAux.append(c)

        # if vendortoreg == []:
        #     for sp in vendorAux:
        #         for c in allVendors:
        #             if str(c) == str(sp):
        #                 idAux = c.id
        #         targetVendor = vendor.objects.get(id=idAux)
        #         spare1.spare_vendor.add(targetVendor)
        # else:
        #     for sp in vendortoreg:
        #         for c in allVendors:
        #             if str(c) == str(sp):
        #                 idAux = c.id
        #         targetVendor = vendor.objects.get(id=idAux)
        #         spare1.spare_vendor.add(targetVendor)

        # Spare targets -----------------------------------------------

        spAux = spare.objects.filter(id=val)
        print("spAux")
        print(spAux)

        for sp in spAux:
            for a in sp.spare_spare.all():
                spare1.spare_spare.remove(a.id)

        # print("sparrr: ")
        # print(sparrr)
        # print("sparrrAux: ")
        # print(sparrrAux)
        if sparrr == []:
            for sp in sparrrAux:
                varId = 0
                aux = str(sp).split(" ")
                code = (aux[0])

                auxSp = spare.objects.filter(spare_code=code)
                for sp in auxSp:
                    varId = sp.id
                targetCode = spare.objects.get(id=varId)
                spare1.spare_spare.add(targetCode)
        else:
            print("Entra en contrario")
            print("sparr")
            print(sparrr)
            for sp in sparrr:
                varId = 0
                aux = sp.split(" ")
                print("Valor de aux")
                print(aux)
                code = (aux[0])
                print("Valor de code")
                print(code)

                auxSp = spare.objects.filter(spare_code=code)
                print("auxSp")
                print(auxSp)
                for sp in auxSp:
                    varId = sp.id
                targetCode = spare.objects.get(id=varId)
                spare1.spare_spare.add(targetCode)
        
        # Reference ------------------------------------------------------
        codesList = request.POST.getlist("refcodes")
        notesList = request.POST.getlist("refcodesnote")
        # quantityList = request.POST.getlist("refquantity")

        auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
        if auxSp:
            reference1 = reference.objects.filter(referenceSpare__spare_code=request.POST.get("cod"))
            reference1.delete()
        else:
            # print("No consigue codigo")
            reference1 = reference()

        i = 0
        for ref in codesList:

            if (ref != "") and (notesList[i] != ""):
                reference1 = reference()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                reference1.referenceSpare = targetSpare
                reference1.referenceCode = ref
                reference1.referenceNote = notesList[i]
                # if quantityList[i]:
                #     reference1.cantidad = quantityList[i]
                # else:
                #     reference1.cantidad = 0
                reference1.save()
            i = i + 1

        # Atributes ------------------------------------------------------

        atrtName = request.POST.getlist("atributName")
        atrtVal = request.POST.getlist("atributVal")

        auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
        if auxSp:
            atribute1 = atribute.objects.filter(atributeSpare__spare_code=request.POST.get("cod"))
            atribute1.delete()
        else:
            # print("No consigue codigo")
            atribute1 = reference()

        i = 0
        for ref in atrtName:

            if (ref != "") and (atrtVal[i] != ""):
                atribute1 = atribute()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                atribute1.atributeSpare = targetSpare
                # print(ref)
                # print(atrtVal[i])
                atribute1.atributeName = ref
                atribute1.atributeVal = atrtVal[i]
                atribute1.save()
            i = i + 1

        # Dimensions ------------------------------------------------------

        dimName = request.POST.getlist("dimensName")
        dimVal = request.POST.getlist("dimensVal")
        # print(dimName)
        # print(dimVal)
        # print(type(dimVal[0]))

        auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
        if auxSp:
            dimension1 = dimension.objects.filter(dimensionSpare__spare_code=request.POST.get("cod"))
            dimension1.delete()
        else:
            dimension1 = dimension()

        i = 0
        for ref in dimName:

            if (ref != "") and (dimVal[i] != ""):
                dimension1 = dimension()
                auxSp = spare.objects.filter(spare_code=request.POST.get("cod"))
                for sp in auxSp:
                    varId = sp.id
                targetSpare = spare.objects.get(id=varId)
                dimension1.dimensionSpare = targetSpare
                dimension1.atributeName = ref
                dimension1.atributeVal = dimVal[i]
                dimension1.save()
            i = i + 1

        allSparesall = spare.objects.all()

        dic = {"allSparesall":allSparesall}
        
        # return render(request,"spareapp/listspare.html",dic)
        return redirect("listspare")
    else:
        return render(request,"spareapp/editspare.html",dic)


def listspare(request):

    allSparesall=spare.objects.all().order_by("spare_code")
    allVendors=vendor.objects.all()
    dic={"allSparesall":allSparesall,"allVendors":allVendors}

    return render(request,"spareapp/listspare.html",dic)

def listList(request):

    cantidad = {}
    acum = 0

    allSparesall=spare.objects.all().order_by("spare_code")
    allVendors=vendor.objects.all()
    allReferences=reference.objects.all().order_by("referenceCode")
    
    if allSparesall:

        for sp in allSparesall:

            acum = 0

            if allReferences:

                for ref in allReferences:

                    if ref.referenceSpare == sp:
                        
                        if ref.cantidad:
                            
                            acum = acum + ref.cantidad

                        else:
                            
                            acum = 0

                        # print(ref.referenceSpare)
           
            cantidad[sp.id] = [acum]

    dic={"cantidad":cantidad,"allReferences":allReferences,"allSparesall":allSparesall,"allVendors":allVendors}

    return render(request,"spareapp/listList.html",dic)

def listengine(request):
    allEngines=engine.objects.all().order_by("engine_manufacturer__manufacturer","engine_ide","engine_l")
    allCarsEngines=car.objects.all().values("id","engine__id","car_manufacturer","car_model","chasis","carfrom","carto").order_by("car_manufacturer")
    allTrans=transmission.objects.all().values("trans","car__id")
    dic={"allEngines":allEngines,"allCarsEngines":allCarsEngines,"allTrans":allTrans}

    return render(request,"spareapp/listengine.html",dic)

def listcar(request):
    allCars=car.objects.all().order_by("car_manufacturer__manufacturer","car_model","carfrom","chasis")
    allCarsAll=car.objects.all().values("id","transmission__id").order_by("car_manufacturer__manufacturer","car_model","carfrom","chasis")
    allTrans=transmission.objects.all()
    allEngines=engine.objects.all().order_by("engine_ide")
    allEnginesCars=engine.objects.all().values("id","car_engine_info__id","engine_ide","engine_l","engine_type","engine_pistons").order_by("engine_ide")
    dic={"allCarsAll":allCarsAll,"allCars":allCars,"allTrans":allTrans,"allEngines":allEngines,"allEnginesCars":allEnginesCars}

    return render(request,"spareapp/listcar.html",dic)

def deleteengine(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    engine1 = engine.objects.get(id=val)
    engine1.delete()
    # sparefind=spare.objects.filter(spare_code=val)
    dic={"val":val,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/listengine.html",dic)

def deletespare(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    spare1 = spare.objects.get(id=val)
    spare1.delete()
    # sparefind=spare.objects.filter(spare_code=val)
    dic={"val":val,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/listspare.html",dic)
    # return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def deletecar(request,val):
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    car1 = car.objects.get(id=val)
    car1.delete()
    # sparefind=spare.objects.filter(spare_code=val)
    dic={"val":val,"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/listcar.html",dic)

def fillcategory(request):

    # print("Entra en fillCategory")

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    # print(request.POST)
    category1 = category()
    # print("categor")
    # print(request.POST.get("categor"))
    category1.category = request.POST.get("categor")
    category1.save()

    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/fillspare.html",dic)

def fillvendor(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    # print(request.POST.get("vendo"))
    vendor1 = vendor()    
    vendor1.vendorName = request.POST.get("vendo")
    vendor1.save()
    # category1 = category()
    # category1.category = request.POST.get("categor")
    # category1.save()

    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/fillspare.html",dic)

def importCar(request):
    
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    checkear = 0
    linea=""
    linea2=""
    lineaTrans=""
    litros=""
    codigo=""
    valve=""
    motor=""
    verificador = False
    auxMake = None
    auxModel = None
    auxFrom = ""
    auxTo = ""
    auxChasis = None
    auxYear = None
    lineaTrans = ""
    trans1 = ""
    targetCar = ""
    auxCar = None
    aux = None

    if "phot" in request.FILES:
        checkear = 0

        FILE_PATH = request.FILES['phot']
        workbook = load_workbook(FILE_PATH)
        sheet = workbook.active

        maxCol = []

        for col in sheet.iter_rows():
            maxCol.append(col)
        
        i=0
        poscar_manufacturer = -1
        poscar_model = -1
        poscarfrom = -1
        poscarto = -1
        postyear = -1
        postchasis = -1
        postransmission = -1
        postengine = -1
        cont = 0
        for fil in maxCol:
            j=0
            for col in fil:
                if col.value == "MAKE":
                    poscar_manufacturer = j
                    cont = cont + 1
                if col.value == "MODEL":
                    poscar_model = j
                    cont = cont + 1
                if col.value == "YEAR":
                    postyear = j
                    cont = cont + 1
                if col.value == "CHASIS":
                    postchasis = j
                    cont = cont + 1
                j=j+1
            i=i+1
        if cont > 0:
            print("Contador mayor a 0")
            i=0
            for fil in maxCol:
                j=0
                for col in fil:
                    if i>0:
                        # COLUMNA MANUFACTURER
                        if j == poscar_manufacturer:
                            # print(col.value)
                            auxMake = col.value
                            # if car.objects.filter(car_manufacturer=col.value):
                            #     checkear = checkear + 1
                        # COLUMNA MODEL
                        if j == poscar_model:
                            auxModel = col.value
                            # if car.objects.filter(car_model=col.value):
                            #     checkear = checkear + 1
                        # COLUMNA YEAR
                        if j == postyear:
                            auxYear = col.value
                        # COLUMNA CHASIS
                        if j == postchasis:
                            auxChasis = col.value
                            # if car.objects.filter(chasis=col.value):
                            #     checkear = checkear + 1
                        
                    j=j+1
                
                if auxMake:
                    # print(auxMake)
                    aux = manufacturer.objects.filter(manufacturer=auxMake)
                    if aux:
                        pass
                    else:
                        newManu = manufacturer()
                        newManu.manufacturer = auxMake
                        newManu.save()
                    aux = manufacturer.objects.get(manufacturer=auxMake)
                    auxCar=car.objects.filter(car_manufacturer__manufacturer=auxMake)
                        # print("Car: "+str(auxMake))
                    # car1.car_manufacturer = auxMake
                if auxModel:
                    # print(auxModel)
                    # aux = car.objects.filter(car_model=auxModel)
                    auxCar=auxCar.filter(car_model=auxModel)
                    if aux:
                        pass
                    else:
                        pass
                    # print("Model: "+str(auxModel))
                    # car1.car_model = auxModel
                if auxYear:
                    # print(auxYear)
                    auxCar=auxCar.filter(carfrom=auxYear)
                    # print("Year: "+str(auxYear))
                    # car1.caryear = auxYear
                if auxChasis:
                    # print(auxChasis)
                    auxCar=auxCar.filter(chasis=auxChasis)
                    # print("Chasis: "+str(auxChasis))
                    # car1.chasis = auxChasis
                if auxCar:
                    pass
                else:
                    if i>0:
                        pass
                        # print(aux)
                        # print(auxModel)
                        # print(auxYear)
                        # print(auxChasis)
                        car1=car()
                        car1.car_manufacturer=aux
                        car1.car_model=auxModel
                        car1.carfrom=auxYear
                        car1.chasis=auxChasis
                        car1.save()
                 
                i=i+1

    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/fillspare.html",dic)

# def importCar(request):
    
#     dim=dimension.objects.values("atributeName").distinct()
#     dim2=dimension.objects.all()
#     atr=atribute.objects.values("atributeName").distinct()
#     atr2=atribute.objects.all()
#     allSparesall=spare.objects.all()
#     allCategories=category.objects.all().order_by("category")
#     allEngines=engine.objects.all()
#     onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
#     allCars=car.objects.all()
#     allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
#     allVendors=vendor.objects.all()
#     ref=reference.objects.all().order_by("referenceSpare")
#     ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
#     checkear = 0
#     linea=""
#     linea2=""
#     lineaTrans=""
#     litros=""
#     codigo=""
#     valve=""
#     motor=""
#     verificador = False
#     auxMake = ""
#     auxModel = ""
#     auxFrom = ""
#     auxTo = ""
#     auxChasis = ""
#     lineaTrans = ""
#     trans1 = ""
#     targetCar = ""

#     if "phot" in request.FILES:
#         checkear = 0

#         FILE_PATH = request.FILES['phot']
#         workbook = load_workbook(FILE_PATH)
#         sheet = workbook.active

#         maxCol = []

#         for col in sheet.iter_rows():
#             maxCol.append(col)
        
#         i=0
#         poscar_manufacturer = -1
#         poscar_model = -1
#         poscarfrom = -1
#         poscarto = -1
#         postchasis = -1
#         postransmission = -1
#         postengine = -1
#         cont = 0
#         for fil in maxCol:
#             j=0
#             for col in fil:
#                 if col.value == "MAKE":
#                     poscar_manufacturer = j
#                     cont = cont + 1
#                 if col.value == "MODEL":
#                     poscar_model = j
#                     cont = cont + 1
#                 if col.value == "FROM":
#                     poscarfrom = j
#                     cont = cont + 1
#                 if col.value == "TO":
#                     poscarto = j
#                     cont = cont + 1
#                 if col.value == "CHASIS":
#                     postchasis = j
#                     cont = cont + 1
#                 if col.value == "TRANSMISION":
#                     postransmission = j
#                     cont = cont + 1
#                 if col.value == "ENGINE(LITERS/CODE/VALVES/OHC)":
#                     postengine = j
#                     cont = cont + 1
#                 j=j+1
#             i=i+1
#         if cont > 0:
#             i=0
#             for fil in maxCol:
#                 j=0
#                 for col in fil:
#                     if i>0:
#                         if j == poscar_manufacturer:
#                             auxMake = col.value
#                             if car.objects.filter(car_manufacturer=col.value):
#                                 checkear = checkear + 1
#                         if j == poscar_model:
#                             auxModel = col.value
#                             if car.objects.filter(car_model=col.value):
#                                 checkear = checkear + 1
#                         if j == poscarfrom:
#                             auxFrom = col.value
#                         if j == poscarto:
#                             auxTo = col.value
#                         if j == postchasis:
#                             auxChasis = col.value
#                             if car.objects.filter(chasis=col.value):
#                                 checkear = checkear + 1
#                         if j == postransmission:
#                             auxTransmision = col.value
#                             if col.value:
#                                 lineaTrans = col.value.split("\n")
#                         if j == postengine:
#                             if col.value:
#                                 linea = col.value.split("\n")
                        
#                     j=j+1
                
#                 if checkear<3 and i>0 and (auxMake or auxModel):
                    
#                     car1 = car()
#                     if auxMake:
#                         car1.car_manufacturer = auxMake
#                     if auxModel:
#                         car1.car_model = auxModel
#                     if auxFrom:
#                         car1.carfrom = auxFrom
#                     if auxTo:
#                         car1.carto = auxTo
#                     if auxChasis:
#                         car1.chasis = auxChasis
#                     car1.save()
#                     for a in lineaTrans:
#                         if transmission.objects.filter(trans=a):
#                             targetCar = transmission.objects.get(trans=a)
#                         else:
#                             trans1 = transmission()
#                             trans1.trans = a
#                             trans1.save()
#                             targetCar = transmission.objects.get(trans=a)
#                         car1.transmission.add(targetCar)
#                 else:
#                     if i>0 and (auxMake or auxModel):
#                         car1 = car.objects.get(car_manufacturer=auxMake,car_model=auxModel,chasis=auxChasis)
#                         if car1.carfrom:
#                             pass
#                         else:
#                             if auxFrom:
#                                 car1.carfrom = auxFrom
#                         car1.save()
#                         for a in lineaTrans:
#                             if transmission.objects.filter(trans=a):
#                                 targetCar = transmission.objects.get(trans=a)
#                             else:
#                                 trans1 = transmission()
#                                 trans1.trans = a
#                                 trans1.save()
#                                 targetCar = transmission.objects.get(trans=a)
#                             car1.transmission.add(targetCar)
#                 checkear = 0
#                 i=i+1
#                 for n in linea:
#                     linea2 = n.split(" ")
#                     if len(linea2)>0:
#                         litros = linea2[0]
#                     else:
#                         litros = None
#                     if len(linea2)>1:
#                         codigo = linea2[1]
#                     else:
#                         codigo = None
#                     if len(linea2)>2:
#                         valve = linea2[2].split("VALVE")[0]
#                     else:
#                         valve = None

#                     if i>0:
#                         if litros:
#                             if codigo:
#                                 if valve:
#                                     if engine.objects.filter(engine_l=litros,engine_ide=codigo,engine_cylinder=valve):
#                                         verificador=True
#                                 else:
#                                     if engine.objects.filter(engine_l=litros,engine_ide=codigo):
#                                         verificador=True
#                             else:
#                                 if valve:
#                                     if engine.objects.filter(engine_l=litros,engine_cylinder=valve):
#                                         verificador=True
#                         else:
#                             if codigo:
#                                 if valve:
#                                     if engine.objects.filter(engine_ide=codigo,engine_cylinder=valve):
#                                         verificador=True
#                                 else:
#                                     if engine.objects.filter(engine_ide=codigo):
#                                         verificador=True
#                             else:
#                                 if valve:
#                                     if engine.objects.filter(engine_cylinder=valve):
#                                         verificador=True

#                         if verificador:
#                             pass
#                         else:
#                             motor = engine()
#                             if litros:
#                                 motor.engine_l = litros
#                             if codigo:
#                                 motor.engine_ide = codigo
#                             if valve:
#                                 motor.engine_cylinder = valve
#                             motor.save()
#                         verificador=False

#     dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

#     return render(request,"spareapp/fillspare.html",dic)

def importEngine(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()

    if "phot" in request.FILES:

        FILE_PATH = request.FILES['phot']

        # FILE_PATH = "prueba2.xlsx"

        workbook = load_workbook(FILE_PATH)
        sheet = workbook.active

        maxCol = []

        for col in sheet.iter_rows():
            maxCol.append(col)
        
        i=0
        engLitres = -1
        engCode = -1
        engType = -1
        engValve = -1
        engPistons = -1
        cont = 0
        for fil in maxCol:
            j=0
            for col in fil:
                if col.value == "litres":
                    engLitres = j
                    cont = cont + 1
                if col.value == "ecode":
                    engCode = j
                    cont = cont + 1
                if col.value == "type":
                    engType = j
                    cont = cont + 1
                if col.value == "valve":
                    engValve = j
                    cont = cont + 1
                if col.value == "pistons":
                    engPistons = j
                    cont = cont + 1
                j=j+1
            i=i+1

        if cont > 0:
            i=0
            for fil in maxCol:
                j=0
                if i>0:
                    engine1 = engine()
                for col in fil:
                    if i>0:
                        if j == engLitres:
                            engine1.engine_l = col.value
                        if j == engCode:
                            engine1.engine_ide = col.value
                        if j == engType:
                            engine1.engine_type = col.value
                        if j == engValve:
                            engine1.engine_cylinder = col.value
                        if j == engPistons:
                            engine1.engine_pistons = col.value
                        engine1.save()
                    j=j+1
                i=i+1

    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/fillspare.html",dic)

def importSpare(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    ref2=reference.objects.values("referenceSpare").order_by("referenceSpare").distinct()
    lineaDesc = ""
    spComp = ""

    if "phot" in request.FILES:

        FILE_PATH = request.FILES['phot']

        workbook = load_workbook(FILE_PATH, data_only = True)
        sheet = workbook.active

        maxCol = []

        for col in sheet.iter_rows():
            maxCol.append(col)
        
        i=0
        spCode = -1
        spTitle = -1
        spPricem = -1
        spPriced = -1
        spReference = -1
        spAttribute = -1
        spDimension = -1
        spNote = -1
        spCategory = -1
        spVendor = -1
        spDesc = ""
        CodeAux = ""
        refName = ""
        refDesc = ""
        AtrName = ""
        AtrDesc = ""
        DimName = ""
        DimDesc = ""
        VenName = ""
        carManu = ""
        carModel = ""
        carChasis = ""
        auxCar = ""
        auxSpare = ""
        car1 = ""
        cont = 0

        for fil in maxCol:
            j=0
            for col in fil:
                if i==0 and col.value:
                    lineaDesc = col.value.split("DESCRIPTION (")
                    if len(lineaDesc)>1:
                        spTitle = j
                        lineaDesc = lineaDesc[1].split(")")[0]
                        spComp = lineaDesc.split(" ")
                        carManu = spComp[0]
                        carModel = spComp[1]
                        carChasis = spComp[2]

                if col.value == "CODE":
                    spCode = j
                    cont = cont + 1
                j=j+1
            i=i+1

        if cont > 0:
            i=0
            

            varRefAux = ""
            varAtrAux = ""
            varDimAux = ""

            for fil in maxCol:
                varRefAux = ""
                varAtrAux = ""
                varDimAux = ""
                bandera = True
                spDesc = None
                j=0
                if i>0:
                    spare1 = spare()
                    
                    for col in fil:

                        if i>0:
                            if j == spCode:
                                CodeAux = col.value.split(" ")[0]
                                if spare.objects.filter(spare_code=CodeAux):
                                    bandera = False
                                spare1.spare_code = CodeAux
                            if j == spTitle:
                                spDesc = col.value

                        j=j+1

                    if spDesc:
                        spare1.spare_name = spDesc
                    if carManu and carModel and carChasis:
                        auxCar = car.objects.filter(car_manufacturer=carManu,car_model=carModel,chasis=carChasis)
                        if auxCar:
                            print("Existe ése make model y chasis")
                        else:
                            car1=car()
                            car1.car_manufacturer=carManu
                            car1.car_model=carModel
                            car1.chasis=carChasis
                            car1.save()
                        auxCar = car.objects.get(car_manufacturer=carManu,car_model=carModel,chasis=carChasis)

                    if bandera:
                        spare1.save()
                        if auxCar:
                            spare1.car_info.add(auxCar)

                i=i+1

    dic={"refSpare":ref2,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/fillspare.html",dic)

def register(request):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")


    if request.method == "POST":
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            # print("Form is valid")
            username = form.cleaned_data["username"]
            userAux=User.objects.get(username=form.cleaned_data["username"])
            # print(username)
            # print(form)
            # print(type(form))
            profileAux = Profile()
            profileAux.ventas = False
            profileAux.bodega = False
            profileAux.mayorista = False
            profileAux.detal = True
            profileAux.user = userAux
            profileAux.save()
            messages.success(request,f"User {username} created")
            return redirect("home")
    else:
        form = UserRegisterForm()
    
    dic={"form":form,"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/register.html",dic)

def cart(request,val):

    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    spCartAll=spareCart.objects.all()
    # spCart=spareCart.objects.values("spareId","nameUser").distinct()
    spares = spare.objects.all().order_by("spare_name","spare_code")

    # print("User: "+str(val))
    spCartPost=spareCart.objects.filter(spareId=request.POST.get("cartN"))
    spCart = spareCart.objects.filter(nameUser=val).values("spareId","nameUser").distinct().order_by("spareId")
    # print("Cart")
    # print(spCartMain)
    cartIdPasar = request.POST.get("cartN")
    
    if request.method == "POST":
        # print("Entra al POST")
        # print(request.POST)
        if request.POST.get("cartN") == None:
            cartIdPasar = request.POST.get("cartPasar")
        else:
            cartIdPasar = request.POST.get("cartN")
        spCartPost = spareCart.objects.filter(spareId=request.POST.get("cartN"),nameUser=val)
        spCart = spareCart.objects.values("spareId","nameUser").filter(nameUser=val).distinct().order_by("spareId")

        # print(request.POST)
        # if request.POST.get("cartOpen"):
        #     print("Manda a abrir")
        if request.POST.get("cartDelete"):
            # print("Manda a borrar")
            cartDelete=spareCart.objects.filter(spareId=request.POST.get("cartN"))
            # print(cartDelete)
            cartDelete.delete()
        # spCart = spareCart.objects.values("spareId","nameUser").filter(spareId=request.POST.get("cartN"),nameUser=val).distinct()

        delist = request.POST.getlist("toDel")
        if delist:
            # print("Entra a delist para borrar individualmente")
            cartElementDelete=spareCart.objects.filter(spareId=request.POST.get("cartPasar"))
            for el in cartElementDelete:
                for de in delist:
                    if el.spareCode == de:
                        # print("Debe borrar")
                        # print(de)
                        el.delete()
            spCartPost = spareCart.objects.filter(spareId=request.POST.get("cartPasar"))
            

    dic={"cartIdPasar":cartIdPasar,"spare":spares,"spCartMain":spCartPost,"spCartAll":spCartAll,"spCart":spCart,"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/cart.html",dic)

def listAdmin(request):
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    spCartAll=spareCart.objects.all()
    allUser=User.objects.all()

    dic={"allUser":allUser,"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/listAdmin.html",dic)

def userProfile(request,val):
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    spCartAll=spareCart.objects.all()
    allUser=User.objects.all()
    userP=User.objects.filter(id=val)

    dic={"userP":userP,"allUser":allUser,"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    if request.method == "POST":

        print(request.POST)
        userP=User.objects.get(id=val)
        print(type(userP))
        print(userP)
        # userPa=User.objects.filter(id=val)
        # print(userP.profile.all)
        profileP=Profile.objects.filter(user=userP)
        print("Obtiene profileP")
        print(profileP)
        # print(userP)
        # print(userP.is_superuser)
        if request.POST.get("superuser"):
            print("Es superuser")
            userP.is_superuser=True
        else:
            userP.is_superuser=False

        if profileP:
            if request.POST.get("ventas"):
                print("Es ventas")
                userP.profile.ventas=True
            else:
                userP.profile.ventas=False
            if request.POST.get("bodega"):
                print("Es bodega")
                userP.profile.bodega=True
            else:
                userP.profile.bodega=False
            if request.POST.get("mayorista"):
                print("Es mayorista")
                userP.profile.mayorista=True
            else:
                userP.profile.mayorista=False
            if request.POST.get("detal"):
                print("Es detal")
                userP.profile.detal=True
            else:
                userP.profile.detal=False
        else:
            profileAux=Profile()
            profileAux.user=userP
            if request.POST.get("ventas"):
                print("Es ventas")
                profileAux.ventas=True
            else:
                profileAux.ventas=False
            if request.POST.get("bodega"):
                print("Es bodega")
                profileAux.bodega=True
            else:
                profileAux.bodega=False
            if request.POST.get("mayorista"):
                print("Es mayorista")
                profileAux.mayorista=True
            else:
                profileAux.mayorista=False
            if request.POST.get("detal"):
                print("Es detal")
                profileAux.detal=True
            else:
                profileAux.detal=False
            profileAux.save()
        userP.save()
        userP.profile.save()

        return render(request,"spareapp/listAdmin.html",dic)


    return render(request,"spareapp/userProfile.html",dic)

def deleteuser(request,val):
    dim=dimension.objects.values("atributeName").distinct()
    dim2=dimension.objects.all()
    atr=atribute.objects.values("atributeName").distinct()
    atr2=atribute.objects.all()
    allSparesall=spare.objects.all()
    allCategories=category.objects.all().order_by("category")
    allEngines=engine.objects.all()
    onlyManufCars=car.objects.all().values("car_manufacturer").order_by("car_manufacturer").distinct()
    manu=car.objects.values("car_manufacturer").order_by("car_manufacturer").distinct()
    allCars=car.objects.all()
    allSpares=spare.objects.values("spare_name","spare_category").order_by("spare_name").distinct()
    allVendors=vendor.objects.all()
    ref=reference.objects.all().order_by("referenceSpare")
    spCartAll=spareCart.objects.all()
    allUser=User.objects.all()

    userAux=User.objects.filter(id=val)
    print(userAux)
    userAux.delete()

    dic={"allUser":allUser,"manu":manu,"reference":ref,"allVendors":allVendors,"allAtributes":atr2,"atribute":atr,"allDimensions":dim2,"dimension":dim,"allSparesall":allSparesall,"allCategories":allCategories,"allCars":allCars,"onlyManufCars":onlyManufCars,"allEngines":allEngines,"allSpares":allSpares}

    return render(request,"spareapp/listAdmin.html",dic)

# ------------------------------ Contabilidad -----------------------------------

def contBase(request):

    return render(request,"spareapp/contBase.html")

@login_required
# @login_required(redirect_field_name='contLogin')
def contDay(request):

    # if not request.user.is_authenticated:
    #     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    # pedido=factura.objects.filter(refPersona__nombre="PODEROSO") | factura.objects.filter(refPersona__nombre="SUPERCAR") | factura.objects.filter(refPersona__nombre="AUTO GLOBAL PARTS #1") | factura.objects.filter(refPersona__nombre="AUTO GLOBAL PARTS #2")
    # # # pedido=factura.objects.filter(refPersona__nombre="Luis Velasco")
    # typeAuxx = factType.objects.get(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")

    # for ped in pedido:

    #     pedAux = factura.objects.get(id=ped.id)

    #     if pedAux.refCategory.nombre == "Factura cobrada (Mayorista)":

    #         pedAux.refType=typeAuxx

    #     pedAux.save()

    # print(pedido)

    # Lleno las categorias basicas -------------------------------
    catAux = factCategory.objects.filter(nombre="Factura cobrada")
    if catAux:
        pass
    else:
        catA = factCategory()
        catA.nombre = "Factura cobrada"
        catA.ingreso = True
        catA.save()
    catAux = factCategory.objects.filter(nombre="Mercancia credito pagada")
    if catAux:
        pass
    else:
        catA = factCategory()
        catA.nombre = "Mercancia credito pagada"
        catA.egreso = True
        catA.save()
    catAuxMay = factCategory.objects.filter(nombre="Factura cobrada (Mayorista)")
    if catAuxMay:
        pass
    else:
        catA = factCategory()
        catA.nombre = "Factura cobrada (Mayorista)"
        catA.ingreso = True
        catA.save()
    
    # Lleno los tipos basicos -------------------------------
    catAux = factType.objects.filter(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")
    if catAux:
        pass
    else:
        catA = factType()
        catA.nombre = "FACTURA CREDITO COBRADA (MAYORISTA)"
        catA.facCobrada = True
        catA.ingreso = True
        catA.save()

    tod = datetime.now().date()
    allTypes = factType.objects.all().order_by("nombre")
    editPrueba = False
    contTotal = 0
    allFactures = factura.objects.filter(fechaCreado__date=tod) | factura.objects.filter(fechaCobrado=tod)
    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)
    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    # Operacion --------------------------------------------

    tableAuxOpEmpty = None
    acum = 0
    totalParcialOp = {}
    cantAuxOp = tableOperacion.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    for nom in cantAuxOp:

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal

        if nom["principal"] == True:

            acumTablaTotales = acumTablaTotales + acum
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabTipo__nombre")

    tableAuxOpEmpty = tableOperacion.objects.all().values("tabNombre","tabTipo__nombre").distinct().order_by("tabTipo__nombre")

    # Operacion category --------------------------------------------

    tableAuxOpEmptyCat = None
    acum = 0
    totalParcialOpCat = {}
    cantAuxOpCat = tableOperacionCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    for nom in cantAuxOpCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal

        if nom["principal"] == True:

            acumTablaTotalesCat = acumTablaTotalesCat + acum
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    tableAuxOpEmptyCat = tableOperacionCat.objects.all().values("tabNombre","tabCat__nombre").distinct().order_by("tabCat__nombre")

    dic = {"acumTablaTotalesCat":acumTablaTotalesCat,"acumTablaTotales":acumTablaTotales,"cantAuxOpCat":cantAuxOpCat,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"tableAuxOpEmptyCat":tableAuxOpEmptyCat,"tableAuxOpEmpty":tableAuxOpEmpty,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"totalParcialOp":totalParcialOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"allFactures":allFactures,"contTotal":contTotal,"editPrueba":editPrueba,"tod":tod,"allTypes":allTypes,"facturesToCollect":facturesToCollect,"facturesToPay":facturesToPay}

    return render(request,"spareapp/contDay.html",dic)

def contEntry(request):

    allTypes = factType.objects.all().order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True).exclude(mercPagar=True)
    allCategories = factCategory.objects.filter(ingreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada")
    allCategoriesEntry = factCategory.objects.filter(ingreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada")
    allCategoriesSpending = factCategory.objects.filter(egreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada")
    allCustomers = persona.objects.all()
    tod = datetime.now().date()
    actualAux=datetime.now().date()
    actualDay=str(actualAux.year)+"-"+str('%02d' % actualAux.month)+"-"+str('%02d' % actualAux.day)
    deadlineDefault=(datetime.now()+timedelta(days=30)).date()
    actual=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0
    acum = 0
    banderaRepetido = False
    banderaNumero = ""
    facAuxAllCat = ""

    dic = {"actualDay":actualDay,"actual":actual,"allCategoriesSpending":allCategoriesSpending,"allCategoriesEntry":allCategoriesEntry,"allCustomers":allCustomers,"allTypes":allTypes,"allCategories":allCategories}

    if request.method == "POST":

        print("Entra a POST de contEntry")

        actualAux=datetime.now().date()
        actualDay=str(actualAux.year)+"-"+str('%02d' % actualAux.month)+"-"+str('%02d' % actualAux.day)

        factAux = factura()

        contFecha = request.POST.get("contFechaCreado")
        factAux.fechaCreado=contFecha

        tod=contFecha

        contNombre = request.POST.get("contNombre")
        nomAux = persona.objects.get(id=contNombre)
        factAux.refPersona = nomAux

        if request.POST.get("contNumFac"):

            auxF = factura.objects.filter(num=request.POST.get("contNumFac"))
            if auxF:
                banderaRepetido = True
                banderaNumero = request.POST.get("contNumFac")
                # return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

            contNumFac = request.POST.get("contNumFac")
            factAux.num = contNumFac

        contTypeIng = request.POST.get("contTypeIng")
        typeAux = factType.objects.get(id=contTypeIng)
        factAux.refType = typeAux

        contCatIng = request.POST.get("contCatIng")
        catAux = factCategory.objects.filter(id=contCatIng)

        for cat in catAux:
            if cat.limite == True:
                factAux.pendiente = True

        catAux = factCategory.objects.get(id=contCatIng)
        factAux.refCategory = catAux

        if request.POST.get("contFechaTope") != "":
            contFechaTope = request.POST.get("contFechaTope")
            factAux.fechaTope = contFechaTope

        contMonto = request.POST.get("contMonto")
        factAux.monto = contMonto

        if request.POST.get("contItbm") == "":
            contIva = float(0)
        else:
            contIva = request.POST.get("contItbm")
        factAux.iva = contIva

        contTotal = request.POST.get("contTotal")
        factAux.total = contTotal

        if request.POST.get("notaCredito"):
            factAux.nc = True
        else:
            factAux.nc = False

        factAux.note = request.POST.get("contNota")

        factAux.save()

        allTypes = factType.objects.all().order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True)

        # ----------- Operacion -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
        factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factType.objects.all()
        totalParcialOp = {}
        
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo")

        if factureAuxOp:

            print("Hay facturas")

            if tableAuxOp:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factType.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                    else:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum + fac.total

                                    else:

                                        custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum - fac.total

                                    else:

                                        custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacion()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factType.objects.get(nombre=ty)
                                costomInd.tabTipo = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                    else:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum + fac.total

                                    else:

                                        custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum - fac.total

                                    else:

                                        custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacion()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        for nom in cantAuxOp:

            suma = 0

            aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal
            
            totalParcialOp[nom["tabNombre"]] = acum

            acum = 0

        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAux2Op = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo__nombre")
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo__nombre")

        # ----------- Categoria -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabNombre","tabCat")

        if factureAuxCat:

            print("Hay facturas")

            if tableAuxCat:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                    for fac in facAuxAllCat:

                        if ty.ingreso == True:
                            if fac.nc == True:
                                custAcum = custAcum - fac.total
                            else:
                                custAcum = custAcum + fac.total
                        else:
                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                    for fac in facAuxAllCat:

                        if ty.ingreso == True:
                            if fac.nc == True:
                                custAcum = custAcum - fac.total
                            else:
                                custAcum = custAcum + fac.total
                        else:
                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        for nom in cantAuxCat:

            aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal
            
            totalParcialOpCat[nom["tabNombre"]] = acum

            acum = 0

        cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabNombre","tabCat__nombre")
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabNombre","tabCat__nombre")

        if request.POST.get("entryOption")=="otro":

            tod = datetime.now().date()
            allTypes = factType.objects.all().order_by("nombre")
            contTotal = 0
            allFactures = factura.objects.filter(fechaCreado__date=tod) | factura.objects.filter(fechaCobrado=tod)
            allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
            allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
            facturesToCollect = len(allFacturesToCollect)
            facturesToPay = len(allFacturesToPay)

            contTotal = 0
            noIncludeTotal = 0
            noIncludeTotalGasto = 0
            contPagadoCobrado = 0

            dic = {"banderaNumero":banderaNumero,"banderaRepetido":banderaRepetido,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxCat":tableAuxCat,"totalParcialOpCat":totalParcialOpCat,"totalParcialOp":totalParcialOp,"tableAuxOp":tableAux2Op,"cantAuxOp":cantAuxOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"allFactures":allFactures,"contTotal":contTotal,"tod":tod,"allTypes":allTypes,"facturesToCollect":facturesToCollect,"facturesToPay":facturesToPay}

            return render(request,"spareapp/contDay.html",dic)

        dic = {"banderaNumero":banderaNumero,"banderaRepetido":banderaRepetido,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotal":noIncludeTotal,"noIncludeTotalGasto":noIncludeTotalGasto,"actualDay":actualDay,"actual":actual,"allCustomers":allCustomers,"tod":tod,"allTypes":allTypes,"allCategories":allCategories}
    
    return render(request,"spareapp/contEntry.html",dic)

def contType(request,val,val2):

    print("Entra en contType")
    print("Val")
    print(val)

    allTypes = factType.objects.all().order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)
    todos = factType.objects.all()
    tod = datetime.now().date()
    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0
    allFacturesAux = None

    allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val).order_by("fechaCreado","id")
    allFacturesAux = allFacturesVal
    typeAux = factType.objects.get(nombre=val)

    if val2 != "today":

        tod = val2

        for t in todos:
            s=t.nombre
            if s:
                out = s.translate(str.maketrans('', '', '/'))
                if val.upper()==out.upper():
                    allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=s).order_by("fechaCreado","id")

    if typeAux.facCobrada == True:

        if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

            allFacturesVal = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
        else:
            allFacturesVal = factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

    if typeAux.mercPagada == True:

        allFacturesVal = factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

    if typeAux.facCobrar == True:

        allFacturesVal = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    if typeAux.mercPagar == True:

        allFacturesVal = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
        allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    itbm7 = {}

    if typeAux.facCobrada == False and typeAux.mercPagada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

        if typeAux.ingreso:

            for fac in allFacturesVal:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

        else:

            for fac in allFacturesVal:

                if fac.nc == True:

                    montoTotal = montoTotal + fac.monto
                    itbmTotal = itbmTotal + float(fac.iva)
                    totalTotal = totalTotal + fac.total

                else:

                    montoTotal = montoTotal - fac.monto
                    itbmTotal = itbmTotal - float(fac.iva)
                    totalTotal = totalTotal - fac.total

    if typeAux.facCobrada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total

    if typeAux.mercPagada == True:

        print("Entra en facPagada")

        for fac in allFacturesVal:

            if fac.nc == True:

                print("Entra en nc")

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

            else:

                print("Entra en false nc")

                montoTotal = montoTotal - fac.monto
                itbmTotal = itbmTotal - float(fac.iva)
                totalTotal = totalTotal - fac.total

    if typeAux.facCobrar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux.mercPagar == True:

        for fac in allFacturesVal:

            if fac.nc == True:

                montoTotal = montoTotal - fac.monto
                itbmTotal = itbmTotal - float(fac.iva)
                totalTotal = totalTotal - fac.total

            else:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

    typeDate = val2

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    allFacturesModal = allFacturesAux

    dic = {"allFacturesModal":allFacturesModal,"typeAux":typeAux,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"montoTotal":montoTotal,"itbmTotal":itbmTotal,"totalTotal":totalTotal,"itbm7":itbm7,"typeDate":typeDate,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contType.html",dic)

def contTypeCat(request,val,val2):

    allTypes = factCategory.objects.all().order_by("nombre")
    todos = factCategory.objects.all()
    tod = val2
    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0
    val = str(factCategory.objects.get(id=val))

    allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refCategory__nombre=val).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True).order_by("fechaCreado","id")

    # if val2 != "today":

    #     tod = val2

    #     for t in todos:
    #         s=t.nombre
    #         if s:
    #             out = s.translate(str.maketrans('', '', '/'))
    #             if val.upper()==out.upper():
    #                 allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refCategory__nombre=s).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True).order_by("fechaCreado","id")

    itbm7 = {}

    typeAux = factType.objects.filter(id=allFacturesVal[0].refType.id)

    if typeAux[0].facCobrada == False and typeAux[0].mercPagada == False and typeAux[0].facCobrar == False and typeAux[0].mercPagar == False:

        if typeAux[0].ingreso:

            for fac in allFacturesVal:

                if fac.nc == True:

                    montoTotal = montoTotal - fac.monto
                    itbmTotal = itbmTotal - float(fac.iva)
                    totalTotal = totalTotal - fac.total

                else:

                    montoTotal = montoTotal + fac.monto
                    itbmTotal = itbmTotal + float(fac.iva)
                    totalTotal = totalTotal + fac.total

        else:

            for fac in allFacturesVal:

                if fac.nc == True:

                    montoTotal = montoTotal + fac.monto
                    itbmTotal = itbmTotal + float(fac.iva)
                    totalTotal = totalTotal + fac.total

                else:

                    montoTotal = montoTotal - fac.monto
                    itbmTotal = itbmTotal - float(fac.iva)
                    totalTotal = totalTotal - fac.total

    if typeAux[0].facCobrada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total

    if typeAux[0].mercPagada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux[0].facCobrar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux[0].mercPagar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total
    
    # for fac in allFacturesVal:

    #     montoTotal = montoTotal + fac.monto
    #     itbmTotal = itbmTotal + float(fac.iva)
    #     totalTotal = totalTotal + fac.total

    typeDate = val2

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"montoTotal":montoTotal,"itbmTotal":itbmTotal,"totalTotal":totalTotal,"itbm7":itbm7,"typeDate":typeDate,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contTypeCat.html",dic)

def contTypeTarjeta(request,val,val2):

    tod = datetime.now().date()
    typeAux = factType.objects.get(nombre=val)

    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0
    interesTotal = 0
    retencionTotal = 0
    netoTotal = 0

    allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val)

    if val2 != "today":

        tod = val2

        allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val)

    itbm7 = {}

    for fac in allFacturesVal:

        itbmMonto = float(fac.iva)

        if typeAux.visa == True:

            neto = float((fac.total)-(float(fac.total)*0.0225*1.07)-(float(itbmMonto)/2))
            interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            retencionTotal = retencionTotal + float(itbmMonto/2)
            netoTotal = netoTotal + neto
            itbm7[fac.id] = [float(itbmMonto),float(float(fac.total)*0.0225*1.07),float(float(itbmMonto)/2),float(neto)]

        else:

            neto = float((fac.total)-(float(fac.total)*0.02*1.07)-(float(itbmMonto)/2))
            interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)
            retencionTotal = retencionTotal + float(itbmMonto/2)
            netoTotal = netoTotal + neto
            itbm7[fac.id] = [float(fac.monto)*0.07,float(float(fac.total)*0.02*1.07),float(float(itbmMonto)/2),float(neto)]

    for fac in allFacturesVal:

        montoTotal = montoTotal + fac.monto
        itbmTotal = itbmTotal + float(fac.iva)
        totalTotal = totalTotal + fac.total

    typeDate = val2

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"netoTotal":netoTotal,"retencionTotal":retencionTotal,"interesTotal":interesTotal,"montoTotal":montoTotal,"itbmTotal":itbmTotal,"totalTotal":totalTotal,"itbm7":itbm7,"typeDate":typeDate,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contTypeTarjeta.html",dic)

def contTypeRange(request,val,val2,val3):

    allTypes = factType.objects.all().order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)
    tod = datetime.now().date()
    typeAux = factType.objects.get(id=val)

    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0

    todos = factType.objects.all()

    allFacturesVal = factura.objects.filter(refType=val,fechaCreado__date__gte=val2,fechaCreado__date__lte=val3).order_by("fechaCreado","id")

    allFacturesModal = allFacturesVal

    if typeAux.facCobrada == True:

        if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

            allFacturesVal = factura.objects.filter(fechaCreado__lte=val3,fechaCreado__gte=val2,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
        else:
            allFacturesVal = factura.objects.filter(refCategory__ingreso=True,fechaCobrado__lte=val3,fechaCobrado__gte=val2,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

    if typeAux.mercPagada == True:

        allFacturesVal = factura.objects.filter(refCategory__egreso=True,fechaCobrado__lte=val3,fechaCobrado__gte=val2,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

    if typeAux.facCobrar == True:

        allFacturesVal = factura.objects.filter(fechaCreado__date__lte=val3,fechaCreado__date__gte=val2,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    if typeAux.mercPagar == True:

        allFacturesVal = factura.objects.filter(fechaCreado__date__lte=val3,fechaCreado__date__gte=val2,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
        allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    itbm7 = {}

    if typeAux.facCobrada == False and typeAux.mercPagada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

        if typeAux.ingreso:

            for fac in allFacturesVal:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

        else:
    
            for fac in allFacturesVal:

                if fac.nc == True:

                    montoTotal = montoTotal + fac.monto
                    itbmTotal = itbmTotal + float(fac.iva)
                    totalTotal = totalTotal + fac.total

                else:

                    montoTotal = montoTotal - fac.monto
                    itbmTotal = itbmTotal - float(fac.iva)
                    totalTotal = totalTotal - fac.total

                # montoTotal = montoTotal - fac.monto
                # itbmTotal = itbmTotal - float(fac.iva)
                # totalTotal = totalTotal - fac.total

    if typeAux.facCobrada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total

    if typeAux.mercPagada == True:

        for fac in allFacturesVal:

            if fac.nc == True:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

            else:

                montoTotal = montoTotal - fac.monto
                itbmTotal = itbmTotal - float(fac.iva)
                totalTotal = totalTotal - fac.total

            # montoTotal = montoTotal - fac.monto
            # itbmTotal = itbmTotal - float(fac.iva)
            # totalTotal = totalTotal - fac.total

    if typeAux.facCobrar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux.mercPagar == True:

        for fac in allFacturesVal:

            if fac.nc == True:

                montoTotal = montoTotal - fac.monto
                itbmTotal = itbmTotal - float(fac.iva)
                totalTotal = totalTotal - fac.total

            else:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

    typeDate = "Desde "+val2+", hasta "+val3

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    val = typeAux
    dateFrom = val2
    dateTo = val3

    dic = {"allFacturesModal":allFacturesModal,"typeAux":typeAux,"dateTo":dateTo,"dateFrom":dateFrom,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"allTypes":allTypes,"totalTotal":totalTotal,"itbmTotal":itbmTotal,"montoTotal":montoTotal,"itbm7":itbm7,"typeDate":typeDate,"val3":val3,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contType.html",dic)

def contTypeRangeCat(request,val,val2,val3):

    allTypes = factCategory.objects.all().order_by("nombre")
    tod = datetime.now().date()
    allFacturesVal = factura.objects.filter(refCategory__id=val,fechaCreado__date__gte=val2,fechaCreado__date__lte=val3).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True).order_by("fechaCreado","id")

    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0

    itbm7 = {}

    for fac in allFacturesVal:

        if fac.refCategory.ingreso == True:

            if fac.monto == fac.total:

                itbm7[fac.id] = float(0)

            else:

                itbm7[fac.id] = float(fac.monto)*0.07

    typeAux = factType.objects.filter(id=allFacturesVal[0].refType.id)

    if typeAux[0].facCobrada == False and typeAux[0].mercPagada == False and typeAux[0].facCobrar == False and typeAux[0].mercPagar == False:

        if typeAux[0].ingreso:

            for fac in allFacturesVal:

                montoTotal = montoTotal + fac.monto
                itbmTotal = itbmTotal + float(fac.iva)
                totalTotal = totalTotal + fac.total

        else:

            for fac in allFacturesVal:

                montoTotal = montoTotal - fac.monto
                itbmTotal = itbmTotal - float(fac.iva)
                totalTotal = totalTotal - fac.total

    if typeAux[0].facCobrada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total

    if typeAux[0].mercPagada == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux[0].facCobrar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal - fac.monto
            itbmTotal = itbmTotal - float(fac.iva)
            totalTotal = totalTotal - fac.total

    if typeAux[0].mercPagar == True:

        for fac in allFacturesVal:

            montoTotal = montoTotal + fac.monto
            itbmTotal = itbmTotal + float(fac.iva)
            totalTotal = totalTotal + fac.total
    
    # for fac in allFacturesVal:

    #     montoTotal = montoTotal + fac.monto
    #     itbmTotal = itbmTotal + float(fac.iva)
    #     totalTotal = totalTotal + fac.total

    typeDate = "From "+val2+", to "+val3

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    val = factCategory.objects.get(id=val)
    dateFrom = val2
    dateTo = val3

    dic = {"dateTo":dateTo,"dateFrom":dateFrom,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"allTypes":allTypes,"totalTotal":totalTotal,"itbmTotal":itbmTotal,"montoTotal":montoTotal,"itbm7":itbm7,"typeDate":typeDate,"val3":val3,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contTypeCat.html",dic)

def contTypeRangeTarjeta(request,val,val2,val3):

    tod = datetime.now().date()
    typeAux = factType.objects.get(nombre=val)

    montoTotal = 0
    itbmTotal = 0
    totalTotal = 0
    interesTotal = 0
    retencionTotal = 0
    netoTotal = 0

    allFacturesVal = factura.objects.filter(refType__nombre=val,fechaCreado__date__gte=val2,fechaCreado__date__lte=val3)

    itbm7 = {}

    netoTotal = 0

    for fac in allFacturesVal:

        if fac.monto == fac.total:

            itbmMonto = float(0)

        else:

            itbmMonto = float(fac.monto)*0.07

        if typeAux.visa == True:

            neto = float((fac.total)-(float(fac.total)*0.0225*1.07)-(float(itbmMonto)/2))
            interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            retencionTotal = retencionTotal + float(itbmMonto/2)
            netoTotal = netoTotal + neto
            itbm7[fac.id] = [float(itbmMonto),float(float(fac.total)*0.0225*1.07),float(float(itbmMonto)/2),float(neto)]

        else:

            neto = float((fac.total)-(float(fac.total)*0.02*1.07)-(float(itbmMonto)/2))
            interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)
            retencionTotal = retencionTotal + float(itbmMonto/2)
            netoTotal = netoTotal + neto
            itbm7[fac.id] = [float(fac.monto)*0.07,float(float(fac.total)*0.02*1.07),float(float(itbmMonto)/2),float(neto)]

    for fac in allFacturesVal:

        montoTotal = montoTotal + fac.monto
        if fac.monto == fac.total:
            itbmTotal = itbmTotal + float(0)
        else:
            itbmTotal = itbmTotal + float(fac.monto)*0.07
        totalTotal = totalTotal + fac.total

    typeDate = "From "+val2+", to "+val3

    dateFrom = val2
    dateTo = val3

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"dateTo":dateTo,"dateFrom":dateFrom,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"netoTotal":netoTotal,"retencionTotal":retencionTotal,"interesTotal":interesTotal,"totalTotal":totalTotal,"itbmTotal":itbmTotal,"montoTotal":montoTotal,"itbm7":itbm7,"typeDate":typeDate,"val3":val3,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    return render(request,"spareapp/contTypeTarjeta.html",dic)

def contToCollect(request):

    tod = datetime.now().date()
    dateFrom = None
    dateTo = None
    dayFrom = None
    dayTo = None
    searchMetodo = "all"
    creadoAuxdateFrom = None
    creadoAuxdateTo = None

    allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
    allFacturesAux = allFacturesPay

    if allFacturesPay:
        dateFrom = allFacturesPay[0].fechaCreado.date()
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    # if request.POST.get("searchDateFrom"):
    #     dateFrom = request.POST.get("searchDateFrom")
    #     dayFrom = dateFrom
    # if request.POST.get("searchDateTo"):
    #     dateTo = request.POST.get("searchDateTo")
    #     dayTo = dateTo

    if request.POST.get("searchDateFrom"):
        dateFrom = request.POST.get("searchDateFrom")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateFrom = creadoAuxdateFrom.date()
        dayFrom = dateFrom
        dateFrom = str(creadoAuxdateFrom.date())
    if request.POST.get("searchDateTo"):
        dateTo = request.POST.get("searchDateTo")
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dateTo = creadoAuxdateTo.date()
        dayTo = dateTo
        dateTo = str(creadoAuxdateTo.date())

    allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)
    
    if request.POST.get("search") == "range":
        searchMetodo = "range"
        allFacturesPay = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

# ------------------------------------------------------------

    filter = None

    filterFactures = factura.objects.none()
    filterFacturesE = factura.objects.none()
    filterAuxFinal = ""
    palabrasErase = []
    palabrasErase2 = []

    if request.POST.get("cod2"):
        filter = request.POST.get("cod2")
        filterAuxFinal = request.POST.get("cod2")
        auxInicio = -1
        auxFin = -1
        acumCom = 0
        filterAux = filter
        filterAuxErase = filter
        palabras = []
        nuevoFilter = ""
        palabraFinal = ""

        for pos,let in enumerate(filter):

            if(let == '-'):

                filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
                # auxFine = pos
                palabraFinalErase = filterErase.split(" ")
                palabrasErase.append(palabraFinalErase[0])
                auxInicioe = -1

            else:

                auxInicioe = pos

        for val in palabrasErase:

            palabrasErase2.append("-"+val)

        for pos,let in enumerate(filter):

            if(let == '"'):

                acumCom = acumCom + 1

                if acumCom == 2:

                    nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                    auxFin = pos
                    palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                    palabras.append(palabraFinal)
                    acumCom = 0
                    auxInicio = -1
                    auxFin = -1

                else:

                    auxInicio = pos

        palabrasAux = []

        for val in palabras:

            if filterAux.find(val) or filterAux == val:

                filterAux2 = filterAux.replace(val,"")
                filterAux = filterAux2
            palabrasAux.append(val.strip('"'))

        filterAux = filterAux.split(" ")
        filterAux = [item for item in filterAux if item]
        filter = filterAux + palabrasAux
        for val in palabrasErase2:
            filter.remove(val)

    if filter:

        for fil in filter:

            if filterFactures:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

            else:

                filterFactures = factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )

        else:

            filterFacturesE = ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )

    if filter:

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
    
    if searchMetodo == "range":
        allFacturesPay = filterFactures & allFacturesPay & filterFacturesE
    else:
        allFacturesPay = filterFactures & filterFacturesE

    # ------------------------------------------------------------

    deadlineDic = {}

    for all in allFacturesPay:

        deadline = datetime.now().date() - all.fechaCreado.date()
        deadline = deadline.days
        deadlineDic[all.id] = deadline

    acum = 0
    acum2 = 0
    iva = 0

    for fac in allFacturesPay:

        acum = acum + fac.monto
        acum2 = acum2 + fac.total
        iva = iva + fac.iva

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    allFacturesModal = allFacturesPay

    if searchMetodo == "range":

        tod = None

    dic = {"allFacturesModal":allFacturesModal,"filtro":filterAuxFinal,"iva":iva,"searchMetodo":searchMetodo,"dayFrom":dayFrom,"dayTo":dayTo,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"deadlineDic":deadlineDic,"allFacturesPay":allFacturesPay,"totalTotal":acum2,"montoTotal":acum}

    return render(request,"spareapp/contToCollect.html",dic)

def contToPay(request):

    filter = None
    tod = datetime.now().date()
    dateFrom = None
    dateTo = None
    dayFrom = None
    dayTo = None
    searchMetodo = "all"
    creadoAuxdateFrom = None
    creadoAuxdateTo = None

    allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    if allFacturesPay:
        dateFrom = allFacturesPay[0].fechaCreado.date()
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    if request.POST.get("searchDateFrom"):
        dateFrom = request.POST.get("searchDateFrom")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateFrom = creadoAuxdateFrom.date()
        dayFrom = dateFrom
        dateFrom = str(creadoAuxdateFrom.date())
    if request.POST.get("searchDateTo"):
        dateTo = request.POST.get("searchDateTo")
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dateTo = creadoAuxdateTo.date()
        dayTo = dateTo
        dateTo = str(creadoAuxdateTo.date())
    # allFacturesPay = factura.objects.filter(fechaCreado__gte=dateFrom,fechaCreado__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
    allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    if request.POST.get("search") == "range":
        searchMetodo = "range"
        allFacturesPay = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    # ------------------------------------------------------------

    checkeado = False

    if request.POST.get("entrySpending2"):

        checkeado = True
    
    else:

        checkeado = False

    if checkeado == True:

        allFactures = factura.objects.filter(nc=checkeado)

    else:

        allFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")

    filterFactures = factura.objects.none()
    filterFacturesE = factura.objects.none()
    palabrasErase = []
    palabrasErase2 = []
    filterAuxFinal = ""

    if request.POST.get("cod2"):
        filter = request.POST.get("cod2")
        filterAuxFinal = request.POST.get("cod2")
        auxInicio = -1
        auxFin = -1
        acumCom = 0
        filterAux = filter
        filterAuxErase = filter
        palabras = []
        nuevoFilter = ""
        palabraFinal = ""

        for pos,let in enumerate(filter):

            if(let == '-'):

                filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
                # auxFine = pos
                palabraFinalErase = filterErase.split(" ")
                palabrasErase.append(palabraFinalErase[0])
                auxInicioe = -1

            else:

                auxInicioe = pos

        for val in palabrasErase:

            palabrasErase2.append("-"+val)

        for pos,let in enumerate(filter):

            if(let == '"'):

                acumCom = acumCom + 1

                if acumCom == 2:

                    nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                    auxFin = pos
                    palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                    palabras.append(palabraFinal)
                    acumCom = 0
                    auxInicio = -1
                    auxFin = -1

                else:

                    auxInicio = pos

        palabrasAux = []

        for val in palabras:

            if filterAux.find(val) or filterAux == val:

                filterAux2 = filterAux.replace(val,"")
                filterAux = filterAux2
            palabrasAux.append(val.strip('"'))

        filterAux = filterAux.split(" ")
        filterAux = [item for item in filterAux if item]
        filter = filterAux + palabrasAux
        for val in palabrasErase2:
            filter.remove(val)

    if filter:

        for fil in filter:

            if filterFactures:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )

            else:

                filterFactures = factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    if filter:

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )

        else:

            filterFacturesE = ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
    
    if searchMetodo == "range":
        allFacturesPay = filterFactures & allFacturesPay & allFactures & filterFacturesE
    else:
        allFacturesPay = filterFactures & allFactures & filterFacturesE

    # ------------------------------------------------------------

    deadlineDic = {}

    for all in allFacturesPay:

        deadline = datetime.now().date() - all.fechaCreado.date()
        deadline = deadline.days
        deadlineDic[all.id] = deadline

    acum = 0
    acum2 = 0
    iva = 0

    for fac in allFacturesPay:
        
        if fac.nc == False:

            acum = acum + fac.monto
            acum2 = acum2 + fac.total
            iva = iva + fac.iva

        else:

            acum = acum - fac.monto
            acum2 = acum2 - fac.total
            iva = iva - fac.iva        

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True).order_by("fechaCreado","id")
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    allFacturesModal = allFacturesToPay

    if searchMetodo == "range":

        tod = None

    dic = {"allFacturesModal":allFacturesModal,"checkeado":checkeado,"filtro":filterAuxFinal,"searchMetodo":searchMetodo,"dayFrom":dayFrom,"dayTo":dayTo,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"deadlineDic":deadlineDic,"allFacturesPay":allFacturesPay,"totalTotal":acum2,"montoTotal":acum,"iva":iva}

    return render(request,"spareapp/contToPay.html",dic)

def contAdmin(request):

    return render(request,"spareapp/contAdmin.html")

def contAddType(request):

    tod = datetime.now().date()

    if request.method == "POST":

        typeNombre = request.POST.get("typeNombre")

        if typeNombre != "":

            listType = factType.objects.filter(nombre=typeNombre)

            if (listType):

                print("Ya existe éste type")

            else:

                factTypeAux = factType()
                factTypeAux.nombre = typeNombre
                if request.POST.get("typeEntry") == "entry":
                    factTypeAux.ingreso = True
                    factTypeAux.gasto = False
                else:
                    factTypeAux.ingreso = False
                    factTypeAux.gasto = True
                factTypeAux.save()

    return render(request,"spareapp/contAddType.html")

def contAddCategory(request):

    if request.method == "POST":

        catNombre = request.POST.get("catNombre")
        catIngreso = request.POST.get("catIngreso")
        catEgreso = request.POST.get("catEgreso")
        catExpirationDate = request.POST.get("catExpirationDate")
        listCategory = factCategory.objects.filter(nombre=catNombre)

        if (listCategory):

                print("Ya existe ésta categoría")
                
        else:

            factCatAux = factCategory()
            factCatAux.nombre = catNombre
            if catIngreso:
                factCatAux.ingreso = True
            if catEgreso:
                factCatAux.egreso = True
            if catExpirationDate:
                factCatAux.limite = True
            factCatAux.save()

    return render(request,"spareapp/contAddCategory.html")

def contListType(request):

    allTypes = factType.objects.all().order_by("nombre")
    facAux = ""
    deleteAux = {}

    for ty in allTypes:

        facAux = factura.objects.filter(refType=ty)
        if facAux:
            deleteAux[ty.id] = "on"
        else:
            deleteAux[ty.id] = "off"

    if request.method == "POST":

        for cat in request.POST.getlist("typId"):

            # Si ahora es un tipo de factura por cobrar
            bandAntesCobrar = False
            # Si antes era un tipo de factura por cobrar
            bandAntesCobrar2 = False
            # Si antes era un tipo de factura cobrada
            bandAntesCobrada2 = False
            # Si ahora es un tipo de factura por pagar
            bandAntesPagar = False
            # Si antes era un tipo de factura por pagar
            bandAntesPagar2 = False

            singleType = factType.objects.get(id=cat)

            if singleType.facCobrar == True:
                bandAntesCobrar2 = True

            if singleType.mercPagar == True:
                bandAntesPagar2 = True

            if request.POST.get("typNom"+cat):

                singleType.nombre = request.POST.get("typNom"+cat)

            if request.POST.get("facCobrar"+cat):

                bandAntesCobrar = False
                if singleType.facCobrar == False:
                    bandAntesCobrar = True

                singleType.facCobrar = True
                if bandAntesCobrar == True and singleType.facCobrar == True:
                    print("Antes era debito y ahora es credito por cobrar")
                    facturaCambiar = factura.objects.filter(refType=singleType)
                    for fac in facturaCambiar:
                        auxCat = factCategory.objects.filter(limite=True,ingreso=True)
                        facCamb = factura.objects.get(id=fac.id)
                        facCamb.refCategory=auxCat[0]
                        facCamb.pendiente = True
                        creado = facCamb.fechaCreado
                        creadoAux = datetime.strptime(str(creado.date()),"%Y-%m-%d")
                        deadlineDefault=(creadoAux+timedelta(days=30)).date()
                        actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)
                        actual = actualAux
                        singleType.ingreso = True
                        singleType.gasto = False
                        facCamb.refType.ingreso = True
                        facCamb.refType.gasto = False
                        facCamb.fechaTope = actual
                        facCamb.save()
            
            else:

                singleType.facCobrar = False

            if request.POST.get("facCobrada"+cat):

                singleType.facCobrada = True
            
            else:

                singleType.facCobrada = False

            if request.POST.get("mercPagar"+cat):

                bandAntesPagar = False
                if singleType.mercPagar == False:
                    bandAntesPagar = True

                singleType.mercPagar = True
                if bandAntesPagar == True and singleType.mercPagar == True:
                    print("Antes era debito y ahora es credito por pagar")
                    facturaCambiar = factura.objects.filter(refType=singleType)
                    for fac in facturaCambiar:
                        auxCat = factCategory.objects.filter(limite=True,egreso=True)
                        facCamb = factura.objects.get(id=fac.id)
                        facCamb.refCategory=auxCat[0]
                        facCamb.pendiente = True
                        creado = facCamb.fechaCreado
                        creadoAux = datetime.strptime(str(creado.date()),"%Y-%m-%d")
                        deadlineDefault=(creadoAux+timedelta(days=30)).date()
                        actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)
                        actual = actualAux
                        singleType.ingreso = False
                        singleType.gasto = True
                        facCamb.refType.ingreso = False
                        facCamb.refType.gasto = True
                        facCamb.fechaTope = actual
                        facCamb.save()
            
            else:

                singleType.mercPagar = False

            if request.POST.get("mercPagada"+cat):

                singleType.mercPagada = True
            
            else:

                singleType.mercPagada = False

            if request.POST.get("visa"+cat):

                singleType.visa = True
            
            else:

                singleType.visa = False

            if request.POST.get("clave"+cat):

                singleType.clave = True
            
            else:

                singleType.clave = False

            if request.POST.get("typInclude"+cat):

                singleType.include = True
            
            else:

                singleType.include = False

            if request.POST.get("ingreso"+cat):

                if bandAntesCobrar == False and bandAntesPagar == False:
                    singleType.ingreso = True
                    singleType.gasto = False
            
            else:

                if bandAntesCobrar == False and bandAntesPagar == False:
                    singleType.ingreso = False
                    singleType.gasto = True

            if bandAntesCobrar2 == True and singleType.facCobrar == False:
                print(singleType.nombre)
                print("Antes era credito por cobrar y ahora es contado")
                facturaCambiar = factura.objects.filter(refType=singleType)
                for fac in facturaCambiar:
                    auxCat = factCategory.objects.filter(limite=False,ingreso=True)
                    facCamb = factura.objects.get(id=fac.id)
                    facCamb.refCategory=auxCat[0]
                    facCamb.pendiente = False
                    facCamb.fechaTope = None
                    facCamb.fechaCobrado = None
                    facCamb.save()
                    
                    facCambCob = factura.objects.filter(num=facCamb.num,refPersona=facCamb.refPersona,pendiente=False).exclude(fechaCobrado=None)
                    for e in facCambCob:
                        e.delete()

            if bandAntesPagar2 == True and singleType.mercPagar == False:
                print(singleType.nombre)
                print("Antes era credito por pagar y ahora es contado")
                facturaCambiar = factura.objects.filter(refType=singleType)
                for fac in facturaCambiar:
                    auxCat = factCategory.objects.filter(limite=False,egreso=True)
                    facCamb = factura.objects.get(id=fac.id)
                    facCamb.refCategory=auxCat[0]
                    facCamb.pendiente = False
                    facCamb.fechaTope = None
                    facCamb.fechaCobrado = None
                    facCamb.save()
                    
                    facCambCob = factura.objects.filter(num=facCamb.num,refPersona=facCamb.refPersona,pendiente=False).exclude(fechaCobrado=None)
                    for e in facCambCob:
                        e.delete()

            singleType.save()

    allTypes = factType.objects.all().order_by("nombre")

    # ----------- Operacion -------------------
    toddy = datetime.now().date()
    allTypesCustom = factType.objects.all()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

    if tableAuxOp:

        print("Hay tabla")

        allTypesCustom = factType.objects.all()
        custAcum = 0
        for ty in allTypesCustom:
            # print("Ty")
            # print(ty)
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                # print("Entra en facCobrar")
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                # print("Entra en mercPagar")
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                # print("Entra en mercPagada")
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)
                # facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

            if ty.facCobrada == True:
                # print("Entra en facCobrada")

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            # print("facAuxAll")
            # print(facAuxAll)

            if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                for fac in facAuxAll:

                    if fac.nc == True:

                        custAcum = custAcum + fac.total

                    else:

                        custAcum = custAcum - fac.total

            else:

                for fac in facAuxAll:

                    # custAcum = custAcum + fac.total


                    if fac.nc == True:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                        else:

                            custAcum = custAcum + fac.total

                    else:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum + fac.total

            custAcum = abs(custAcum)
                        
            customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                    if prob2:

                        costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        costomInd.tabTotal = custAcum
                        costomInd.save()

                    else:

                        costomInd = tableOperacion()
                        costomInd.fecha = toddy
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
            custAcum = 0

    else:

        print("No hay tabla")
        custAcum = 0
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                for fac in facAuxAll:

                    if fac.nc == True:

                        custAcum = custAcum + fac.total

                    else:

                        custAcum = custAcum - fac.total

            else:

                for fac in facAuxAll:

                    if fac.nc == True:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                        else:

                            custAcum = custAcum + fac.total

                    else:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum + fac.total

            custAcum = abs(custAcum)
                        
            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:
                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:
                    costomInd = tableOperacion()
                    costomInd.fecha = toddy
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factType.objects.get(nombre=ty)
                    costomInd.tabTipo = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
            
            custAcum = 0

    dic = {"deleteAux":deleteAux,"allTypes":allTypes}

    return render(request,"spareapp/contListType.html",dic)

def contListCategory(request):

    allCategories = factCategory.objects.all().order_by("nombre")
    deleteAux = {}
    facAux = ""

    for cat in allCategories:

        facAux = factura.objects.filter(refCategory=cat)
        if facAux:
            deleteAux[cat.id] = "on"
        else:
            deleteAux[cat.id] = "off"

    if request.method == "POST":

        allCategories = factCategory.objects.all().order_by("nombre")

        for cat in request.POST.getlist("catId"):

            singleCategory = factCategory.objects.get(id=cat)

            if request.POST.get("contNom"+cat):

                singleCategory.nombre = request.POST.get("contNom"+cat)

            if request.POST.get("catEntry"+cat):

                singleCategory.ingreso = True
                singleCategory.egreso = False
            
            else:

                singleCategory.ingreso = False
                singleCategory.egreso = True

            if request.POST.get("catExpiration"+cat):
                singleCategory.limite = True
            else:
                singleCategory.limite = False
            
            singleCategory.save()

    dic = {"deleteAux":deleteAux,"allCategories":allCategories}

    return render(request,"spareapp/contListCategory.html",dic)

def contDeleteType(request,val):

    singleType = factType.objects.get(id=val)

    singleType.delete()

    allTypes = factType.objects.all()

    dic = {"allTypes":allTypes}

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def contDeleteCategory(request,val):

    singleCategory = factCategory.objects.get(id=val)

    singleCategory.delete()

    allCategories = factCategory.objects.all().order_by("nombre")

    dic = {"allCategories":allCategories}

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def contByDay(request):

    toddy = request.POST.get("searchDate")
    tod = request.POST.get("searchDate")
    acumTablaTotales = 0

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    facAuxAllCat = ""

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    # ----------- Operacion -------------------
    print("ENtra en operacion")

    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=toddy).values("tabNombre","principal").order_by("tabNombre").distinct()
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

    if tableAuxOp:

        print("Hay tabla")

        allTypesCustom = factType.objects.all()
        custAcum = 0
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                for fac in facAuxAll:

                    if fac.nc == True:

                        custAcum = custAcum + fac.total

                    else:

                        custAcum = custAcum - fac.total

            else:

                for fac in facAuxAll:

                    if fac.nc == True:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                        else:

                            custAcum = custAcum + fac.total

                    else:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum + fac.total

            custAcum = abs(custAcum)
            customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                    if prob2:

                        costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        costomInd.tabTotal = custAcum
                        costomInd.save()

                    else:

                        costomInd = tableOperacion()
                        costomInd.fecha = toddy
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
            custAcum = 0

    else:

        print("No hay tabla")

        custAcum = 0
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                for fac in facAuxAll:

                    if fac.nc == True:

                        custAcum = custAcum + fac.total

                    else:

                        custAcum = custAcum - fac.total

            else:

                for fac in facAuxAll:

                    if fac.nc == True:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                        else:

                            custAcum = custAcum + fac.total

                    else:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum + fac.total

            custAcum = abs(custAcum)

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:
                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:
                    costomInd = tableOperacion()
                    costomInd.fecha = toddy
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factType.objects.get(nombre=ty)
                    costomInd.tabTipo = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
            
            custAcum = 0

    for nom in cantAuxOp:

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=toddy)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal

        if nom["principal"] == True:

            acumTablaTotales = acumTablaTotales + acum

        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=toddy).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy).order_by("tabTipo__nombre")

    # ----------- Categoria -------------------

    toddy = request.POST.get("searchDate")
    tod = request.POST.get("searchDate")
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    acum = 0

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    todDate = datetime.strptime(tod, '%Y-%m-%d')
    tod = todDate.date()

    dic = {"acumTablaTotales":acumTablaTotales,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"totalParcialOp":totalParcialOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"contTotal":contTotal,"tod":tod,"allFacturesToPay":allFacturesToPay,"allFacturesToCollect":allFacturesToCollect,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    return render(request,"spareapp/contDay.html",dic)

def contByDayCustom(request):

    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    facAuxAllCat = ""

    if request.method == "POST":

        acumTablaTotales = 0
        acumTablaTotalesCat = 0

        tod = request.POST.get("searchDate")
        tods = request.POST.get("fecha")
        toddy = request.POST.get("searchDate")

        if tods:

            tod = tods
            toddy = tods

        allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
        allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
        
        facturesToCollect = len(allFacturesToCollect)
        facturesToPay = len(allFacturesToPay)

        # ----------- Operacion -------------------

        acum = 0
        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factType.objects.all()
        totalParcialOp = {}
        
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

        if factureAuxOp:

            print("Hay facturas")

            if tableAuxOp:

                print("Hay tabla")
                allTypesCustom = factType.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total
                    customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacion()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factType.objects.get(nombre=ty)
                                costomInd.tabTipo = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0

            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")


                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total
                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacion()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        for nom in cantAuxOp:

            suma = 0

            aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal

            if nom["principal"] == False:

                acumTablaTotales = acumTablaTotales + acum
            
            totalParcialOp[nom["tabNombre"]] = acum

            acum = 0

        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabTipo__nombre")

        # ----------- Categoria -------------------

        toddy = request.POST.get("searchDate")
        tod = request.POST.get("searchDate")
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

        if factureAuxCat:

            print("Hay facturas")

            if tableAuxCat:

                print("Hay tabla")
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

                    for fac in facAuxAllCat:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        acum = 0

        for nom in cantAuxCat:

            aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal

            if nom["principal"] == False:

                acumTablaTotalesCat = acumTablaTotalesCat + acum
            
            totalParcialOpCat[nom["tabNombre"]] = acum

            acum = 0

        cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

        todDate = datetime.strptime(tod, '%Y-%m-%d')
        tod = todDate.date()

    dic = {"acumTablaTotalesCat":acumTablaTotalesCat,"acumTablaTotales":acumTablaTotales,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"totalParcialOp":totalParcialOp,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"tod":tod,"allFacturesToPay":allFacturesToPay,"allFacturesToCollect":allFacturesToCollect,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    return render(request,"spareapp/customTables.html",dic)

def contByRangeCustom(request):

    tod = request.POST.get("searchDate")
    sumaLista = ""
    restaLista = ""
    sumaRestaTotal = 0
    cantAuxEmpty = False
    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    if request.method == "POST":

        acumTablaTotales = 0
        acumTablaTotalesCat = 0

        tod1 = request.POST.get("fecha1")
        tod2 = request.POST.get("fecha2")

        dateFrom = request.POST.get("searchDateFrom")
        dateTo = request.POST.get("searchDateTo")

        if tod1 or tod2:

            dateFrom = tod1
            dateTo = tod2

        allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
        allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
        
        facturesToCollect = len(allFacturesToCollect)
        facturesToPay = len(allFacturesToPay)

        custAcum = 0

        # ----------- Operacion -------------------

        tableAuxOpAux = tableOperacionAux.objects.all()
        totalParcialOp = {}
        for all in tableAuxOpAux:

            all.delete()
        
        acum = 0
        allTypesCustom = factType.objects.all()
        totalParcialOp = {}
        
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty).exclude(pendiente=False,refType__facCobrar=True)
            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            for fac in facAuxAll:
                custAcum = custAcum + fac.total

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    costomInd = tableOperacionAux()
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factType.objects.get(nombre=ty)
                    costomInd.tabTipo = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
                
            custAcum = 0
            
        cantAuxOp = tableOperacion.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

        realAcum = 0

        for nom in cantAuxOp:

            aux1 = tableOperacion.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabTipo__nombre","suma").distinct()
            aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).exclude(pendiente=False,refType__facCobrar=True)

            for a in aux1:

                for b in aux2:

                    if a["tabTipo__nombre"] == b.refType.nombre:

                        acum = acum + b.total

                if a["suma"]==True:
                    realAcum = realAcum + acum
                else:
                    realAcum = realAcum - acum
                    
                acum = 0

            if nom["principal"] == False:

                acumTablaTotales = acumTablaTotales + realAcum

            totalParcialOp[nom["tabNombre"]] = realAcum

            realAcum = 0

        cantAuxOp = tableOperacionAux.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOp = tableOperacionAux.objects.all().order_by("tabTipo__nombre")

        # ----------- Categoria -------------------

        dateFrom = request.POST.get("searchDateFrom")
        dateTo = request.POST.get("searchDateTo")
        custAcum = 0
        tableAuxOpAuxCat = tableOperacionAuxCat.objects.all()
        for all in tableAuxOpAuxCat:

            all.delete()
        
        acum = 0
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)
            
            for fac in facAuxAll:
                custAcum = custAcum + fac.total

            lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    costomInd = tableOperacionAuxCat()
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factCategory.objects.get(nombre=ty)
                    costomInd.tabCat = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
                
            custAcum = 0
           
        cantAuxOpCat = tableOperacionCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

        realAcum = 0

        for nom in cantAuxOpCat:

            aux1 = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabCat__nombre","suma").distinct()
            aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

            for a in aux1:

                for b in aux2:

                    if a["tabCat__nombre"] == b.refCategory.nombre:

                        acum = acum + b.total

                if a["suma"]==True:
                    realAcum = realAcum + acum
                else:
                    realAcum = realAcum - acum
                    
                acum = 0

            if nom["principal"] == False:

                acumTablaTotalesCat = acumTablaTotalesCat + realAcum

            totalParcialOpCat[nom["tabNombre"]] = realAcum

            realAcum = 0

        cantAuxOpCat = tableOperacionAuxCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOpCat = tableOperacionAuxCat.objects.all().order_by("tabCat__nombre")
        
        todFrom = datetime.strptime(dateFrom, '%Y-%m-%d')
        dateFrom = todFrom.date()
        todTo = datetime.strptime(dateTo, '%Y-%m-%d')
        dateTo = todTo.date()

        dic = {"acumTablaTotales":acumTablaTotales,"acumTablaTotalesCat":acumTablaTotalesCat,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxOp":tableAuxOp,"tableAuxOpAux":tableAuxOpAux,"totalParcialOp":totalParcialOp,"cantAuxOp":cantAuxOp,"cantAuxEmpty":cantAuxEmpty,"sumaRestaTotal":sumaRestaTotal,"restaLista":restaLista,"sumaLista":sumaLista,"tod":tod,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    return render(request,"spareapp/customTables.html",dic)

def contByRange(request):

    cantAuxEmpty = False
    custAcum = 0
    acumTablaTotales = 0
    acumTablaTotalesCat = 0
    totalcito = 0

    tod = request.POST.get("searchDate")

    if request.method == "POST":

        custAcum = 0
        cantAuxEmpty = False
        dateFrom = request.POST.get("searchDateFrom")
        dateTo = request.POST.get("searchDateTo")

        contTotal = 0
        noIncludeTotal = 0
        noIncludeTotalGasto = 0
        contPagadoCobrado = 0

        allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
        allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
        
        facturesToCollect = len(allFacturesToCollect)
        facturesToPay = len(allFacturesToPay)

        # ----------- Operacion -------------------

        tableAuxOpAux = tableOperacionAux.objects.all()
        for all in tableAuxOpAux:

            all.delete()
        
        acum = 0
        allTypesCustom = factType.objects.all()
        totalParcialOp = {}

        # ----------- Recorrido largo ----------------
        
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty).exclude(pendiente=False,refType__facCobrar=True)
            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                for fac in facAuxAll:

                    if fac.nc == True:

                        custAcum = custAcum + fac.total

                    else:

                        custAcum = custAcum - fac.total

            else:

                for fac in facAuxAll:

                    if fac.nc == True:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                        else:

                            custAcum = custAcum + fac.total

                    else:

                        if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada":

                                custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum + fac.total

            custAcum = abs(custAcum)

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    costomInd = tableOperacionAux()
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factType.objects.get(nombre=ty)
                    costomInd.tabTipo = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
                
            custAcum = 0

        # ------- Termina recorrido largo -----------------
           
        cantAuxOp = tableOperacion.objects.all().values("tabNombre","principal","suma").order_by("tabNombre").distinct()

        realAcum = 0

        for nom in cantAuxOp:

            aux1 = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date__gte=dateFrom,fecha__date__lte=dateTo).all()
            print(aux1)
            print("...")
            # print(aux1[0])
            for a in aux1:
                print("Tipo: "+str(a.tabTipo)+" Total: "+str(a.tabTotal))
            # print(aux1["tabNombre"])
            totalcito = aux1.aggregate(Sum("tabTotal"))
            print("Total: "+str(totalcito["tabTotal__sum"]))
            print(totalcito["tabTotal__sum"])
            if nom["suma"]:
                totalParcialOp[nom["tabNombre"]] = totalcito["tabTotal__sum"]
            else:
                totalParcialOp[nom["tabNombre"]] = totalcito["tabTotal__sum"]*(-1)
            print("-------------------------------------------")

        cantAuxOp = tableOperacionAux.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOp = tableOperacionAux.objects.all().order_by("tabTipo__nombre")

        # ----------- Categoria -------------------

        dateFrom = request.POST.get("searchDateFrom")
        dateTo = request.POST.get("searchDateTo")
        custAcum = 0
        tableAuxOpAuxCat = tableOperacionAuxCat.objects.all()
        for all in tableAuxOpAuxCat:

            all.delete()
        
        acum = 0
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory=ty)
            
            for fac in facAuxAll:
                custAcum = custAcum + fac.total

            lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    costomInd = tableOperacionAuxCat()
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factCategory.objects.get(nombre=ty)
                    costomInd.tabCat = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
                
            custAcum = 0
           
        cantAuxOpCat = tableOperacionCat.objects.all().values("tabNombre","principal","suma").distinct()

        realAcum = 0

        for nom in cantAuxOpCat:

            aux1 = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date__gte=dateFrom,fecha__date__lte=dateTo).all().distinct()
            totalcito = aux1.aggregate(Sum("tabTotal"))
            print("Tabla: "+nom["tabNombre"])
            print(aux1)
            for a in aux1:
                print(a.tabNombre)
                print(a.tabCat)
                print(a.tabTotal)
                # print("Cat: "+str(aux1.tabCat__nombre))
                # print("Total: "+str(aux1.suma))
            print("..")
            print("totalcito: "+str(totalcito))
            if nom["suma"]:
                totalParcialOpCat[nom["tabNombre"]] = totalcito["tabTotal__sum"]
            else:
                totalParcialOpCat[nom["tabNombre"]] = totalcito["tabTotal__sum"]*(-1)

        # for nom in cantAuxOpCat:

        #     aux1 = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabCat__nombre","suma").distinct()
        #     aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo)

        #     for a in aux1:

        #         for b in aux2:

        #             if a["tabCat__nombre"] == b.refCategory.nombre:

        #                 acum = acum + b.total

        #         if a["suma"]==True:
        #             realAcum = realAcum + acum
        #         else:
        #             realAcum = realAcum - acum
                    
        #         acum = 0

        #     if nom["principal"] == True:

        #         acumTablaTotalesCat = acumTablaTotalesCat + realAcum

        #     totalParcialOpCat[nom["tabNombre"]] = realAcum

        #     realAcum = 0

        cantAuxOpCat = tableOperacionAuxCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOpCat = tableOperacionAuxCat.objects.all().order_by("tabCat__nombre")

        todFrom = datetime.strptime(dateFrom, '%Y-%m-%d')
        dateFrom = todFrom.date()
        todTo = datetime.strptime(dateTo, '%Y-%m-%d')
        dateTo = todTo.date()
        
        dic = {"acumTablaTotalesCat":acumTablaTotalesCat,"acumTablaTotales":acumTablaTotales,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"totalParcialOp":totalParcialOp,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"cantAuxEmpty":cantAuxEmpty,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"tod":tod,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"contTotal":contTotal}

    return render(request,"spareapp/contByRange.html",dic)

def contCollectFac(request,val):

    acum = 0

    tod = datetime.now().date()
    allTypes = factType.objects.all().order_by("nombre")
    factAux = factura.objects.filter(id=val)
    typeAux = factType.objects.filter(facCobrada=True).exclude(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")
    typeAux = typeAux[0]

    # print(val)
    # print(request.POST.get("filtro"+val))
    filtro = request.POST.get("filtro"+val)

    facAuxAllCat = ""

    if factAux:

        factErase = factura.objects.get(id=val)
        factErase.pendiente = False
        factErase.fechaCobrado = tod
        factErase.save()

        reciboCollect = factura()
        reciboCollect.fechaCreado=tod
        reciboCollect.num = factErase.num
        reciboCollect.refPersona = factErase.refPersona
        auxCat = factCategory.objects.get(nombre="Factura cobrada")
        if factAux[0].refCategory.nombre.lower() == "venta mayorista":
            auxCat = factCategory.objects.get(nombre="Factura cobrada (Mayorista)")
            typeAux = factType.objects.get(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")
        reciboCollect.refCategory = auxCat
        reciboCollect.refType = typeAux
        reciboCollect.fechaTope = factErase.fechaTope
        reciboCollect.fechaCobrado = tod
        reciboCollect.iva = factErase.iva
        reciboCollect.monto = factErase.monto
        reciboCollect.total = factErase.total
        reciboCollect.note=request.POST.get("contNota")
        reciboCollect.pendiente = False
        reciboCollect.save()

    contTotal = 0

    # ----------- Operacion -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
    factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

    if factureAuxOp:

        print("Hay facturas")

        if tableAuxOp:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0

        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxOp:

        suma = 0

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre")

    # ----------- Categoria -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

    allFacturesModal = allFacturesToCollect
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    editPrueba = False

    # contToCollect -------------------------------------------

    filterFactures = factura.objects.none()
    
    allFacturesPay = None
    dateFrom = None
    dateTo = None
    dayFrom = None
    dayTo = None

    searchMetodo = "all"
    if request.POST.get("search") == "all":
        searchMetodo = "all"
        allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())
    else:
        searchMetodo = "range"
        dateFrom = request.POST.get("dateFrom")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateTo = request.POST.get("dateTo")
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dateFrom = creadoAuxdateFrom.date()
        dateTo = creadoAuxdateTo.date()
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []

    filter = request.POST.get("filtro"+val)
    filter = filter.split(" ")

    filterFacturesDate = ""

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    for fil in filter:

        if filterFactures:

            filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

        else:

            filterFactures = factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    if filter != "":

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")


    if searchMetodo == "range":
        allFacturesPay = filterFactures & filterFacturesDate
    else:
        allFacturesPay = filterFactures

    deadlineDic = {}

    for all in allFacturesPay:

        deadline = datetime.now().date() - all.fechaCreado.date()
        deadline = deadline.days
        deadlineDic[all.id] = deadline

    for fac in allFacturesPay:

        deadline = datetime.now().date() - all.fechaCreado.date()
        deadline = deadline.days
        deadlineDic[all.id] = deadline

        acum = acum + fac.monto
        acum2 = acum2 + fac.total
        acumIva = acumIva + fac.iva

    if searchMetodo == "range":

        tod = None

    # allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaTope","id")
    allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    dic = {"allFacturesModal":allFacturesModal,"iva":acumIva,"filtro":filtro,"dateFrom":dateFrom,"dateTo":dateTo,"dayFrom":dayFrom,"dayTo":dayTo,"searchMetodo":searchMetodo,"filtro":filtro,"totalTotal":acum2,"montoTotal":acum,"deadlineDic":deadlineDic,"allTypes":allTypes,"allFacturesPay":allFacturesPay,"totalParcialOpCat":totalParcialOpCat,"tableAuxCat":tableAuxCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"contTotal":contTotal,"factAux":factAux,"tod":tod,"allTypes":allTypes,"editPrueba":editPrueba,"allFacturesToPay":allFacturesToPay,"allFacturesToCollect":allFacturesToCollect,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    # return redirect("/contToPay",dic)
    # return redirect("/contToCollect")
    # return render(request,"spareapp/contToCollect.html",dic)

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def contPayFac(request,val):

    print("------------------------Entra a pagar")

    # print(request.POST)
    # print(request.GET)
    # filtro3685 
    # entryspendingfiltro3685 on
    # print(val)
    # print(request.POST.get("filtro"+val))
    filtro = request.POST.get("filtro"+val)
    facAuxAllCat = ""
    # print(request.POST.get("entrySpendingFiltro"+val))

    acum = 0

    tod = datetime.now().date()
    allTypes = factType.objects.all().order_by("nombre")
    factAux = factura.objects.filter(id=val)

    typeAux = request.POST.get("contTypeIng")
    # print("typeAux")
    # print(typeAux)
    typeAux = factType.objects.get(id=typeAux)

    if factAux:

        factErase = factura.objects.get(id=val)
        factErase.pendiente = False
        factErase.fechaCobrado = tod
        factErase.save()

        reciboCollect = factura()
        reciboCollect.fechaCreado = tod
        reciboCollect.num = factErase.num
        reciboCollect.refPersona = factErase.refPersona
        reciboCollect.refType = typeAux
        if factAux[0].nc == True:
            reciboCollect.nc = True
        else:
            reciboCollect.nc = False
        auxCat = factCategory.objects.get(nombre="Mercancia credito pagada")
        reciboCollect.refCategory = auxCat
        reciboCollect.fechaTope = factErase.fechaTope
        reciboCollect.fechaCobrado = tod
        reciboCollect.iva = factErase.iva
        reciboCollect.monto = factErase.monto
        reciboCollect.note = request.POST.get("contNota")
        reciboCollect.total = factErase.total
        reciboCollect.pendiente = False
        reciboCollect.save()

    contTotal = 0

    # ----------- Operacion -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
    factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

    if factureAuxOp:

        print("Hay facturas")

        if tableAuxOp:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0

        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxOp:

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabTipo__nombre")

    # ----------- Categoria -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    editPrueba = False

    # contToPay ------------------------------------------

    searchMetodo = "all"
    dateFrom = ""
    dateTo = ""
    dayFrom = ""
    dayTo = ""

    if request.POST.get("search") == "all":
        searchMetodo = "all"
        allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())
    else:
        searchMetodo = "range"
        dateFrom = request.POST.get("dateFrom")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dateTo = request.POST.get("dateTo")
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dateFrom = creadoAuxdateFrom.date()
        dateTo = creadoAuxdateTo.date()
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    filterFactures = None
    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []

    checkeado = None

    if request.POST.get("entrySpendingFiltro"+val):

        checkeado = True
    
    else:

        checkeado = False

    filter = request.POST.get("filtro"+val)

    auxInicio = -1
    auxFin = -1
    acumCom = 0
    filterAux = filter
    palabras = []
    nuevoFilter = ""
    palabraFinal = ""

    for pos,let in enumerate(filter):

        if(let == '"'):

            acumCom = acumCom + 1

            if acumCom == 2:

                nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                auxFin = pos
                palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                palabras.append(palabraFinal)
                acumCom = 0
                auxInicio = -1
                auxFin = -1

            else:

                auxInicio = pos

    palabrasAux = []

    for val in palabras:

        if filterAux.find(val) >= 0:

            filterAux2 = filterAux.replace(val,"")
            filterAux = filterAux2
        palabrasAux.append(val.strip('"'))

    filterAux = filterAux.split(" ")
    filterAux = [item for item in filterAux if item]
    filter = filterAux + palabrasAux

    filterFacturesDate = ""

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    for fil in filter:

        if filterFactures:

            filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        else:

            filterFactures = ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )

    if checkeado == True:

        allFactures = factura.objects.filter(nc=checkeado)

    else:

        allFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")

    if filter:

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    if searchMetodo == "range":
        allFacturesPay = allFactures & filterFactures & filterFacturesDate
    else:
        allFacturesPay = allFactures & filterFactures
    allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

    deadlineDic = {}

    creadoAuxdateFrom = ""
    creadoAuxdateTo = ""

    for all in allFacturesPay:

        deadline = datetime.now().date() - all.fechaCreado.date()
        deadline = deadline.days
        deadlineDic[all.id] = deadline

    acum = 0
    acum2 = 0

    for fac in allFacturesPay:

        acum = acum + fac.monto
        acum2 = acum2 + fac.total
        acumIva = acumIva + fac.iva

    allFacturesModal = allFacturesToPay

    if searchMetodo == "range":

        tod = None

    dic = {"allFacturesModal":allFacturesModal,"dateFrom":dateFrom,"dateTo":dateTo,"dayFrom":dayFrom,"dayTo":dayTo,"iva":acumIva,"searchMetodo":searchMetodo,"filtro":filtro,"checkeado":checkeado,"deadlineDic":deadlineDic,"totalTotal":acum2,"montoTotal":acum,"allTypes":allTypes,"allFacturesPay":allFacturesPay,"totalParcialOpCat":totalParcialOpCat,"tableAuxCat":tableAuxCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"contTotal":contTotal,"factAux":factAux,"tod":tod,"allTypes":allTypes,"editPrueba":editPrueba,"allFacturesToPay":allFacturesToPay,"allFacturesToCollect":allFacturesToCollect,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    # return render(request,"spareapp/contToPay.html",dic)
    # return redirect("/contToPay",dic)
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def contPayFacType(request,val,val2,val3):

    facAuxAllCat = ""

    if request.method == "POST":

        # print("filtro"+val3)
        # print(request.POST)
        # print(request.GET)

        filtro = request.POST.get("filtro"+val3)
        # if filtro:
        #     pass
        # else:
        filtro = request.POST.get("cod2")

        acum = 0

        tod = datetime.now().date()
        factAux = factura.objects.filter(id=val3)
        typeAux = request.POST.get("contTypeIng")
        typeAux = factType.objects.get(id=typeAux)

        if typeAux.ingreso == True:
            allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)
        else:
            allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

        if factAux:

            factErase = factura.objects.get(id=val3)
            factErase.pendiente = False
            factErase.fechaCobrado = tod
            factErase.save()

            reciboCollect = factura()
            reciboCollect.fechaCreado = tod
            reciboCollect.num = factErase.num
            reciboCollect.refPersona = factErase.refPersona
            reciboCollect.refType = typeAux
            if factAux[0].nc == True:
                reciboCollect.nc = True
            else:
                reciboCollect.nc = False
            auxCat = factCategory.objects.get(nombre="Mercancia credito pagada")
            reciboCollect.refCategory = auxCat
            reciboCollect.fechaTope = factErase.fechaTope
            reciboCollect.fechaCobrado = tod
            reciboCollect.iva = factErase.iva
            reciboCollect.monto = factErase.monto
            reciboCollect.note = request.POST.get("contNota")
            reciboCollect.total = factErase.total
            reciboCollect.pendiente = False
            reciboCollect.save()

        # ----------- Operacion -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
        factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factType.objects.all()
        totalParcialOp = {}
        
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

        if factureAuxOp:

            if tableAuxOp:

                toddy = datetime.now().date()
                allTypesCustom = factType.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:

                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total

                    else:

                        for fac in facAuxAll:
                            if fac.nc == True:
                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:
                                    if fac.refCategory.nombre == "Mercancia credito pagada":
                                        custAcum = custAcum + fac.total
                                    else:
                                        custAcum = custAcum - fac.total
                                else:
                                    custAcum = custAcum + fac.total
                            else:
                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:
                                    if fac.refCategory.nombre == "Mercancia credito pagada":
                                        custAcum = custAcum - fac.total
                                    else:
                                        custAcum = custAcum + fac.total
                                else:
                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacion()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factType.objects.get(nombre=ty)
                                costomInd.tabTipo = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0

            else:

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:
                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total
                    else:

                        for fac in facAuxAll:
                            if fac.nc == True:
                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:
                                    if fac.refCategory.nombre == "Mercancia credito pagada":
                                        custAcum = custAcum + fac.total
                                    else:
                                        custAcum = custAcum - fac.total
                                else:
                                    custAcum = custAcum + fac.total
                            else:
                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:
                                    if fac.refCategory.nombre == "Mercancia credito pagada":
                                        custAcum = custAcum - fac.total
                                    else:
                                        custAcum = custAcum + fac.total
                                else:
                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacion()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        for nom in cantAuxOp:

            aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:
                    acum = acum + a.tabTotal
                else:
                    acum = acum - a.tabTotal
            
            totalParcialOp[nom["tabNombre"]] = acum

            acum = 0

        cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabTipo__nombre")

        # ----------- Categoria -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

        if factureAuxCat:

            if tableAuxCat:

                toddy = datetime.now().date()
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                    for fac in facAuxAllCat:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        for nom in cantAuxCat:

            aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

            for a in aux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal
            
            totalParcialOpCat[nom["tabNombre"]] = acum

            acum = 0

        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

        allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
        allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
        
        facturesToCollect = len(allFacturesToCollect)
        facturesToPay = len(allFacturesToPay)

        # contType --------------------------------------------

        tod = val2

        allFactures = factura.objects.none()
        allFacturesVal = None
        filterFactures = factura.objects.none()
        filterPersonas = factura.objects.none()
        filterCategorys = factura.objects.none()
        filterFacturesE = factura.objects.none()
        filterCategorysE = factura.objects.none()
        filterFacturesDate = factura.objects.none()
        acum = 0
        acum2 = 0
        acumIva = 0
        deadline = ""
        deadlineDic = []
        dateDic = []
        auxFac = None
        auxPer = None
        auxCat = None
        dayFrom = None
        dayTo = None
        dateFrom = None
        dateTo = None

        allFacturesAux = None

        # print("Antes del error: val2")
        # print(request.POST)
        # print(request.GET)
        # print(val2)

        # diaAux = datetime.strptime(str(val2),"%Y-%m-%d")

        if request.POST.get("dateTo"):

            searchMetodo = "range"
            dateFrom = request.POST.get("dateFrom")
            dateTo = request.POST.get("dateTo")
            dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
            dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            diaAux = creadoAuxdateFrom
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())

        else:

            searchMetodo = "all"
            diaAux = datetime.strptime(str(val2),"%Y-%m-%d")
            allFacturesPay = factura.objects.filter(fechaCreado__date=diaAux.date(),refType__nombre=val).order_by("fechaTope")
            if allFacturesPay:
                dateFrom = allFacturesPay[0].fechaCreado.date()
                creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
                dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
                dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
                dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
                dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
                dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
                creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
                dayFrom = dateFrom
                dayTo = dateTo
                dateFrom = str(creadoAuxdateFrom.date())
                dateTo = str(creadoAuxdateTo.date())

        # print(request.POST)
        # print(request.GET)
        # print("dateFrom")
        # print(dateFrom)
        # print("dateTo")
        # print(dateTo)

        if request.POST.get("dateTo"):
            allFacturesVal = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val).order_by("fechaCreado","id")
        else:
            allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val).order_by("fechaCreado","id")
        allFacturesAux = allFacturesVal

        allFacturesModal = allFacturesAux

        acum = 0
        acum2 = 0
        acumIva = 0
        deadline = ""
        deadlineDic = []
        dateDic = []
        palabrasErase = []
        palabrasErase2 = []

        filter = request.POST.get("filtro"+val3)

        # if filter:
        #     pass
        # else:
        filter = request.POST.get("cod2")
        # print("filter")
        # print(filter)
        filtro = None
        filtro = filter
        # print("filtro")
        # print(filtro)

        auxInicio = -1
        auxInicioe = -1
        auxFin = -1
        acumCom = 0
        filterAux = filter
        filterAuxErase = filter
        palabras = []
        nuevoFilter = ""
        palabraFinal = ""

        for pos,let in enumerate(filter):

            if(let == '-'):

                filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
                # auxFine = pos
                palabraFinalErase = filterErase.split(" ")
                palabrasErase.append(palabraFinalErase[0])
                auxInicioe = -1

            else:

                auxInicioe = pos

        for val in palabrasErase:

            palabrasErase2.append("-"+val)

        for pos,let in enumerate(filter):

            if(let == '"'):

                acumCom = acumCom + 1

                if acumCom == 2:

                    nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                    auxFin = pos
                    palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                    palabras.append(palabraFinal)
                    acumCom = 0
                    auxInicio = -1
                    auxFin = -1

                else:

                    auxInicio = pos

        palabrasAux = []

        for val in palabras:

            if filterAux.find(val) >= 0:

                filterAux2 = filterAux.replace(val,"")
                filterAux = filterAux2
            palabrasAux.append(val.strip('"'))

        filterAux = filterAux.split(" ")
        filterAux = [item for item in filterAux if item]
        filter = filterAux + palabrasAux
        for val in palabrasErase2:
            filter.remove(val)
        
        tod = val2
        typeAux = factType.objects.get(nombre=val)
        # val1 = None
        val1 = val

        if searchMetodo == "range":
            filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
            filterFacturesDatePersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
            filterFacturesDateCategories = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

        tod = diaAux.date()

        for fil in filter:

            if dateTo:

                if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )

                if typeAux.mercPagada == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

                if typeAux.facCobrada == True:

                    if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                        auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                        auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                        auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                    else:

                        auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                        auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                        auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

                if typeAux.facCobrar == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

                if typeAux.mercPagar == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

            else:

                if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )

                if typeAux.mercPagada == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

                if typeAux.facCobrada == True:

                    if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                        auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                        auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                        auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                    else:

                        auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                        auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                        auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

                if typeAux.facCobrar == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

                if typeAux.mercPagar == True:

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

            if filter.index(fil) == 0:

                filterFactures = auxFac
                filterPersonas = auxPer
                filterCategorys = auxCat

            if filterFactures:

                filterFactures = filterFactures & auxFac
                filterPersonas = filterPersonas & auxPer
                filterCategorys = filterCategorys & auxCat

            else:

                filterFactures = auxFac
                filterPersonas = auxPer
                filterCategorys = auxCat

        if filter:

            pass

        else:

            if dateTo:

                if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                    filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

                if typeAux.mercPagada == True:

                    filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

                if typeAux.facCobrada == True:

                    if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                        filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                        filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")

                    else:

                        filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                        filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

                if typeAux.facCobrar == True:

                    filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

                if typeAux.mercPagar == True:

                    filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")

            else:

                if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                    filterFactures = factura.objects.filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")

                if typeAux.mercPagada == True:

                    filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

                if typeAux.facCobrada == True:

                    if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                        filterFactures = factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                        filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                
                    else:

                        filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                        filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

                if typeAux.facCobrar == True:

                    filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

                if typeAux.mercPagar == True:

                    filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                
        for fil in palabrasErase:

            if filterFacturesE:

                filterFacturesE = filterFacturesE & ( factura.objects.all().exclude(num__icontains=fil))
                filterPersonasE = filterPersonasE & ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
                filterCategorysE = filterCategorysE & ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

            else:

                filterFacturesE = ( factura.objects.all().exclude(num__icontains=fil))
                filterPersonasE = ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
                filterCategorysE = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

        if palabrasErase:

            pass

        else:

            filterFacturesE = factura.objects.all().order_by("fechaCreado","id")
            filterPersonasE = factura.objects.values("refPersona__nombre").all().order_by("fechaCreado","id")
            filterCategorysE = factura.objects.values("refCategory__nombre","refCategory__ingreso").all().order_by("fechaCreado","id")

        if searchMetodo == "range":
            allFactures = filterFactures & filterFacturesDate & filterFacturesE
        else:
            allFactures = filterFactures & filterFacturesE

        if typeAux.facCobrada == False and typeAux.mercPagada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

            if typeAux.ingreso:

                for fac in allFactures:

                    deadline = datetime.now().date() - fac.fechaCreado.date()
                    deadlineDic.append(deadline.days)
                    dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                    acum = acum + fac.monto
                    acumIva = acumIva + fac.iva
                    acum2 = acum2 + fac.total

            else:

                for fac in allFactures:

                    deadline = datetime.now().date() - fac.fechaCreado.date()
                    deadlineDic.append(deadline.days)
                    dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                    if fac.nc == True:

                        acum = acum + fac.monto
                        acumIva = acumIva + fac.iva
                        acum2 = acum2 + fac.total

                    else:

                        acum = acum - fac.monto
                        acumIva = acumIva - fac.iva
                        acum2 = acum2 - fac.total

        if typeAux.facCobrada == True:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

        if typeAux.mercPagada == True:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                if fac.nc == True:

                    acum = acum + fac.monto
                    acumIva = acumIva + fac.iva
                    acum2 = acum2 + fac.total

                else:

                    acum = acum - fac.monto
                    acumIva = acumIva - fac.iva
                    acum2 = acum2 - fac.total

        if typeAux.facCobrar == True:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

        if typeAux.mercPagar == True:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                if fac.nc == True:

                    acum = acum - fac.monto
                    acumIva = acumIva - fac.iva
                    acum2 = acum2 - fac.total

                else:

                    acum = acum + fac.monto
                    acumIva = acumIva + fac.iva
                    acum2 = acum2 + fac.total

        montoTotal = 0
        montoTotal = acum
        itbmTotal = 0
        itbmTotal = acumIva
        totalTotal = 0
        totalTotal = acum2
        typeDate = None
        typeDate = val2
        
        allFacturesVal = allFactures

        if searchMetodo == "range":

            dateFrom = dateFrom
            dateTo = dateTo
            tod = None

        else:

            dateFrom = None
            dateTo = None

        # dic = {"typeAux":typeAux,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"val2":val2,"val":val}
        dic = {"dateFrom":dateFrom,"dateTo":dateTo,"allFacturesModal":allFacturesModal,"filtro":filtro,"typeAux":typeAux,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"montoTotal":montoTotal,"itbmTotal":itbmTotal,"totalTotal":totalTotal,"typeDate":typeDate,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    # return redirect("contType")
    return render(request,"spareapp/contType.html",dic)

def contCollectFacType(request,val,val2,val3):

    acum = 0

    tod = datetime.now().date()
    allTypes = factType.objects.all().order_by("nombre")
    factAux = factura.objects.filter(id=val3)
    typeAux = factType.objects.filter(facCobrada=True).exclude(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")
    typeAux = typeAux[0]

    # allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val).order_by("fechaCreado","id")
    # allFacturesAux = allFacturesVal

    filtro = request.POST.get("filtro"+val3)
    print("filtro")
    print(filtro)

    facAuxAllCat = ""

    if factAux:

        factErase = factura.objects.get(id=val3)
        factErase.pendiente = False
        factErase.fechaCobrado = tod
        factErase.save()

        reciboCollect = factura()
        reciboCollect.fechaCreado=tod
        reciboCollect.num = factErase.num
        reciboCollect.refPersona = factErase.refPersona
        auxCat = factCategory.objects.get(nombre="Factura cobrada")
        if factAux[0].refCategory.nombre.lower() == "venta mayorista":
            auxCat = factCategory.objects.get(nombre="Factura cobrada (Mayorista)")
            typeAux = factType.objects.get(nombre="FACTURA CREDITO COBRADA (MAYORISTA)")
        reciboCollect.refCategory = auxCat
        reciboCollect.refType = typeAux
        reciboCollect.fechaTope = factErase.fechaTope
        reciboCollect.fechaCobrado = tod
        reciboCollect.iva = factErase.iva
        reciboCollect.monto = factErase.monto
        reciboCollect.total = factErase.total
        reciboCollect.note=request.POST.get("contNota")
        reciboCollect.pendiente = False
        reciboCollect.save()

    contTotal = 0

    # ----------- Operacion -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
    factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

    if factureAuxOp:

        print("Hay facturas")

        if tableAuxOp:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0

        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxOp:

        suma = 0

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre")

    # ----------- Categoria -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

    allFacturesModal = allFacturesToCollect
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    editPrueba = False

    # contType --------------------------------------------

    tod = val2

    allFactures = factura.objects.none()
    allFacturesVal = None
    filterFactures = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterCategorys = factura.objects.none()
    filterFacturesE = factura.objects.none()
    filterCategorysE = factura.objects.none()
    filterFacturesDate = factura.objects.none()
    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    auxFac = None
    auxPer = None
    auxCat = None
    dayFrom = None
    dayTo = None
    dateFrom = None
    dateTo = None

    allFacturesAux = None

    # print("Antes del error: val2")
    # print(request.POST)
    # print(request.GET)
    # print(val2)

    # diaAux = datetime.strptime(str(val2),"%Y-%m-%d")

    if request.POST.get("dateTo"):

        searchMetodo = "range"
        dateFrom = request.POST.get("dateFrom")
        dateTo = request.POST.get("dateTo")
        dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
        dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        diaAux = creadoAuxdateFrom
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    else:

        searchMetodo = "all"
        diaAux = datetime.strptime(str(val2),"%Y-%m-%d")
        allFacturesPay = factura.objects.filter(fechaCreado__date=diaAux.date(),refType__nombre=val).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
            dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())

    # print(request.POST)
    # print(request.GET)
    # print("dateFrom")
    # print(dateFrom)
    # print("dateTo")
    # print(dateTo)

    if request.POST.get("dateTo"):
        allFacturesVal = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val).order_by("fechaCreado","id")
    else:
        allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refType__nombre=val).order_by("fechaCreado","id")
    allFacturesAux = allFacturesVal

    allFacturesModal = allFacturesAux

    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    palabrasErase = []
    palabrasErase2 = []

    filter = request.POST.get("filtro"+val3)

    # if filter:
    #     pass
    # else:
    filter = request.POST.get("cod2")
    # print("filter")
    # print(filter)
    filtro = None
    filtro = filter
    # print("filtro")
    # print(filtro)

    auxInicio = -1
    auxInicioe = -1
    auxFin = -1
    acumCom = 0
    filterAux = filter
    filterAuxErase = filter
    palabras = []
    nuevoFilter = ""
    palabraFinal = ""

    for pos,let in enumerate(filter):

        if(let == '-'):

            filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
            # auxFine = pos
            palabraFinalErase = filterErase.split(" ")
            palabrasErase.append(palabraFinalErase[0])
            auxInicioe = -1

        else:

            auxInicioe = pos

    for val in palabrasErase:

        palabrasErase2.append("-"+val)

    for pos,let in enumerate(filter):

        if(let == '"'):

            acumCom = acumCom + 1

            if acumCom == 2:

                nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                auxFin = pos
                palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                palabras.append(palabraFinal)
                acumCom = 0
                auxInicio = -1
                auxFin = -1

            else:

                auxInicio = pos

    palabrasAux = []

    for val in palabras:

        if filterAux.find(val) >= 0:

            filterAux2 = filterAux.replace(val,"")
            filterAux = filterAux2
        palabrasAux.append(val.strip('"'))

    filterAux = filterAux.split(" ")
    filterAux = [item for item in filterAux if item]
    filter = filterAux + palabrasAux
    for val in palabrasErase2:
        filter.remove(val)
    
    tod = val2
    typeAux = factType.objects.get(nombre=val)
    # val1 = None
    val1 = val

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
        filterFacturesDatePersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
        filterFacturesDateCategories = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

    tod = diaAux.date()

    for fil in filter:

        if dateTo:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )

            if typeAux.mercPagada == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                else:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

            if typeAux.mercPagar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        else:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,refType__nombre=val1).order_by("fechaCreado","id") )

            if typeAux.mercPagada == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                else:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

            if typeAux.mercPagar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        if filter.index(fil) == 0:

            filterFactures = auxFac
            filterPersonas = auxPer
            filterCategorys = auxCat

        if filterFactures:

            filterFactures = filterFactures & auxFac
            filterPersonas = filterPersonas & auxPer
            filterCategorys = filterCategorys & auxCat

        else:

            filterFactures = auxFac
            filterPersonas = auxPer
            filterCategorys = auxCat

    if filter:

        pass

    else:

        if dateTo:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

            if typeAux.mercPagada == True:

                filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")

                else:

                    filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

            if typeAux.facCobrar == True:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

            if typeAux.mercPagar == True:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")

        else:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                filterFactures = factura.objects.filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")

            if typeAux.mercPagada == True:

                filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    filterFactures = factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
            
                else:

                    filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

            if typeAux.facCobrar == True:

                filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

            if typeAux.mercPagar == True:

                filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
            
    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.all().exclude(num__icontains=fil))
            filterPersonasE = filterPersonasE & ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
            filterCategorysE = filterCategorysE & ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

        else:

            filterFacturesE = ( factura.objects.all().exclude(num__icontains=fil))
            filterPersonasE = ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
            filterCategorysE = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.all().order_by("fechaCreado","id")
        filterPersonasE = factura.objects.values("refPersona__nombre").all().order_by("fechaCreado","id")
        filterCategorysE = factura.objects.values("refCategory__nombre","refCategory__ingreso").all().order_by("fechaCreado","id")

    if searchMetodo == "range":
        allFactures = filterFactures & filterFacturesDate & filterFacturesE
    else:
        allFactures = filterFactures & filterFacturesE

    if typeAux.facCobrada == False and typeAux.mercPagada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

        if typeAux.ingreso:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

        else:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                if fac.nc == True:

                    acum = acum + fac.monto
                    acumIva = acumIva + fac.iva
                    acum2 = acum2 + fac.total

                else:

                    acum = acum - fac.monto
                    acumIva = acumIva - fac.iva
                    acum2 = acum2 - fac.total

    if typeAux.facCobrada == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum + fac.monto
            acumIva = acumIva + fac.iva
            acum2 = acum2 + fac.total

    if typeAux.mercPagada == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            if fac.nc == True:

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

            else:

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

    if typeAux.facCobrar == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum - fac.monto
            acumIva = acumIva - fac.iva
            acum2 = acum2 - fac.total

    if typeAux.mercPagar == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            if fac.nc == True:

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

            else:

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

    montoTotal = 0
    montoTotal = acum
    itbmTotal = 0
    itbmTotal = acumIva
    totalTotal = 0
    totalTotal = acum2
    typeDate = None
    typeDate = val2
    
    allFacturesVal = allFactures

    if searchMetodo == "range":

        dateFrom = dateFrom
        dateTo = dateTo
        tod = None

    else:

        dateFrom = ""
        dateTo = ""

    # dic = {"typeAux":typeAux,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"val2":val2,"val":val}
    dic = {"dateFrom":dateFrom,"dateTo":dateTo,"allFacturesModal":allFacturesModal,"filtro":filtro,"typeAux":typeAux,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes,"montoTotal":montoTotal,"itbmTotal":itbmTotal,"totalTotal":totalTotal,"typeDate":typeDate,"val2":val2,"allFacturesVal":allFacturesVal,"val":val}

    # return redirect("/contToPay",dic)
    # return redirect("/contToCollect")
    return render(request,"spareapp/contType.html",dic)

def contAddPerson(request):

    if request.method == "POST":

        nombreaux = request.POST.get("custName")
        identificacionaux = request.POST.get("custId")

        personProbar = persona.objects.filter(nombre=nombreaux,documento=identificacionaux)

        if personProbar:

            print("Ya existe")

        else:

            personAux = persona()
            personAux.nombre = nombreaux
            personAux.documento = identificacionaux
            personAux.gasto = True
            personAux.save()

            personLast = persona.objects.filter(id=personAux.id)
            lastPerson = list(personLast.values())
            
            return JsonResponse({'lastPerson':lastPerson})

    return render(request,"spareapp/contDay.html")

def accountStat(request):

    allCustomers = persona.objects.all()
    cont = 0
    pos = 0
    balance = {}
    balanceTotal = 0
    factureName = None
    dayFrom = ""
    dayTo = ""
    dateFrom = None
    dateTo = None
    balanceFacMerc = 0
    acumTotal = 0
    auxNombre = ""
    searchMetodo = "all"

    if request.method == "POST":

        auxNombre = request.POST.get("contNombre")
        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()
        cont = 0
        
        # for fac in factureName:

        #     if fac.refCategory.ingreso:

        #         cont = cont

        #         if fac.refType.facCobrar==True:

        #             cont = cont + fac.total
        #             if fac.pendiente == True:
        #                 balanceFacMerc = balanceFacMerc + fac.total
                
        #         if fac.refCategory.nombre=="Factura cobrada":

        #             cont = cont - fac.total
            
        #     else:

        #         cont = cont

        #         if fac.refType.mercPagar==True:

        #             cont = cont - fac.total
        #             if fac.pendiente == True:
        #                 balanceFacMerc = balanceFacMerc - fac.total
                
        #         if fac.refCategory.nombre=="Mercancia credito pagada":

        #             cont = cont + fac.total

        #     # if fac.pendiente == True and fac.refType.gasto == True:
        #     #     balance[fac.id] = [cont,fac.total*(-1)]
        #     # else:
        #     balance[fac.id] = [cont,fac.total]



        




        balanceTotal = cont

        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

        if request.POST.get("search") == "balance":

            print("Entra en balance")

            searchMetodo = "balance"

            for key in balance:

                if balance[key][0]==0:
                    pos = key
            
            facActAux = factura.objects.filter(id=pos)

            if facActAux:
            
                facAct = factura.objects.get(id=pos)
                factureName = factura.objects.filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
            
            else:

                factureName = None

            if factureName:
                pass
            else:
                factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")


        if request.POST.get("search") == "month":

            searchMetodo = "month"

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()

            mes = datetime.now().date().month
            dayFrom = dateFrom
            dayTo = datetime.now().date()
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")

        if request.POST.get("search") == "range":

            print("Entra en range")

            searchMetodo = "range"

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refPersona__id=auxNombre).order_by("fechaCreado")
    
        cont = 0

        # if factureName:

        #     # for fac in factureName:

        #     #     # print(fac)
        #     #     # print(cont)
        #     #     # print(fac.total)

        #     #     if fac.refCategory.ingreso:

        #     #         cont = cont

        #     #         if fac.refType.facCobrar==True:

        #     #             cont = cont + fac.total
        #     #             if fac.pendiente == True:
        #     #                 balanceFacMerc = balanceFacMerc + fac.total
                    
        #     #         if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

        #     #             cont = cont - fac.total
                
        #     #     else:

        #     #         cont = cont

        #     #         if fac.nc == True:

        #     #             if fac.refType.mercPagar==True:

        #     #                 cont = cont + fac.total
        #     #                 if fac.pendiente == True:
        #     #                     balanceFacMerc = balanceFacMerc - fac.total
                    
        #     #             if fac.refCategory.nombre=="Mercancia credito pagada":

        #     #                 cont = cont - fac.total

        #     #         else:

        #     #             if fac.refType.mercPagar==True:

        #     #                 cont = cont - fac.total
        #     #                 if fac.pendiente == True:
        #     #                     balanceFacMerc = balanceFacMerc - fac.total
                        
        #     #             if fac.refCategory.nombre=="Mercancia credito pagada":

        #     #                 cont = cont + fac.total

        #     #     # if fac.pendiente == True and fac.refType.gasto == True:
        #     #     #     balance[fac.id] = [cont,fac.total*(-1)]
        #     #     # else:
        #     #     balance[fac.id] = [cont,fac.total]

        #     for fac in factureName:

        #         # deadline = datetime.now().date() - fac.fechaCreado.date()
        #         # deadlineDic.append(deadline.days)
        #         # dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

        #         if fac.refCategory.ingreso:

        #             cont = cont

        #             if fac.refType.facCobrar==True:

        #                 cont = cont + fac.total
        #                 if fac.pendiente == True:
        #                     balanceFacMerc = balanceFacMerc + fac.total

        #             if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

        #                 cont = cont - fac.total
                
        #         else:

        #             cont = cont

        #             if fac.refType.mercPagar==True:

        #                 if fac.nc == True:
        #                     cont = cont + fac.total
        #                 else:
        #                     cont = cont - fac.total

        #                 # cont = cont - fac.total
        #                 if fac.pendiente == True:
        #                     balanceFacMerc = balanceFacMerc - fac.total
                    
        #             if fac.refCategory.nombre=="Mercancia credito pagada":

        #                 if fac.nc == False:
        #                     cont = cont + fac.total
        #                 else:
        #                     cont = cont - fac.total

        #                 # cont = cont + fac.total

        #         balance[fac.id] = [cont,fac.total]

        #         acumTotal = cont

        # balanceTotal = cont

        # ---------------------------------------
        acumTotal = 0

        if factureName:

            for fac in factureName:

                # deadline = datetime.now().date() - fac.fechaCreado.date()
                # deadlineDic.append(deadline.days)
                # dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                if fac.refCategory.ingreso:

                    # print("Ingreso")

                    cont = cont

                    if fac.refType.facCobrar==True:

                        cont = cont + fac.total
                        if fac.pendiente == True:
                            balanceFacMerc = balanceFacMerc + fac.total

                    if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

                        cont = cont - fac.total
                
                else:
                    # print("Egreso")

                    cont = cont

                    if fac.refType.mercPagar==True:

                        if fac.nc == True:
                            cont = cont + fac.total
                        else:
                            cont = cont - fac.total

                        if fac.pendiente == True:
                            balanceFacMerc = balanceFacMerc - fac.total
                    
                    if fac.refCategory.nombre=="Mercancia credito pagada":

                        if fac.nc == False:
                            cont = cont + fac.total
                        else:
                            cont = cont - fac.total

                if fac.refType.ingreso == True and fac.refType.facCobrar == False or fac.refType.mercPagar == True:

                    acumTotal = acumTotal + abs(fac.total)
                
                else:

                    acumTotal = acumTotal - abs(fac.total)

                balance[fac.id] = [cont,fac.total]

        # print("Valor de acumtotal: "+str(acumTotal))
        # ---------------------------------------

        # acumTotal = 0

        # for facT in factureName:

        #     if facT.refType.ingreso == True:
        #         if (facT.refType.facCobrar == False):

        #             acumTotal = acumTotal + abs(facT.total)

        #         else:

        #             acumTotal = acumTotal - abs(facT.total)        
            
        #     else:

        #         if facT.nc == False:

        #             if facT.refType.mercPagar == True:

        #                 acumTotal = acumTotal + abs(facT.total)

        #             else:

        #                 acumTotal = acumTotal - abs(facT.total)

        #         else:

        #             if facT.refType.mercPagar == True:

        #                 acumTotal = acumTotal - abs(facT.total)

        #             else:

        #                 acumTotal = acumTotal + abs(facT.total)

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    tod = datetime.now().date()

    # print(acumTotal)
    # print(balance)

    dic = {"dateFrom":dateFrom,"dateTo":dateTo,"searchMetodo":searchMetodo,"auxNombre":auxNombre,"acumTotal":acumTotal,"tod":tod,"balanceFacMerc":balanceFacMerc,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"dayFrom":dayFrom,"dayTo":dayTo,"balanceTotal":balanceTotal,"balance":balance,"allCustomers":allCustomers,"factureName":factureName}

    return render(request,"spareapp/accountStat.html",dic)

def contLastMonth(request):

    tod = datetime.now().date()
    mes = datetime.now().date().month
    anio = datetime.now().date().year
    if mes == 1:
        mes = 12
        anio = anio - 1
    else:
        mes = mes - 1
    editPrueba = False
    allTypes = factType.objects.all().order_by("nombre")
    tableAux = mainTableAux.objects.all()

    date_today = datetime.now()
    dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
    dateFrom = dateFrom.date()

    now = datetime.now()
    start_month = datetime(now.year, now.month-1, 1)
    date_on_next_month = start_month + timedelta(35)
    start_next_month = datetime(date_on_next_month.year, date_on_next_month.month, 1)
    dateTo = start_next_month - timedelta(1)
    dateTo = dateTo.date()

    for all in tableAux:

        all.delete()

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    for typ in allTypes:

        tableAuxGet = mainTableAux()
        tableAuxGet.tabTipo = typ
        cont = 0

        tableAuxGet.tabTotal = cont
        tableAuxGet.save()

    tableAux = mainTableAux.objects.all().order_by("tabTipo__nombre")

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    editPrueba = False

    dic = {"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"editPrueba":editPrueba,"contTotal":contTotal,"dateTo":dateTo,"dateFrom":dateFrom,"editPrueba":editPrueba,"tableAux":tableAux}

    return render(request,"spareapp/contByRange.html",dic)

def contThisMonth(request):

    tod = datetime.now().date()
    mes = datetime.now().date().month
    anio = datetime.now().date().year
    allTypes = factType.objects.all().order_by("nombre")
    tableAux = mainTableAux.objects.all()

    editPrueba = False
    allTypes = factType.objects.all().order_by("nombre")
    tableAux = mainTableAux.objects.all()

    date_today = datetime.now()
    dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
    dateFrom = dateFrom.date()

    dateTo = tod

    for all in tableAux:

        all.delete()

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    for typ in allTypes:

        tableAuxGet = mainTableAux()
        tableAuxGet.tabTipo = typ
        cont = 0

        tableAuxGet.tabTotal = cont
        tableAuxGet.save()

    tableAux = mainTableAux.objects.all().order_by("tabTipo__nombre")

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"contTotal":contTotal,"editPrueba":editPrueba,"dateTo":dateTo,"dateFrom":dateFrom,"editPrueba":editPrueba,"tableAux":tableAux}

    return render(request,"spareapp/contByRange.html",dic)

def editeFact(request,val,val2):

    # http://localhost:8000/contType/CASH%3F2021-11-13

    # contType/CASH%3F2021-11-13
    # http://localhost:8000/editeFact/485/contTypeCASH%3F2021-11-13

    # http://localhost:8000/editeFact/485/contTypeCASH%3F2021-11-13

    # contCollectFac1917

    urlFinal = ""
    if val2.find("contType")>-1:
        typeA = val2.replace("contType","")
        urlFinal = "/contType/"+typeA
    if val2.find("contTypeCat")>-1:
        typeA = val2.replace("contTypeCat","")
        urlFinal = "/contTypeCat/"+typeA
    if val2.find("contToCollect")>-1:
        typeA = val2.replace("contToCollect","")
        urlFinal = "/contToCollect"
    if val2.find("contCollectFac")>-1:
        typeA = val2.replace("contCollectFac","")
        urlFinal = "/contToCollect"
    if val2.find("contPayFac")>-1:
        typeA = val2.replace("contPayFac","")
        urlFinal = "/contToPay"
    if val2.find("contToPay")>-1:
        typeA = val2.replace("contToPay","")
        urlFinal = "/contToPay"
    if val2.find("searchTable")>-1:
        # typeA = val2.replace("searchTable","")
        # urlFinal = "/searchTable/"+typeA
        fAux = factura.objects.get(id=val)
        pAux = fAux.refPersona.id
        typeA = val2.replace("accountStat","")
        urlFinal = "/contIndividual/"+str(pAux)
        # urlFinal = "/accountDay"
        # urlFinal = "/contDay"
    if val2.find("contTypeTarjeta")>-1:
        typeA = val2.replace("contTypeTarjeta","")
        urlFinal = "/contTypeTarjeta/"+typeA
    if val2.find("accountDay")>-1:
        typeA = val2.replace("accountDay","")
        urlFinal = "/accountDay"
    if val2.find("accountStat")>-1:
        fAux = factura.objects.get(id=val)
        pAux = fAux.refPersona.id
        typeA = val2.replace("accountStat","")
        urlFinal = "/contIndividual/"+str(pAux)
    if val2.find("contIndividual")>-1:
        fAux = factura.objects.get(id=val)
        pAux = fAux.refPersona.id
        typeA = val2.replace("contIndividual","")
        urlFinal = "/contIndividual/"+str(pAux)
    if val2.find("contListByType")>-1:
        fAux = factura.objects.get(id=val)
        typeA = val2.replace("contListByType","")
        urlFinal = "/contListByType/"+typeA
    if val2.find("contListByCategory")>-1:
        typeA = val2.replace("contListByCategory","")
        urlFinal = "/contListByCategory/"+typeA
    check = False
    allCustomers = persona.objects.all()
    allTypes = factType.objects.all()
    # DEPENDE DEL TYPE SALEN UNOS U OTROS
    facAux = factura.objects.filter(id=val)
    facTotal = 0
    auxFacGet = factura.objects.get(id=val)
    tope=""
    # if auxFacGet.pendiente == True:
    #     tope=str(auxFacGet.fechaTope.year)+"-"+str('%02d' % auxFacGet.fechaTope.month)+"-"+str('%02d' % auxFacGet.fechaTope.day)
    todold = auxFacGet.fechaCreado.date()
    # actual = ""
    actual=str(auxFacGet.fechaCreado.date().year)+"-"+str('%02d' % auxFacGet.fechaCreado.date().month)+"-"+str('%02d' % auxFacGet.fechaCreado.date().day)
    if auxFacGet.monto == auxFacGet.total:
        check = False
    else:
        check = True

    if facAux[0].refCategory.ingreso == True:

        allCategories = factCategory.objects.filter(ingreso=True)

    else:

        allCategories = factCategory.objects.filter(egreso=True)

    facAuxAllCat = ""

    if request.method == "POST":

        returnPath = request.POST.get("returnPath")
        fechaAct = request.POST.get("searchDateFrom")
        contNombre = request.POST.get("contNombre")
        nomAux = persona.objects.get(id=contNombre)
        factAux = factura.objects.get(id=request.POST.get("facId"))
        fechaAct = datetime.strptime(fechaAct,'%Y-%m-%d')
        if factAux.fechaCreado.date() != fechaAct.date():
            factAux.fechaCreado = fechaAct
        else:
            print("Iguales")
        factAux.refPersona = nomAux

        if request.POST.get("contNumFac"):
            contNumFac = request.POST.get("contNumFac")
            factAux.num = contNumFac
        else:
            factAux.num = None

        contTypeIng = request.POST.get("contTypeIng")
        typeAux = factType.objects.get(id=contTypeIng)
        factAux.refType = typeAux

        contCatIng = request.POST.get("contCatIng")
        catAux = factCategory.objects.filter(id=contCatIng)

        catAux = factCategory.objects.get(id=contCatIng)
        factAux.refCategory = catAux

        if request.POST.get("contFechaTope") != "":
            contFechaTope = request.POST.get("contFechaTope")
            factAux.fechaTope = contFechaTope

        if factAux.refCategory.limite == True:
            factAux.pendiente = True
        else:
            factAux.pendiente = False

        contMonto = request.POST.get("contMonto")
        factAux.monto = str(contMonto).replace(',','.')

        if request.POST.get("contItbm") == "":
            contIva = float(0)
        else:
            contIva = request.POST.get("contItbm")
        factAux.iva = str(contIva).replace(',','.')

        contTotal = request.POST.get("contTotal")
        factAux.total = str(contTotal).replace(',','.')

        if request.POST.get("notaCredito"):
            factAux.nc = True
        else:
            factAux.nc = False

        factAux.note = request.POST.get("contNota")

        factAux.save()

        tod = factAux.fechaCreado.date()
        allTypes = factType.objects.all().order_by("nombre")

        tod = todold
        allTypes = factType.objects.all().order_by("nombre")

        toddy = datetime.now().date()
        allTypesCustom = factType.objects.all()
        custAcum = 0
        # ----------- Operacion -------------------

        tod = datetime.now().date()
        factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factType.objects.all()
        
        tableAuxOp = tableOperacion.objects.filter(fecha__date=tod)

        if factureAuxOp:

            print("Hay facturas")

            if tableAuxOp:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factType.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                    else:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum + fac.total

                                    else:

                                        custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum - fac.total

                                    else:

                                        custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacion()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factType.objects.get(nombre=ty)
                                costomInd.tabTipo = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum - fac.total

                    else:

                        for fac in facAuxAll:

                            if fac.nc == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum + fac.total

                                    else:

                                        custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                    if fac.refCategory.nombre == "Mercancia credito pagada":

                                        custAcum = custAcum - fac.total

                                    else:

                                        custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum + fac.total

                    custAcum = abs(custAcum)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacion()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        # ----------- Categoria -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

        if factureAuxCat:

            print("Hay facturas")

            if tableAuxCat:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                    for fac in facAuxAllCat:

                        if ty.ingreso == True:
                            if fac.nc == True:
                                custAcum = custAcum - fac.total
                            else:
                                custAcum = custAcum + fac.total
                        else:
                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                    for fac in facAuxAllCat:

                        if ty.ingreso == True:
                            if fac.nc == True:
                                custAcum = custAcum - fac.total
                            else:
                                custAcum = custAcum + fac.total
                        else:
                            if fac.nc == True:
                                custAcum = custAcum + fac.total
                            else:
                                custAcum = custAcum - fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        tod = datetime.now().date()

        dic = {"actual":actual,"allCustomers":allCustomers,"tod":tod,"allTypes":allTypes,"allCategories":allCategories}
    
        return redirect(urlFinal)

    dic = {"tope":tope,"urlFinal":urlFinal,"actual":actual,"check":check,"allCategories":allCategories,"allTypes":allTypes,"allCustomers":allCustomers,"facAux":facAux}

    return render(request,"spareapp/editeFact.html",dic)

# Borrar?
def contTotalDay(request):

    allTypes = factType.objects.all().order_by("nombre")
    editPrueba = False

    if request.method == "POST":

        # print(request.POST)

        fechaAux = request.POST.get("searchDate")
        tableAuxPro = mainTable.objects.filter(fecha__date=fechaAux)

        if tableAuxPro:

            print("Ya existe la tabla")

        else:

            print("No existe la tabla")

            for typ in allTypes:

                tableAux = mainTable()
                tableAux.fecha = fechaAux
                tableAux.tabTipo = typ

                if typ.manual:

                    monto = request.POST.get(str(typ.nombre).replace(" ", "")+"Total")
                    if monto:
                        tableAux.tabTotal = float(monto)
                    else:
                        tableAux.tabTotal = float(0)
                else:

                    if typ.facCobrado == True:

                        print("Factura cobrado")
                    
                    tableAux.tabTotal = float(0)
                
                tableAux.save()
                tableChange = mainTable.objects.get(id=tableAux.id)
                tableChange.fecha=fechaAux
                tableChange.save()

    dic = {"editPrueba":editPrueba,"allTypes":allTypes}

    return render(request,"spareapp/contTotalDay.html",dic)

def contIndividual(request,val):

    print("Id: "+str(val))
    dayFrom = ""
    dayTo = ""

    tod = datetime.now().date()
    acumTotal = 0
    balanceFacMerc = 0
    auxNombre = val
    searchMetodo = "all"
    dateFrom = ""
    dateTo = ""
    pos = 0

    allCustomers = persona.objects.all().order_by("nombre")
    personaAux = persona.objects.get(id=val)
    print(personaAux)
    print(personaAux.id)
    print(type(personaAux))
    factureName = factura.objects.filter(refPersona=personaAux).order_by("fechaCreado","id")
    print(factureName)
    balance = {}
    cont = 0
    balanceTotal = 0

    for fac in factureName:

        if fac.refCategory.ingreso:

            cont = cont

            if fac.refType.facCobrar==True:

                cont = cont + fac.total
            
            if fac.refCategory.nombre=="Factura cobrada":

                cont = cont - fac.total
        
        else:

            cont = cont

            if fac.refType.mercPagar==True:

                if fac.nc == True:
                    cont = cont + fac.total
                else:
                    cont = cont - fac.total

                # cont = cont - fac.total
            
            if fac.refCategory.nombre=="Mercancia credito pagada":

                if fac.nc == False:
                    cont = cont + fac.total
                else:
                    cont = cont - fac.total

                # cont = cont + fac.total

        # if fac.pendiente == True and  fac.refType.gasto == True:
        #     balance[fac.id] = [cont,fac.total*(-1)]
        # else:
        balance[fac.id] = [cont,fac.total]

        acumTotal = cont

    balanceTotal = cont

    if factureName:
        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    # acumTotal = 0

    # for facT in factureName:

    #         if facT.refType.ingreso == True:
    #             if (facT.refType.facCobrar == False):

    #                 acumTotal = acumTotal + abs(facT.total)

    #             else:

    #                 acumTotal = acumTotal - abs(facT.total)        
            
    #         else:

    #             if facT.nc == False:

    #                 if facT.refType.mercPagar == True:

    #                     acumTotal = acumTotal + abs(facT.total)

    #                 else:

    #                     acumTotal = acumTotal - abs(facT.total)

    #             else:

    #                 if facT.refType.mercPagar == True:

    #                     acumTotal = acumTotal - abs(facT.total)

    #                 else:

    #                     acumTotal = acumTotal + abs(facT.total)

    if request.method == "POST":

        auxNombre = request.POST.get("contNombre")
        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()
        cont = 0
        
        for fac in factureName:

            if fac.refCategory.ingreso:

                cont = cont

                if fac.refType.facCobrar==True:

                    cont = cont + fac.total
                    if fac.pendiente == True:
                        balanceFacMerc = balanceFacMerc + fac.total
                
                if fac.refCategory.nombre=="Factura cobrada":

                    cont = cont - fac.total
            
            else:

                cont = cont

                if fac.refType.mercPagar==True:

                    if fac.nc == True:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total

                    # cont = cont - fac.total
                    if fac.pendiente == True:
                        balanceFacMerc = balanceFacMerc - fac.total
                
                if fac.refCategory.nombre=="Mercancia credito pagada":

                    if fac.nc == False:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total

                    # cont = cont + fac.total

            # if fac.pendiente == True and fac.refType.gasto == True:
            #     balance[fac.id] = [cont,fac.total*(-1)]
            # else:
            balance[fac.id] = [cont,fac.total]

        balanceTotal = cont

        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

        if request.POST.get("search") == "balance":

            print("Entra en balance")

            searchMetodo = "balance"

            for key in balance:

                if balance[key][0]==0:
                    pos = key
            
            facActAux = factura.objects.filter(id=pos)

            if facActAux:
            
                facAct = factura.objects.get(id=pos)
                factureName = factura.objects.filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
            
            else:

                factureName = None

            if factureName:
                pass
            else:
                factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")


        if request.POST.get("search") == "month":

            searchMetodo = "month"

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()

            mes = datetime.now().date().month
            dayFrom = dateFrom
            dayTo = datetime.now().date()
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")

        if request.POST.get("search") == "range":

            print("Entra en range")

            searchMetodo = "range"

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")
            print(dateFrom)
            print(dateTo)

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refPersona__id=auxNombre).order_by("fechaCreado")
    
        cont = 0

        if factureName:

            for fac in factureName:

                # print(fac)
                # print(cont)
                # print(fac.total)

                if fac.refCategory.ingreso:

                    cont = cont

                    if fac.refType.facCobrar==True:

                        cont = cont + fac.total
                        if fac.pendiente == True:
                            balanceFacMerc = balanceFacMerc + fac.total
                    
                    if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

                        cont = cont - fac.total
                
                else:

                    cont = cont

                    if fac.refType.mercPagar==True:

                        if fac.nc == True:
                            cont = cont + fac.total
                        else:
                            cont = cont - fac.total

                        # cont = cont - fac.total
                        if fac.pendiente == True:
                            balanceFacMerc = balanceFacMerc - fac.total
                    
                    if fac.refCategory.nombre=="Mercancia credito pagada":

                        if fac.nc == False:
                            cont = cont + fac.total
                        else:
                            cont = cont - fac.total

                # if fac.pendiente == True and fac.refType.gasto == True:
                #     balance[fac.id] = [cont,fac.total*(-1)]
                # else:
                balance[fac.id] = [cont,fac.total]

                print(cont)

                acumTotal = cont

        balanceTotal = cont

        # acumTotal = 0

        # for facT in factureName:

        #     if facT.refType.ingreso == True:
        #         if (facT.refType.facCobrar == False):

        #             acumTotal = acumTotal + abs(facT.total)

        #         else:

        #             acumTotal = acumTotal - abs(facT.total)        
            
        #     else:

        #         if facT.nc == False:

        #             if facT.refType.mercPagar == True:

        #                 acumTotal = acumTotal + abs(facT.total)

        #             else:

        #                 acumTotal = acumTotal - abs(facT.total)

        #         else:

        #             if facT.refType.mercPagar == True:

        #                 acumTotal = acumTotal - abs(facT.total)

        #             else:

        #                 acumTotal = acumTotal + abs(facT.total)

    balanceTotal = cont

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"dateTo":dateTo,"dateFrom":dateFrom,"searchMetodo":searchMetodo,"auxNombre":auxNombre,"acumTotal":acumTotal,"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"dayFrom":dayFrom,"dayTo":dayTo,"allCustomers":allCustomers,"balanceTotal":balanceTotal,"balance":balance,"factureName":factureName}

    return render(request,"spareapp/accountStat.html",dic)

def factTypeES(request):

    cobrarPagar = None
    cate = None

    if request.GET.get("fecha") == "change":

        creado = request.GET.get("creado")
        creadoAux = datetime.strptime(creado,"%Y-%m-%d")
        deadlineDefault=(creadoAux+timedelta(days=30)).date()
        actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)

        actual = actualAux

        return JsonResponse({'actual':actual})


    cateAux = factCategory.objects.filter(nombre=request.GET.get("cat"))

    if cateAux:
        cate = factCategory.objects.get(nombre=request.GET.get("cat"))

    if request.GET.get("val") == "entry":

        if cateAux and (request.GET.get("cat") == "Factura cobrada" or request.GET.get("cat") == "Factura cobrada (Mayorista)"):
            # print("Existe factura y es factura cobrada")
            allCategories = factCategory.objects.filter(ingreso=True).order_by("nombre")
        else:
            # print("No existe factura o no es factura cobrada")
            allCategories = factCategory.objects.filter(ingreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada").exclude(nombre="Factura cobrada (Mayorista)")
        if cateAux and cate.limite == True:
            # print("Existe factura y la categoria limite es True")
            allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True)
        else:
            if cateAux and cate.limite == False:
                # print("Existe factura y la categoría limite es False")
                allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
            else:
                # print("No existe factura o la categoria limite es False")
                allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
        cobrarPagar = factType.objects.filter(facCobrar=True)
    
    else:

        if cateAux and request.GET.get("cat") == "Mercancia credito pagada":
            allCategories = factCategory.objects.filter(egreso=True).order_by("nombre")
        else:
            allCategories = factCategory.objects.filter(egreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada").exclude(nombre="Factura cobrada (Mayorista)")
        if cateAux and cate.limite == True:
            allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(facCobrar=True)
        else:
            allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
        cobrarPagar = factType.objects.filter(mercPagar=True)

    allCategories = list(allCategories.values())
    allTypes = list(allTypes.values())
    cobrarPagar = list(cobrarPagar.values())

    return JsonResponse({'cobrarPagar':cobrarPagar,'allTypes':allTypes,'allCategories': allCategories})

def contListByTypeZero(request):

    allTypes = factType.objects.all().order_by("nombre")
    dayFrom = ""
    dayTo = ""
    factureName = None
    balanceTotal = 0
    itbmTotal = 0
    interesTotal = 0
    retencionTotal = 0
    tod = datetime.now().date()

    if request.method == "POST":

        val = request.POST.get("contNombre")
        factureName = factura.objects.filter(refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()

        balanceTotal = 0

        if request.POST.get("search") == "year":

            mes = datetime.now().date().year
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__year=anio,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)


        if request.POST.get("search") == "month":

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()
            dayFrom = dateFrom
            dayTo = datetime.now().date()

            mes = datetime.now().date().month
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

        if request.POST.get("search") == "range":

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
        
        for fac in factureName:
            balanceTotal = balanceTotal + fac.total
            itbmTotal = itbmTotal + fac.iva
            retencionTotal = retencionTotal + float(fac.iva/2)
            if fac.refType.visa == True:
                interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            if fac.refType.clave == True:
                interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"retencionTotal":retencionTotal,"interesTotal":interesTotal,"itbmTotal":itbmTotal,"balanceTotal":balanceTotal,"dayFrom":dayFrom,"dayTo":dayTo,"factureName":factureName,"allTypes":allTypes}

    return render(request,"spareapp/contListByType.html",dic)

def contListByType(request,val):

    tod = datetime.now().date()
    dayFrom = ""
    dayTo = ""
    balanceTotal = 0
    itbmTotal = 0
    interesTotal = 0
    retencionTotal = 0
    allTypes = factType.objects.all().order_by("nombre")
    factureName = factura.objects.filter(refType__nombre=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
    if factureName:
        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    for fac in factureName:
            balanceTotal = balanceTotal + fac.total
            itbmTotal = itbmTotal + fac.iva
            retencionTotal = retencionTotal + float(fac.iva/2)
            if fac.refType.visa == True:
                interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            if fac.refType.clave == True:
                interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)

    if request.method == "POST":

        val = request.POST.get("contNombre")
        factureName = factura.objects.filter(refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()

        balanceTotal = 0

# ----------------------------------------------

        if request.POST.get("search") == "year":

            mes = datetime.now().date().year
            # date_today = datetime.now()
            # dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            # dateFrom = dateFrom.date()
            # dayFrom = dateFrom
            # dayTo = datetime.now().date()

            # mes = datetime.now().date().month
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__year=anio,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

# ----------------------------------------------

        if request.POST.get("search") == "month":

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()
            dayFrom = dateFrom
            dayTo = datetime.now().date()

            mes = datetime.now().date().month
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

        if request.POST.get("search") == "range":

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
            
        for fac in factureName:
            balanceTotal = balanceTotal + fac.total
            itbmTotal = itbmTotal + fac.iva
            retencionTotal = retencionTotal + float(fac.iva/2)
            if fac.refType.visa == True:
                interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            if fac.refType.clave == True:
                interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)
    
    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"retencionTotal":retencionTotal,"interesTotal":interesTotal,"itbmTotal":itbmTotal,"balanceTotal":balanceTotal,"dayFrom":dayFrom,"dayTo":dayTo,"factureName":factureName,"allTypes":allTypes}

    return render(request,"spareapp/contListByType.html",dic)

def contListByCategoryZero(request):

    tod = datetime.now().date()
    allCategorys = factCategory.objects.all().order_by("nombre")
    dayFrom = ""
    dayTo = ""
    balanceTotal = 0
    factureName = None

    if request.method == "POST":

        val = request.POST.get("contNombre")
        factureName = factura.objects.filter(refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()

        balanceTotal = 0
        
        for fac in factureName:
            balanceTotal = balanceTotal + fac.total

        if request.POST.get("search") == "year":

            mes = datetime.now().date().year
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__year=anio,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)


        if request.POST.get("search") == "month":

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()
            dayFrom = dateFrom
            dayTo = datetime.now().date()

            mes = datetime.now().date().month
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

        if request.POST.get("search") == "range":

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
    
    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"tod":tod,"facturesToCollect":facturesToCollect,"facturesToPay":facturesToPay,"balanceTotal":balanceTotal,"dayFrom":dayFrom,"dayTo":dayTo,"factureName":factureName,"allCategorys":allCategorys}

    return render(request,"spareapp/contListByCategory.html",dic)

def contListByCategory(request,val):

    tod = datetime.now().date()
    dayFrom = ""
    dayTo = ""
    balanceTotal = 0
    allCategorys = factCategory.objects.all().order_by("nombre")
    factureName = factura.objects.filter(refCategory__nombre=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

    if factureName:
        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    for fac in factureName:
        balanceTotal = balanceTotal + fac.total

    if request.method == "POST":

        val = request.POST.get("contNombre")
        factureName = factura.objects.filter(refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()
        
        balanceTotal = 0
        
        for fac in factureName:
            balanceTotal = balanceTotal + fac.total

        if request.POST.get("search") == "year":

            mes = datetime.now().date().year
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__year=anio,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)


        if request.POST.get("search") == "month":

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()
            dayFrom = dateFrom
            dayTo = datetime.now().date()

            mes = datetime.now().date().month
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refType__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)

        if request.POST.get("search") == "range":

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__id=val).order_by("fechaCreado","id").exclude(refType__facCobrar=True,pendiente=False).exclude(refType__mercPagar=True,pendiente=False)
    
    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)
            
    dic = {"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"balanceTotal":balanceTotal,"dayFrom":dayFrom,"dayTo":dayTo,"factureName":factureName,"allCategorys":allCategorys}

    return render(request,"spareapp/contListByCategory.html",dic)

def contEditPerson(request,val):

    newCust = persona.objects.filter(id=val)
    newCust = persona.objects.get(id=val)

    newCust.nombre = request.POST.get("custName")
    if request.POST.get("custId"):
        newCust.documento = request.POST.get("custId")
    newCust.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    
def accountDay(request):

    tod = datetime.now().date()

    dayAux = tod

    dateFrom = tod
    dateTo = tod

    searchMetodo = "today"

    factureName = factura.objects.filter(fechaCreado=tod).order_by("fechaCreado","id","refType__nombre","refCategory__nombre","total")

    if request.method == "POST":

        if request.POST.get("search") == "byDay":

            searchMetodo = "day"

            tod = request.POST.get("searchDateFrom")

            # tod = dayAux

            factureName = factura.objects.filter(fechaCreado=tod).order_by("fechaCreado","id","refType__nombre","refCategory__nombre","total")

            # dayAux = factureName[0].fechaCreado.date()
            dayAux = tod
            # tod = request.POST.get("searchDateFrom")

    cont = 0
    cont2 = 0
    balance = {}
    balanceTotal = 0
    
    for fac in factureName:

        if (fac.refType.ingreso == True and fac.refType.facCobrar == False) or fac.refType.mercPagar == True:

            cont = cont + float(str(fac.total).replace(",","."))
            cont2 = cont2 + abs(float(str(fac.total).replace(",",".")))

        else:

            cont = cont - float(str(fac.total).replace(",","."))
            cont2 = cont2 - abs(float(str(fac.total).replace(",",".")))
        
        balance[fac.id] = cont

    balanceTotal = cont2

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"dayAux":dayAux,"searchMetodo":searchMetodo,"dateFrom":dateFrom,"dateTo":dateTo,"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"balanceTotal":balanceTotal,"balance":balance,"factureName":factureName}

    return render(request,"spareapp/accountDay.html",dic)

def filterAccountDay(request):

    print("Entra en filterAccountDay --------------------------------")

    print(request.GET)

    dateFrom = request.GET.get("dateFrom")
    dateTo = request.GET.get("dateTo")
    tod = datetime.now().date()
    dayOption=request.GET.get("searchDay")
    filtro=request.GET.get("filter")
    # frag=filtro.split(" ")
    frag=filtro
    factureNameVal=""

    dayFrom = datetime.now().date()
    dayTo = datetime.now().date()

    print("filter: "+str(filtro))

    if len(filtro.split(" "))>1:
        print("Se puede separar")
    else:
        print("No se puede separar")

    if request.GET.get("searchDay"):
        tod = dateFrom
    else:
        tod = tod

    # fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
    # fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

    # Filtro por fecha
    factureName = factura.objects.filter(fechaCreado__date=tod).order_by("fechaCreado","id").order_by("fechaCreado","id")
    # Filtro por filtro
    factureName = factureName.filter(Q(refPersona__nombre__icontains=frag) | Q(num__icontains=frag) | Q(refCategory__nombre__icontains=frag) | Q(refType__nombre__icontains=frag) | Q(note__icontains=frag)).order_by("fechaCreado","id")

    total = 0
    cont = 0
    cont2 = 0
    balanceFacMerc = 0
    balance = {}
    acumTotal = 0
    deadlineDic = []
    dateDic = []

    for fac in factureName:

        deadline = datetime.now().date() - fac.fechaCreado.date()
        deadlineDic.append(deadline.days)
        dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

        if (fac.refType.ingreso == True and fac.refType.facCobrar == False) or fac.refType.mercPagar == True:

            cont = cont + float(str(fac.total).replace(",","."))
            cont2 = cont2 + abs(float(str(fac.total).replace(",",".")))

        else:

            cont = cont - float(str(fac.total).replace(",","."))
            cont2 = cont2 - abs(float(str(fac.total).replace(",",".")))
        
        balance[fac.id] = cont

    acumTotal = cont2


    # for fac in factureName:

        # deadline = datetime.now().date() - fac.fechaCreado.date()
        # deadlineDic.append(deadline.days)
        # dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

    #     if fac.refCategory.ingreso:

    #         # cont = cont
    #         cont = cont + fac.total

    #         if fac.refType.facCobrar==True:

    #             cont = cont + fac.total
    #             if fac.pendiente == True:
    #                 balanceFacMerc = balanceFacMerc + fac.total

    #         if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

    #             cont = cont - fac.total
        
    #     else:

    #         cont = cont - fac.total
    #         # cont = cont

    #         if fac.refType.mercPagar==True:

    #             if fac.nc == True:
    #                 cont = cont + fac.total
    #             else:
    #                 cont = cont - fac.total

    #             # cont = cont - fac.total
    #             if fac.pendiente == True:
    #                 balanceFacMerc = balanceFacMerc - fac.total
            
    #         if fac.refCategory.nombre=="Mercancia credito pagada":

    #             cont = cont + fac.total

    #     balance[fac.id] = [cont,fac.total]
    #     acumTotal = cont

    if factureName:

        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    allFacturesQuery = list(factureName.values())

    factureNameVal = factureName.values("refPersona__nombre","refCategory__nombre","refType__nombre").order_by("fechaCreado","id")

    allTypesQuery = list(factureNameVal)

    # print(allTypesQuery)
    # print("..........")
    # for fac in allTypesQuery:
    #     print(fac.refPersona)

    return JsonResponse({"allTypesQuery":allTypesQuery,"dateDic":dateDic,"dayFrom":dayFrom,"dayTo":dayTo,"dateTo":dateTo,"dateFrom":dateFrom,"acumTotal":acumTotal,"allFacturesQuery":allFacturesQuery})

    

def deleteFac(request,val):

    facAux = factura.objects.get(id=val)

    if facAux.refCategory.nombre == "Factura cobrada":

        facAnt = factura.objects.filter(num=facAux.num,refPersona=facAux.refPersona,fechaCobrado=facAux.fechaCreado.date()).exclude(refCategory__nombre="Factura cobrada")

        if len(facAnt)==1:

            facArreglar = factura.objects.exclude(refCategory__nombre="Factura cobrada").get(num=facAux.num,refPersona=facAux.refPersona,fechaCobrado=facAux.fechaCreado.date())
            facArreglar.pendiente = True
            facArreglar.fechaCobrado = None
            facArreglar.save()
    
    if facAux.refCategory.nombre == "Mercancia credito pagada":

        facAnt = factura.objects.filter(num=facAux.num,refPersona=facAux.refPersona,fechaCobrado=facAux.fechaCreado.date()).exclude(refCategory__nombre="Mercancia credito pagada")

        if len(facAnt)==1:

            facArreglar = factura.objects.exclude(refCategory__nombre="Mercancia credito pagada").get(num=facAux.num,refPersona=facAux.refPersona,fechaCobrado=facAux.fechaCreado.date())
            facArreglar.pendiente = True
            facArreglar.fechaCobrado = None
            facArreglar.save()

    facAux.delete()

    facAuxAllCat = ""

    tod = datetime.now().date()
    # ----------- Operacion -------------------
    toddy = datetime.now().date()
    allTypesCustom = factType.objects.all()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

    if tableAuxOp:

        print("Hay tabla")

        allTypesCustom = factType.objects.all()
        custAcum = 0
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

            for fac in facAuxAll:
                custAcum = custAcum + fac.total
            customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:

                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:

                    prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                    if prob2:

                        costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        costomInd.tabTotal = custAcum
                        costomInd.save()

                    else:

                        costomInd = tableOperacion()
                        costomInd.fecha = toddy
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
            custAcum = 0

    else:

        print("No hay tabla")
        custAcum = 0
        for ty in allTypesCustom:
            facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

            if ty.facCobrar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

            if ty.mercPagar == True:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

            if ty.mercPagada == True:
                facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

            if ty.facCobrada == True:

                if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                else:
                    facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")


            for fac in facAuxAll:
                custAcum = custAcum + fac.total
            lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
            for nom in lista:
                prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                if prob:
                    costomInd = tableOperacion()
                    costomInd.fecha = toddy
                    costomInd.tabNombre = nom["tabNombre"]
                    typeAux = factType.objects.get(nombre=ty)
                    costomInd.tabTipo = typeAux
                    costomInd.principal = principalAux[0]["principal"]
                    if sumaAux[0]["suma"]==True:
                        costomInd.suma = True
                        costomInd.resta = False
                    else:
                        costomInd.suma = False
                        costomInd.resta = True
                    costomInd.tabTotal = custAcum
                    costomInd.save()
            
            custAcum = 0

    # ----------- Categoria -------------------

    tod = datetime.now().date()
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def contAddTable(request):

    tod = datetime.now().date()
    acum = 0

    allTypes=factType.objects.all().order_by("nombre")

    if request.method == "POST":

        lista=request.POST.getlist("type")
        searchTable = customTable.objects.filter(tabNombre=request.POST.get("tabNombre"))

        if searchTable:

            print("Ya existe")

        else:

            for val in lista:

                tableAux = customTable()
                tableAux.fecha = tod
                tableAux.tabNombre = request.POST.get("tabNombre")
                typeAux = factType.objects.get(id=val)
                facAux = factura.objects.filter(fechaCreado__date=tod,refType=typeAux)
                for fac in facAux:
                    acum = acum + fac.total
                tableAux.tabTipo = typeAux
                if request.POST.get("tabPrincipal"):
                    tableAux.principal = True
                else:
                    tableAux.principal = False
                tableAux.tabTotal = acum
                acum = 0
                tableAux.save()

    dic={"allTypes":allTypes}

    return render(request,"spareapp/contAddTable.html",dic)

def customTables(request,val):

    print("Entra antes del POST")

    borrarAux = ""
    totalParcial = {}
    totalParcial2 = {}
    totalParcialOp = {}
    acum = 0
    sumaLista = ""
    restaLista = ""
    sumaRestaTotal = 0
    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    toddy = val
    tod = val

    facAuxAllCat = ""

    # ----------- Operacion -------------------

    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
    factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo")

    if factureAuxOp:

        print("Hay facturas")

        if tableAuxOp:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxOp:

        suma = 0

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=toddy).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy).order_by("tabTipo__nombre")

    # ----------- Categoria -------------------

    toddy = val
    tod = val
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

                for fac in facAuxAllCat:

                    if ty.ingreso == True:
                        if fac.nc == True:
                            custAcum = custAcum - fac.total
                        else:
                            custAcum = custAcum + fac.total
                    else:
                        if fac.nc == True:
                            custAcum = custAcum + fac.total
                        else:
                            custAcum = custAcum - fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

                for fac in facAuxAllCat:

                    if ty.ingreso == True:
                        if fac.nc == True:
                            custAcum = custAcum - fac.total
                        else:
                            custAcum = custAcum + fac.total
                    else:
                        if fac.nc == True:
                            custAcum = custAcum + fac.total
                        else:
                            custAcum = custAcum - fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    custAcum = abs(custAcum)

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal

        if nom["principal"] == False:

            acumTablaTotalesCat = acumTablaTotalesCat + acum
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    allTypes=factType.objects.all()
    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    if request.method == "POST":

        print("Entra despues del POST")

    todDate = datetime.strptime(tod, '%Y-%m-%d')
    tod = todDate.date()

    dic={"acumTablaTotales":acumTablaTotales,"acumTablaTotalesCat":acumTablaTotalesCat,"totalParcialOpCat":totalParcialOpCat,"tableAuxCat":tableAuxCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"cantAuxOp":cantAuxOp,"tableAuxOp":tableAuxOp,"totalParcialOp":totalParcialOp,"borrarAux":borrarAux,"totalParcial2":totalParcial2,"sumaRestaTotal":sumaRestaTotal,"restaLista":restaLista,"sumaLista":sumaLista,"totalParcial":totalParcial,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"tod":tod,"allTypes":allTypes}

    return render(request,"spareapp/customTables.html",dic)

def contListCustomTables(request):

    allTables = customTable.objects.all().order_by("tabNombre")
    allTablesNombres = customTable.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    if request.method == "POST":

        typeA = ""

        for val in request.POST:
            
            if val.find("nombre")>-1:

                typeA = val.replace("nombre","")

                if typeA == request.POST.get(val):

                    print("Son iguales")
                
                else:

                    auxChange = customTable.objects.filter(tabNombre=typeA)
                    for a in auxChange:
                        aux = customTable.objects.get(id=a.id)
                        aux.tabNombre = request.POST.get(val)
                        aux.save()

            typeA = ""

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTables.html",dic)

# Borrar?
def editeCustomTable(request,val):

    tod = datetime.now().date()
    customNombre = val
    allTypes = factType.objects.all().order_by("nombre")
    customAux = customTable.objects.filter(tabNombre=val)

    if request.method == "POST":

        allTypes = factType.objects.all()
        lista=request.POST.getlist("type")
        bandType = 0
        auxType = 0
        bandType2 = 0
        auxtype2 = 0

        if lista:

            for ty in allTypes:

                tableAntes = customTable.objects.filter(tabTipo=ty,tabNombre=val)

                for valor in lista:

                    if tableAntes:

                        if str(ty.id) == str(valor):

                            bandType = 1

                        else:

                            auxType = ty.id

                    else:

                        if str(ty.id) == str(valor):
                            
                            bandType2 = 1
                            auxType2 = valor

                if bandType == 0 and tableAntes:
                            
                    typeAux = factType.objects.get(id=auxType)
                    tableChange = customTable.objects.filter(tabTipo=typeAux,tabNombre=val)
                    tableChange.delete()

                if bandType2 == 1 and not(tableAntes):

                    typeAux = factType.objects.get(id=auxType2)
                    tableChange = customTable()
                    tableChange.tabNombre = val
                    tableChange.tabTipo = typeAux
                    if request.POST.get("tabPrincipal"):
                        tableChange.principal = True
                    else:
                        tableChange.principal = False
                    tableChange.tabTotal = 0
                    tableChange.save()

                bandType = 0
                bandType2 = 0

            # ------------------------------------------------------------------------
            # Para las tablas custom
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total
                customType = customTable.objects.filter(tabTipo=ty,tabNombre=val)
                for cus in customType:
                    costomInd = customTable.objects.get(id=cus.id)
                    if request.POST.get("tabPrincipal"):
                        costomInd.principal = True
                    else:
                        costomInd.principal = False
                    costomInd.tabTotal = custAcum
                    costomInd.save()
                custAcum = 0
            # ------------------------------------------------------------------------

        else:

            customAux = customTable.objects.filter(tabNombre=val)

            for cust in customAux:

                aux = customTable.objects.get(id=cust.id)
                aux.delete()

            allTables = customTable.objects.all().order_by("tabNombre")
            allTablesNombres = customTable.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

            dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

            return render(request,"spareapp/contListCustomTables.html",dic)

        allTables = customTable.objects.all().order_by("tabNombre")
        allTablesNombres = customTable.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

        dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

        return render(request,"spareapp/contListCustomTables.html",dic)

    dic = {"customNombre":customNombre,"customAux":customAux,"allTypes":allTypes,"val":val}

    return render(request,"spareapp/editeCustomTable.html",dic)

# Borrar?
def deleteCustom(request,val):

    customErase = customTable.objects.filter(tabNombre=val)
    customErase.delete()

    allTables = customTable.objects.all()
    allTablesNombres = customTable.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTables.html",dic)

def customTablesRange(request,val,val2):

    dateFrom = val
    dateTo = val2
    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0
    acumTablaTotalesCat = 0
    acumTablaTotales = 0

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    custAcum = 0
    # ----------- Operacion -------------------

    tableAuxOpAux = tableOperacionAux.objects.all()
    for all in tableAuxOpAux:

        all.delete()
    
    acum = 0
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    for ty in allTypesCustom:
        facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty).exclude(pendiente=False,refType__facCobrar=True)
        if ty.facCobrar == True:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

        if ty.mercPagar == True:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

        if ty.mercPagada == True:
            facAuxAll = factura.objects.filter(fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

        if ty.facCobrada == True:

            if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
            else:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

        for fac in facAuxAll:
            custAcum = custAcum + fac.total

        lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
        for nom in lista:

            prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
            principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
            sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
            if prob:

                costomInd = tableOperacionAux()
                costomInd.tabNombre = nom["tabNombre"]
                typeAux = factType.objects.get(nombre=ty)
                costomInd.tabTipo = typeAux
                costomInd.principal = principalAux[0]["principal"]
                if sumaAux[0]["suma"]==True:
                    costomInd.suma = True
                    costomInd.resta = False
                else:
                    costomInd.suma = False
                    costomInd.resta = True
                costomInd.tabTotal = custAcum
                costomInd.save()
            
        custAcum = 0
        
    cantAuxOp = tableOperacion.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    realAcum = 0

    for nom in cantAuxOp:

        aux1 = tableOperacion.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabTipo__nombre","suma").distinct()
        aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).exclude(pendiente=False,refType__facCobrar=True)

        for a in aux1:

            for b in aux2:

                if a["tabTipo__nombre"] == b.refType.nombre:

                    acum = acum + b.total

            if a["suma"]==True:
                realAcum = realAcum + acum
            else:
                realAcum = realAcum - acum
                
            acum = 0

        if nom["principal"] == False:

            acumTablaTotales = acumTablaTotales + realAcum

        totalParcialOp[nom["tabNombre"]] = realAcum

        realAcum = 0

    cantAuxOp = tableOperacionAux.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacionAux.objects.all().order_by("tabTipo__nombre")

    # ----------- Categoria -------------------

    dateFrom = val
    dateTo = val2
    custAcum = 0
    tableAuxOpAuxCat = tableOperacionAuxCat.objects.all()
    for all in tableAuxOpAuxCat:

        all.delete()
    
    acum = 0
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    
    for ty in allTypesCustom:
        facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory=ty).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__mercPagar=True)
        
        for fac in facAuxAll:
            custAcum = custAcum + fac.total

        lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
        for nom in lista:

            prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
            principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
            sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
            if prob:

                costomInd = tableOperacionAuxCat()
                costomInd.tabNombre = nom["tabNombre"]
                typeAux = factCategory.objects.get(nombre=ty)
                costomInd.tabCat = typeAux
                costomInd.principal = principalAux[0]["principal"]
                if sumaAux[0]["suma"]==True:
                    costomInd.suma = True
                    costomInd.resta = False
                else:
                    costomInd.suma = False
                    costomInd.resta = True
                costomInd.tabTotal = custAcum
                costomInd.save()
            
        custAcum = 0
        
    cantAuxOpCat = tableOperacionCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    realAcum = 0

    for nom in cantAuxOpCat:

        aux1 = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabCat__nombre","suma").distinct()
        aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo)

        for a in aux1:

            for b in aux2:

                if a["tabCat__nombre"] == b.refCategory.nombre:

                    acum = acum + b.total

            if a["suma"]==True:
                realAcum = realAcum + acum
            else:
                realAcum = realAcum - acum
                
            acum = 0

        if nom["principal"] == False:

            acumTablaTotalesCat = acumTablaTotalesCat + realAcum

        totalParcialOpCat[nom["tabNombre"]] = realAcum

        realAcum = 0

    cantAuxOpCat = tableOperacionAuxCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionAuxCat.objects.all().order_by("tabCat__nombre")
    
    todFrom = datetime.strptime(dateFrom, '%Y-%m-%d')
    dateFrom = todFrom.date()
    todTo = datetime.strptime(dateTo, '%Y-%m-%d')
    dateTo = todTo.date()

    dic = {"acumTablaTotalesCat":acumTablaTotalesCat,"acumTablaTotales":acumTablaTotales,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"totalParcialOp":totalParcialOp,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"contTotal":contTotal}

    return render(request,"spareapp/customTables.html",dic)

def contDayBack(request,val):

    tod = val
    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    tod = val

    facAuxAllCat = ""

    # ----------- Operacion -------------------

    # tod = datetime.now().date()
    toddy = val
    tod = val
    acum = 0
    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre").order_by("tabNombre").distinct()
    factureAuxOp = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo")

    if factureAuxOp:

        print("Hay facturas")

        if tableAuxOp:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=tod,pendiente=False,refCategory__egreso=True,refCategory__limite=False)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=tod,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                if ty.mercPagada == False and ty.mercPagar == False and ty.facCobrar == False and ty.facCobrada == False:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            custAcum = custAcum + fac.total

                        else:

                            custAcum = custAcum - fac.total

                else:

                    for fac in facAuxAll:

                        if fac.nc == True:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum + fac.total

                                else:

                                    custAcum = custAcum - fac.total

                            else:

                                custAcum = custAcum + fac.total

                        else:

                            if fac.refCategory.nombre == "Mercancia credito pagada" or fac.refType.mercPagar == True:

                                if fac.refCategory.nombre == "Mercancia credito pagada":

                                    custAcum = custAcum - fac.total

                                else:

                                    custAcum = custAcum + fac.total

                            else:

                                custAcum = custAcum + fac.total

                custAcum = abs(custAcum)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxOp:

        suma = 0

        aux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOp[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOp = tableOperacion.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAux2Op = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo__nombre")
    tableAuxOp = tableOperacion.objects.filter(fecha__date=tod).order_by("tabNombre","tabTipo__nombre")

    # ----------- Categoria -------------------

    tod = val
    acum = 0
    cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
    factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    custAcum = 0
    
    tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabNombre","tabCat")

    if factureAuxCat:

        print("Hay facturas")

        if tableAuxCat:

            print("Hay tabla")
            toddy = datetime.now().date()
            allTypesCustom = factCategory.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                for fac in facAuxAllCat:

                    if ty.ingreso == True:
                        if fac.nc == True:
                            custAcum = custAcum - fac.total
                        else:
                            custAcum = custAcum + fac.total
                    else:
                        if fac.nc == True:
                            custAcum = custAcum + fac.total
                        else:
                            custAcum = custAcum - fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacionCat()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0
        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                for fac in facAuxAllCat:

                    if ty.ingreso == True:
                        if fac.nc == True:
                            custAcum = custAcum - fac.total
                        else:
                            custAcum = custAcum + fac.total
                    else:
                        if fac.nc == True:
                            custAcum = custAcum + fac.total
                        else:
                            custAcum = custAcum - fac.total

                lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacionCat()
                        costomInd.fecha = tod
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factCategory.objects.get(nombre=ty)
                        costomInd.tabCat = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    for nom in cantAuxCat:

        aux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],fecha__date=tod)

        for a in aux:

            if a.suma == True:

                acum = acum + a.tabTotal

            else:

                acum = acum - a.tabTotal
        
        totalParcialOpCat[nom["tabNombre"]] = acum

        acum = 0

    cantAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionCat.objects.filter(fecha__date=tod).order_by("tabCat__nombre")

    todDate = datetime.strptime(tod, '%Y-%m-%d')
    tod = todDate.date()
    
    dic = {"acumTablaTotales":acumTablaTotales,"acumTablaTotalesCat":acumTablaTotalesCat,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"totalParcialOp":totalParcialOp,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"tod":tod,"allFacturesToPay":allFacturesToPay,"allFacturesToCollect":allFacturesToCollect,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    return render(request,"spareapp/contDay.html",dic)

def contDayBackRange(request,val,val2):

    dateFrom = val
    dateTo = val2

    acumTablaTotales = 0
    acumTablaTotalesCat = 0

    contTotal = 0
    noIncludeTotal = 0
    noIncludeTotalGasto = 0
    contPagadoCobrado = 0

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    custAcum = 0

    # ----------- Operacion -------------------

    tableAuxOpAux = tableOperacionAux.objects.all()
    for all in tableAuxOpAux:

        all.delete()
    
    acum = 0
    allTypesCustom = factType.objects.all()
    totalParcialOp = {}
    
    for ty in allTypesCustom:
        facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty).exclude(pendiente=False,refType__facCobrar=True)
        if ty.facCobrar == True:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

        if ty.mercPagar == True:
            facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

        if ty.mercPagada == True:
            facAuxAll = factura.objects.filter(fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

        if ty.facCobrada == True:

            if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
            else:
                facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

        for fac in facAuxAll:
            custAcum = custAcum + fac.total

        lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
        for nom in lista:

            prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
            principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
            sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
            if prob:

                costomInd = tableOperacionAux()
                costomInd.tabNombre = nom["tabNombre"]
                typeAux = factType.objects.get(nombre=ty)
                costomInd.tabTipo = typeAux
                costomInd.principal = principalAux[0]["principal"]
                if sumaAux[0]["suma"]==True:
                    costomInd.suma = True
                    costomInd.resta = False
                else:
                    costomInd.suma = False
                    costomInd.resta = True
                costomInd.tabTotal = custAcum
                costomInd.save()
            
        custAcum = 0
        
    cantAuxOp = tableOperacion.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    realAcum = 0

    for nom in cantAuxOp:

        aux1 = tableOperacion.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabTipo__nombre","suma").distinct()
        aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).exclude(pendiente=False,refType__facCobrar=True)

        for a in aux1:

            for b in aux2:

                if a["tabTipo__nombre"] == b.refType.nombre:

                    acum = acum + b.total

            if a["suma"]==True:
                realAcum = realAcum + acum
            else:
                realAcum = realAcum - acum
                
            acum = 0

        if nom["principal"] == True:

            acumTablaTotales = acumTablaTotales + realAcum

        totalParcialOp[nom["tabNombre"]] = realAcum

        realAcum = 0

    cantAuxOp = tableOperacionAux.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOp = tableOperacionAux.objects.all().order_by("tabTipo__nombre")
    print(tableAuxOp)

    # ----------- Categoria -------------------

    dateFrom = val
    dateTo = val2
    custAcum = 0
    tableAuxOpAuxCat = tableOperacionAuxCat.objects.all()
    for all in tableAuxOpAuxCat:

        all.delete()
    
    acum = 0
    allTypesCustom = factCategory.objects.all()
    totalParcialOpCat = {}
    
    for ty in allTypesCustom:
        facAuxAll = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory=ty)
        
        for fac in facAuxAll:
            custAcum = custAcum + fac.total

        lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
        for nom in lista:

            prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
            principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
            sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
            if prob:

                costomInd = tableOperacionAuxCat()
                costomInd.tabNombre = nom["tabNombre"]
                typeAux = factCategory.objects.get(nombre=ty)
                costomInd.tabCat = typeAux
                costomInd.principal = principalAux[0]["principal"]
                if sumaAux[0]["suma"]==True:
                    costomInd.suma = True
                    costomInd.resta = False
                else:
                    costomInd.suma = False
                    costomInd.resta = True
                costomInd.tabTotal = custAcum
                costomInd.save()
            
        custAcum = 0
        
    cantAuxOpCat = tableOperacionCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()

    realAcum = 0

    for nom in cantAuxOpCat:

        aux1 = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"]).values("tabNombre","tabCat__nombre","suma").distinct()
        aux2 = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo)

        for a in aux1:

            for b in aux2:

                if a["tabCat__nombre"] == b.refCategory.nombre:

                    acum = acum + b.total

            if a["suma"]==True:
                realAcum = realAcum + acum
            else:
                realAcum = realAcum - acum
                
            acum = 0

        if nom["principal"] == True:

            acumTablaTotalesCat = acumTablaTotalesCat + realAcum

        totalParcialOpCat[nom["tabNombre"]] = realAcum

        realAcum = 0

    cantAuxOpCat = tableOperacionAuxCat.objects.all().values("tabNombre","principal").order_by("tabNombre").distinct()
    tableAuxOpCat = tableOperacionAuxCat.objects.all().order_by("tabCat__nombre")
    
    todFrom = datetime.strptime(dateFrom, '%Y-%m-%d')
    dateFrom = todFrom.date()
    todTo = datetime.strptime(dateTo, '%Y-%m-%d')
    dateTo = todTo.date()

    dic = {"acumTablaTotalesCat":acumTablaTotalesCat,"acumTablaTotales":acumTablaTotales,"totalParcialOpCat":totalParcialOpCat,"tableAuxOpCat":tableAuxOpCat,"cantAuxOpCat":cantAuxOpCat,"tableAuxOp":tableAuxOp,"cantAuxOp":cantAuxOp,"totalParcialOp":totalParcialOp,"contPagadoCobrado":contPagadoCobrado,"noIncludeTotalGasto":noIncludeTotalGasto,"noIncludeTotal":noIncludeTotal,"dateFrom":dateFrom,"dateTo":dateTo,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"contTotal":contTotal}

    return render(request,"spareapp/contByRange.html",dic)

def searchTable(request):

    pos = 0
    cont = 0
    personaVarios = ""

    factureName = ""
    auxNombre = ""
    searchMetodo = ""

    balanceFacMerc = 0

    if request.method == "POST":

        balanceFacMerc = 0

        tod = datetime.now().date()
        print("Entra a searchTable POST")
        searchMetodo = "all"

        allCustomers = persona.objects.all().order_by("nombre")
        balance = {}
        cont = 0
        balanceTotal = 0

        auxNombre = request.POST.get("contNombre")
        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
        if factureName:
            dayFrom = factureName[0].fechaCreado.date()
            dayTo = factureName[len(factureName)-1].fechaCreado.date()
        cont = 0
        
        for fac in factureName:

            if fac.refCategory.ingreso:

                cont = cont

                if fac.refType.facCobrar==True:

                    cont = cont + fac.total
                
                if fac.refCategory.nombre=="Factura cobrada":

                    cont = cont - fac.total
            
            else:

                cont = cont

                if fac.refType.mercPagar==True:

                    if fac.nc == True:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total

                    # cont = cont - fac.total
                    if fac.pendiente == True:
                        balanceFacMerc = balanceFacMerc - fac.total
                
                if fac.refCategory.nombre=="Mercancia credito pagada":

                    if fac.nc == False:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total

            balance[fac.id] = [cont,fac.total]

        balanceTotal = cont

        factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

        if request.POST.get("search") == "balance":

            searchMetodo = "balance"

            for key in balance:

                if balance[key][0]==0:
                    pos = key
            
            facActAux = factura.objects.filter(id=pos)

            if facActAux:
            
                facAct = factura.objects.get(id=pos)
                factureName = factura.objects.filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
            
            else:

                factureName = None

            if factureName:

                pass

            else:

                factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

        if request.POST.get("search") == "month":

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()
            searchMetodo = "month"

            mes = datetime.now().date().month
            dayFrom = dateFrom
            dayTo = datetime.now().date()
            anio = datetime.now().date().year
            factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")

        if request.POST.get("search") == "range":

            dateFrom = request.POST.get("searchDateFrom")
            dateTo = request.POST.get("searchDateTo")

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:
                
                factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refPersona__id=auxNombre).order_by("fechaCreado")

        # BUSQUEDA ---------------------------------------------------------

        dayFrom = ""
        dayTo = ""

        busqueda = request.POST.get("Tsearch")

        if busqueda:

            tod = datetime.now().date()
            factureName = ""
            personaVarios = None
            searchMetodo = "all"
                    
            factureName = factura.objects.filter(num__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=busqueda).order_by("fechaCreado","id")
            facPersona = factura.objects.filter(num__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refPersona__nombre__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(note__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refType__nombre__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refCategory__nombre__icontains=busqueda).values("refPersona").distinct()

            if factureName:

                if len(facPersona)>1:
                    personaVarios = "Varios"
                else:
                    auxNombre = factureName[0].refPersona.id

                for fac in factureName:

                    if fac.refCategory.ingreso:

                        cont = cont

                        if fac.refType.facCobrar==True:

                            cont = cont + fac.total
                        
                        if fac.refCategory.nombre=="Factura cobrada":

                            cont = cont - fac.total
                    
                    else:

                        cont = cont

                        if fac.refType.mercPagar==True:

                            if fac.nc == True:
                                cont = cont + fac.total
                            else:
                                cont = cont - fac.total

                            # cont = cont - fac.total
                            if fac.pendiente == True:
                                balanceFacMerc = balanceFacMerc - fac.total
                        
                        if fac.refCategory.nombre=="Mercancia credito pagada":

                            if fac.nc == False:
                                cont = cont + fac.total
                            else:
                                cont = cont - fac.total

                    balance[fac.id] = [cont,fac.total]

                balanceTotal = cont

                dayFrom = factureName[0].fechaCreado.date()
                dayTo = factureName[len(factureName)-1].fechaCreado.date()

    acumTotal = 0

    for facT in factureName:

        if facT.refType.ingreso == True and facT.refType.facCobrar == False or facT.refType.mercPagar == True:

            acumTotal = acumTotal + abs(facT.total)
        
        else:

            acumTotal = acumTotal - abs(facT.total)

    balanceTotal = cont

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"searchMetodo":searchMetodo,"auxNombre":auxNombre,"acumTotal":acumTotal,"busqueda":busqueda,"personaVarios":personaVarios,"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"dayFrom":dayFrom,"dayTo":dayTo,"allCustomers":allCustomers,"balanceTotal":balanceTotal,"balance":balance,"factureName":factureName}

    return render(request,"spareapp/accountStat.html",dic)


def contAddCliente(request):

    if request.method == "POST":

        nombreCliente = request.POST.get("catNombre")
        identificacionCliente = request.POST.get("catId")

        nuevaPersona = persona()
        nuevaPersona.nombre = nombreCliente
        if identificacionCliente:
            nuevaPersona.documento = identificacionCliente
        nuevaPersona.gasto = True
        nuevaPersona.save()

        allTablesNombres = persona.objects.all().order_by("nombre")

        borrar = {}

        for per in allTablesNombres:

            if factura.objects.filter(refPersona__id=per.id):

                print("Se consiguió persona")

            else:

                borrar[per.id] = True

        dic = {"borrar":borrar,"allTablesNombres":allTablesNombres}

        return redirect("contListClienteTables")

        # return render(request,"spareapp/contListClienteTables.html",dic)

    return render(request,"spareapp/contAddCliente.html")

def contListClienteTables(request):

    if request.method == "POST":

        lista = persona.objects.all().order_by("nombre")

        for li in lista:

            principal = persona.objects.get(id=li.id)

            comparar = request.POST.get("nombre"+str(li.id))
            compIdentificacion = request.POST.get("identificacion"+str(li.id))
            compIngreso = request.POST.get("ingreso"+str(li.id))
            compEgreso = request.POST.get("egreso"+str(li.id))
            perId = request.POST.get("identificador"+str(li.id))

            if principal.nombre == comparar:
                pass
            else:
                principal.nombre = comparar

            if principal.documento == compIdentificacion:
                pass
            else:
                principal.documento = compIdentificacion

            if compIngreso:
                principal.ingreso = True
            else:
                principal.ingreso = False

            if compEgreso:
                principal.gasto = True
            else:
                principal.gasto = False

            principal.save()

            # if compIngreso:
            #     # print("Hay ingreso")
            #     cambiar = persona.objects.get(id=li.id)
            #     cambiar.ingreso = True
            #     cambiar.save()
            # else:
            #     cambiar = persona.objects.get(id=li.id)
            #     cambiar.ingreso = False
            #     cambiar.save()
            # if compEgreso:
            #     # print("Hay egreso")
            #     cambiar = persona.objects.get(id=li.id)
            #     cambiar.gasto = True
            #     cambiar.save()
            # else:
            #     cambiar = persona.objects.get(id=li.id)
            #     cambiar.gasto = False
            #     cambiar.save()

            # if str(li.nombre) == str(comparar):

            #     pass

            # else:

            #     cambiar = persona.objects.get(id=li.id)
            #     cambiar.nombre = str(comparar)
            #     cambiar.save()

            # if str(li.documento) == str(compIdentificacion):

            #     pass

            # else:

            #     cambiar = persona.objects.get(id=li.id)
            #     if compIdentificacion:
            #         cambiar.documento = str(compIdentificacion)
            #     else:
            #         cambiar.documento = ""
            #     cambiar.save()

    allTablesNombres = persona.objects.all().order_by("nombre")

    borrar = {}

    for per in allTablesNombres:

        if factura.objects.filter(refPersona__id=per.id):

            pass

        else:

            borrar[per.id] = True

    dic = {"borrar":borrar,"allTablesNombres":allTablesNombres}

    return render(request,"spareapp/contListClienteTables.html",dic)

def editeClienteTable(request,val):

    personaEditar = persona.objects.get(id=val)

    nombreCliente = personaEditar.nombre
    identificacionCliente = personaEditar.documento

    if request.method == "POST":

        nombreCliente = request.POST.get("catNombre")
        identificacionCliente = request.POST.get("catId")

        personaEditar.nombre = nombreCliente
        if identificacionCliente:
            personaEditar.documento = identificacionCliente
        personaEditar.save()

        allTablesNombres = persona.objects.all().order_by("nombre")

        borrar = {}

        for per in allTablesNombres:

            if factura.objects.filter(refPersona__id=per.id):

                pass

            else:

                borrar[per.id] = True

        dic = {"borrar":borrar,"allTablesNombres":allTablesNombres}

        return render(request,"spareapp/contListClienteTables.html",dic)

    dic = {"personaEditar":personaEditar}

    return render(request,"spareapp/contAddCliente.html",dic)

def deleteClienteTable(request,val):

    allTablesNombres = persona.objects.all().order_by("nombre")

    eraseCliente = persona.objects.filter(id=val)

    eraseCliente.delete()

    borrar = {}

    for per in allTablesNombres:

        if factura.objects.filter(refPersona__id=per.id):

            print("Se consiguió persona")

        else:

            borrar[per.id] = True

    dic = {"borrar":borrar,"allTablesNombres":allTablesNombres}

    return redirect("contListClienteTables")

    # return render(request,"spareapp/contListClienteTables.html",dic)

def contAddOperacion(request):

    tod = datetime.now().date()
    acum = 0
    cantAux = tableOperacion.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    allTypes=factType.objects.all().order_by("nombre")

    if request.method == "POST":

        # lista=request.POST.getlist("type")
        # nombreTabla = request.POST.get("tabNombre")
        # principal=request.POST.get("tabPrincipal")
        sumas = request.POST.getlist("TsumaVal")
        restas = request.POST.getlist("TrestaVal")
        searchTable = tableOperacion.objects.filter(tabNombre=request.POST.get("tabNombre"))

        if searchTable:

            print("Ya existe")

        else:

            for val in sumas:

                tableAux = tableOperacion()
                tableAux.suma = True
                tableAux.resta = False
                tableAux.fecha = tod
                tableAux.tabNombre = request.POST.get("tabNombre")
                typeAux = factType.objects.get(nombre=val)
                facAux = factura.objects.filter(fechaCreado__date=tod,refType=typeAux)
                for fac in facAux:
                    acum = acum + fac.total
                tableAux.tabTipo = typeAux
                if request.POST.get("tabPrincipal"):
                    tableAux.principal = True
                else:
                    tableAux.principal = False
                tableAux.tabTotal = acum
                acum = 0
                tableAux.save()

            for val in restas:

                tableAux = tableOperacion()
                tableAux.suma = False
                tableAux.resta = True
                tableAux.fecha = tod
                tableAux.tabNombre = request.POST.get("tabNombre")
                typeAux = factType.objects.get(nombre=val)
                facAux = factura.objects.filter(fechaCreado__date=tod,refType=typeAux)
                for fac in facAux:
                    acum = acum + fac.total
                tableAux.tabTipo = typeAux
                if request.POST.get("tabPrincipal"):
                    tableAux.principal = True
                else:
                    tableAux.principal = False
                tableAux.tabTotal = acum
                acum = 0
                tableAux.save()

    dic={"cantAux":cantAux,"allTypes":allTypes}

    return render(request,"spareapp/contAddOperacion.html",dic)

def contListCustomTablesOp(request):

    allTables = tableOperacion.objects.all().order_by("tabNombre")
    allTablesNombres = tableOperacion.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    if request.method == "POST":

        typeA = ""

        for val in request.POST:

            if val.find("nombre")>-1:

                typeA = val.replace("nombre","")

                auxPrincipal = request.POST.get("principal"+typeA)
                change = tableOperacion.objects.filter(tabNombre=typeA)
                if auxPrincipal:
                    for a in change:
                        aux = tableOperacion.objects.get(id=a.id)
                        aux.principal = True
                        aux.save()
                else:
                    for a in change:
                        aux = tableOperacion.objects.get(id=a.id)
                        aux.principal = False
                        aux.save()

                if typeA == request.POST.get(val):

                    pass
                
                else:

                    auxChange = tableOperacion.objects.filter(tabNombre=typeA)
                    for a in auxChange:
                        aux = tableOperacion.objects.get(id=a.id)
                        aux.tabNombre = request.POST.get(val)
                        aux.save()

            typeA = ""

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTablesOp.html",dic)

def editeCustomTableOp(request,val):

    tod = datetime.now().date()
    customNombre = val
    allTypes = factType.objects.all().order_by("nombre")
    customAux = tableOperacion.objects.filter(tabNombre=val)

    if request.method == "POST":

        allTypes = factType.objects.all()
        bandOriginal = False
        bandRevisar = False
        sumas=request.POST.getlist("TsumaVal")
        restas=request.POST.getlist("TrestaVal")
        customAux = tableOperacion.objects.filter(tabNombre=val)

        lista = tableOperacion.objects.values("tabTipo__nombre").filter(tabNombre=val).distinct()

        for a in customAux:

            b = tableOperacion.objects.get(id=a.id)

            if request.POST.get("tabPrincipal"):

                b.principal = True
            
            else:

                b.principal = False

            b.save()

        if lista:

            nombreTabla = val

        else:

            nombreTabla = request.POST.get("tabNombre")

        for sum in sumas:

            for li in lista:

                if li["tabTipo__nombre"] == sum:

                    bandOriginal = True

            if bandOriginal == True:

                print("Consigue")

            else:

                print("No consigue")
                agregar = tableOperacion()
                agregar.fecha = tod
                agregar.tabNombre = nombreTabla
                tableType = factType.objects.get(nombre=sum)
                agregar.tabTipo = tableType
                if request.POST.get("tabPrincipal"):
                    agregar.principal = True
                else:
                    agregar.principal = False
                agregar.suma = True
                agregar.resta = False
                agregar.tabTotal = 0
                agregar.save()

            bandOriginal = False

        for res in restas:

            for li in lista:

                if li["tabTipo__nombre"] == res:

                    bandOriginal = True

            if bandOriginal == True:

                print("Consigue")

            else:

                print("No consigue")
                agregar = tableOperacion()
                agregar.fecha = tod
                agregar.tabNombre = nombreTabla
                tableType = factType.objects.get(nombre=res)
                agregar.tabTipo = tableType
                if request.POST.get("tabPrincipal"):
                    agregar.principal = True
                else:
                    agregar.principal = False
                agregar.suma = False
                agregar.resta = True
                agregar.tabTotal = 0
                agregar.save()

            bandOriginal = False

        bandRevisar == False
        lista2 = tableOperacion.objects.all().filter(tabNombre=val)

        for li in lista2:

            for sum in sumas:

                if li.tabTipo.nombre == sum and li.suma == True:

                    bandRevisar = True

            if bandRevisar == False:

                eliminar = tableOperacion.objects.filter(tabNombre=val,tabTipo__nombre=li.tabTipo.nombre,suma=True)
                if eliminar:
                    eliminar.delete()

            else:

                print("Posee ese type en el form")

            bandRevisar = False

            for res in restas:

                if li.tabTipo.nombre == res:

                    bandRevisar = True

            if bandRevisar == False:

                eliminar = tableOperacion.objects.filter(tabNombre=val,tabTipo__nombre=li.tabTipo.nombre,resta=True)
                if eliminar:
                    eliminar.delete()

            else:

                print("Posee ese type en el form")

            bandRevisar = False

        allTables = tableOperacion.objects.all().order_by("tabNombre")
        allTablesNombres = tableOperacion.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

        dic = {"allTables":allTables,"allTablesNombres":allTablesNombres,"allTypes":allTypes}
        return render(request,"spareapp/contListCustomTablesOp.html",dic)

    dic = {"customNombre":customNombre,"customAux":customAux,"allTypes":allTypes,"val":val}

    return render(request,"spareapp/editeCustomTableOp.html",dic)

def deleteCustomOp(request,val):

    customErase = tableOperacion.objects.filter(tabNombre=val)
    customErase.delete()

    allTables = tableOperacion.objects.all()
    allTablesNombres = tableOperacion.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTablesOp.html",dic)

def deleteCustomOpCat(request,val):

    customErase = tableOperacionCat.objects.filter(tabNombre=val)
    customErase.delete()

    allTables = tableOperacionCat.objects.all()
    allTablesNombres = tableOperacionCat.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTablesCat.html",dic)

def editeFactAccount(request,val,val1,val2):

    cont = 0
    print("Entra en editeFactAccount")
    print(val)
    print(val1)
    print(val2)

    balanceFacMerc = 0

    dayFrom = ""
    dayTo = ""
    auxNombre = ""
    searchMetodo = ""

    check = False
    allCustomers = persona.objects.all()
    allTypes = factType.objects.all()
    # DEPENDE DEL TYPE SALEN UNOS U OTROS
    facAux = factura.objects.filter(id=val)
    auxFacGet = factura.objects.get(id=val)
    tope=""
    if auxFacGet.pendiente == True:
        tope=str(auxFacGet.fechaTope.year)+"-"+str('%02d' % auxFacGet.fechaTope.month)+"-"+str('%02d' % auxFacGet.fechaTope.day)
    todold = auxFacGet.fechaCreado.date()
    actual=str(auxFacGet.fechaCreado.date().year)+"-"+str('%02d' % auxFacGet.fechaCreado.date().month)+"-"+str('%02d' % auxFacGet.fechaCreado.date().day)
    if auxFacGet.monto == auxFacGet.total:
        check = False
    else:
        check = True

    if facAux[0].refCategory.ingreso == True:

        allCategories = factCategory.objects.filter(ingreso=True)

    else:

        allCategories = factCategory.objects.filter(egreso=True)

    facAuxAllCat = ""

    if request.method == "POST":

        balanceFacMerc = 0

        searchMetodo = "all"

        fechaAct = request.POST.get("searchDateFrom")
        contNombre = request.POST.get("contNombre")
        nomAux = persona.objects.get(id=contNombre)
        factAux = factura.objects.get(id=request.POST.get("facId"))
        fechaAct = datetime.strptime(fechaAct,'%Y-%m-%d')
        if factAux.fechaCreado.date() != fechaAct.date():
            factAux.fechaCreado = fechaAct
        else:
            print("Iguales")
        factAux.refPersona = nomAux

        if request.POST.get("contNumFac"):
            contNumFac = request.POST.get("contNumFac")
            factAux.num = contNumFac

        contTypeIng = request.POST.get("contTypeIng")
        typeAux = factType.objects.get(id=contTypeIng)
        factAux.refType = typeAux

        contCatIng = request.POST.get("contCatIng")
        catAux = factCategory.objects.filter(id=contCatIng)

        catAux = factCategory.objects.get(id=contCatIng)
        factAux.refCategory = catAux

        if request.POST.get("contFechaTope") != "":
            contFechaTope = request.POST.get("contFechaTope")
            factAux.fechaTope = contFechaTope

        if factAux.refCategory.limite == True:
            factAux.pendiente = True
        else:
            factAux.pendiente = False

        contMonto = request.POST.get("contMonto")
        factAux.monto = str(contMonto).replace(',','.')

        if request.POST.get("contItbm") == "":
            contIva = float(0)
        else:
            contIva = request.POST.get("contItbm")
        factAux.iva = str(contIva).replace(',','.')

        contTotal = request.POST.get("contTotal")
        factAux.total = str(contTotal).replace(',','.')

        factAux.note = request.POST.get("contNota")

        factAux.save()

        # Para la fecha nueva -----------------------------------------------------------

        tod = factAux.fechaCreado.date()

        allTypes = factType.objects.all().order_by("nombre")

        tod = todold
        allTypes = factType.objects.all().order_by("nombre")

        toddy = datetime.now().date()

        # BUSQUEDA ---------------------------------------------------------

        balance = {}
        balanceTotal = 0
        cont = 0

        busqueda = val1

        if busqueda:

            tod = datetime.now().date()
            factureName = ""
            personaVarios = None
                    
            factureName = factura.objects.filter(num__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=busqueda).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=busqueda).order_by("fechaCreado","id")
            facPersona = factura.objects.filter(num__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refPersona__nombre__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(note__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refType__nombre__icontains=busqueda).values("refPersona").distinct() | factura.objects.filter(refCategory__nombre__icontains=busqueda).values("refPersona").distinct()

            if len(facPersona)>1:
                personaVarios = "Varios"
            else:
                auxNombre = factureName[0].refPersona.id

            if factureName:

                for fac in factureName:

                    if fac.refCategory.ingreso:

                        cont = cont

                        if fac.refType.facCobrar==True:

                            cont = cont + fac.total
                        
                        if fac.refCategory.nombre=="Factura cobrada":

                            cont = cont - fac.total
                    
                    else:

                        cont = cont

                        if fac.refType.mercPagar==True:

                            if fac.nc == True:
                                cont = cont + fac.total
                            else:
                                cont = cont - fac.total

                            # cont = cont - fac.total
                            if fac.pendiente == True:
                                balanceFacMerc = balanceFacMerc - fac.total
                        
                        if fac.refCategory.nombre=="Mercancia credito pagada":

                            if fac.nc == False:
                                cont = cont + fac.total
                            else:
                                cont = cont - fac.total

                    # if fac.pendiente == True and  fac.refType.gasto == True:
                    #     balance[fac.id] = [cont,fac.total*(-1)]
                    # else:
                    balance[fac.id] = [cont,fac.total]

                balanceTotal = cont

                dayFrom = factureName[0].fechaCreado.date()
                dayTo = factureName[len(factureName)-1].fechaCreado.date()

        balanceTotal = cont

        allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
        allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
        facturesToCollect = len(allFacturesToCollect)
        facturesToPay = len(allFacturesToPay)

        # ----------- Operacion -------------------

        toddy = datetime.now().date()
        factureAuxOp = factura.objects.filter(fechaCreado__date=toddy)
        allTypesCustom = factType.objects.all()
        
        tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

        if factureAuxOp:

            if tableAuxOp:

                allTypesCustom = factType.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total
                    customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacion()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factType.objects.get(nombre=ty)
                                costomInd.tabTipo = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0

            else:

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                    if ty.facCobrar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                    if ty.mercPagar == True:
                        facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                    if ty.mercPagada == True:
                        facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                    if ty.facCobrada == True:

                        if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                        else:
                            facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")


                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total
                    lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy).order_by("tabTipo__nombre")

        # ----------- Categoria -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

        if factureAuxCat:

            print("Hay facturas")

            if tableAuxCat:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                    for fac in facAuxAllCat:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

        dic = {"busqueda":busqueda,"personaVarios":personaVarios,"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect,"dayFrom":dayFrom,"dayTo":dayTo,"allCustomers":allCustomers,"balanceTotal":balanceTotal,"balance":balance,"factureName":factureName}

        return render(request,"spareapp/accountStat.html",dic)

    dic = {"tope":tope,"actual":actual,"check":check,"allCategories":allCategories,"allTypes":allTypes,"allCustomers":allCustomers,"facAux":facAux}

    return render(request,"spareapp/editeFact.html",dic)

def contAddOperacionCat(request):

    tod = datetime.now().date()
    acum = 0
    cantAux = tableOperacionCat.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    allTypes=factCategory.objects.all().order_by("nombre")

    if request.method == "POST":

        sumas = request.POST.getlist("TsumaVal")
        restas = request.POST.getlist("TrestaVal")
        searchTable = tableOperacion.objects.filter(tabNombre=request.POST.get("tabNombre"))

        if searchTable:

            print("Ya existe")

        else:

            for val in sumas:

                tableAux = tableOperacionCat()
                tableAux.suma = True
                tableAux.resta = False
                tableAux.fecha = tod
                tableAux.tabNombre = request.POST.get("tabNombre")
                typeAux = factCategory.objects.get(nombre=val)
                facAux = factura.objects.filter(fechaCreado__date=tod,refCategory=typeAux)
                for fac in facAux:
                    acum = acum + fac.total
                tableAux.tabCat = typeAux
                if request.POST.get("tabPrincipal"):
                    tableAux.principal = True
                else:
                    tableAux.principal = False
                tableAux.tabTotal = acum
                acum = 0
                tableAux.save()

            for val in restas:

                tableAux = tableOperacionCat()
                tableAux.suma = False
                tableAux.resta = True
                tableAux.fecha = tod
                tableAux.tabNombre = request.POST.get("tabNombre")
                typeAux = factCategory.objects.get(nombre=val)
                facAux = factura.objects.filter(fechaCreado__date=tod,refCategory=typeAux)
                for fac in facAux:
                    acum = acum + fac.total
                tableAux.tabCat = typeAux
                if request.POST.get("tabPrincipal"):
                    tableAux.principal = True
                else:
                    tableAux.principal = False
                tableAux.tabTotal = acum
                acum = 0
                tableAux.save()

    dic={"cantAux":cantAux,"allTypes":allTypes}

    return render(request,"spareapp/contAddOperacionCat.html",dic)

def contListCustomTablesCat(request):

    allTables = tableOperacionCat.objects.all().order_by("tabNombre")
    allTablesNombres = tableOperacionCat.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

    if request.method == "POST":

        typeA = ""

        for val in request.POST:
            
            if val.find("nombre")>-1:

                typeA = val.replace("nombre","")

                auxPrincipal = request.POST.get("principal"+typeA)
                change = tableOperacionCat.objects.filter(tabNombre=typeA)
                if auxPrincipal:
                    for a in change:
                        aux = tableOperacionCat.objects.get(id=a.id)
                        aux.principal = True
                        aux.save()
                else:
                    for a in change:
                        aux = tableOperacionCat.objects.get(id=a.id)
                        aux.principal = False
                        aux.save()

                if typeA == request.POST.get(val):

                    print("Son iguales")
                
                else:

                    auxChange = tableOperacionCat.objects.filter(tabNombre=typeA)
                    for a in auxChange:
                        aux = tableOperacionCat.objects.get(id=a.id)
                        aux.tabNombre = request.POST.get(val)
                        aux.save()

            typeA = ""

    dic = {"allTablesNombres":allTablesNombres,"allTables":allTables}

    return render(request,"spareapp/contListCustomTablesCat.html",dic)

def editeCustomTableCat(request,val):

    tod = datetime.now().date()
    customNombre = val
    allTypes = factCategory.objects.all().order_by("nombre")
    customAux = tableOperacionCat.objects.filter(tabNombre=val)

    if request.method == "POST":

        allTypes = factCategory.objects.all()
        bandOriginal = False
        bandRevisar = False
        sumas=request.POST.getlist("TsumaVal")
        restas=request.POST.getlist("TrestaVal")
        customAux = tableOperacionCat.objects.filter(tabNombre=val)

        for a in customAux:

            b = tableOperacionCat.objects.get(id=a.id)

            if request.POST.get("tabPrincipal"):

                b.principal = True
            
            else:

                b.principal = False

            b.save()

        lista = tableOperacionCat.objects.values("tabCat__nombre").filter(tabNombre=val).distinct()

        if lista:

            nombreTabla = val

        else:

            nombreTabla = request.POST.get("tabNombre")

        for sum in sumas:

            for li in lista:

                if li["tabCat__nombre"] == sum:

                    bandOriginal = True

            if bandOriginal == True:

                print("Consigue")

            else:

                print("No consigue")
                agregar = tableOperacionCat()
                agregar.fecha = tod
                agregar.tabNombre = nombreTabla
                tableType = factCategory.objects.get(nombre=sum)
                agregar.tabCat = tableType
                if request.POST.get("tabPrincipal"):
                    agregar.principal = True
                else:
                    agregar.principal = False
                agregar.suma = True
                agregar.resta = False
                agregar.tabTotal = 0
                agregar.save()

            bandOriginal = False

        for res in restas:

            for li in lista:

                if li["tabCat__nombre"] == res:

                    bandOriginal = True

            if bandOriginal == True:

                print("Consigue")

            else:

                print("No consigue")
                agregar = tableOperacionCat()
                agregar.fecha = tod
                agregar.tabNombre = nombreTabla
                tableType = factCategory.objects.get(nombre=res)
                agregar.tabCat = tableType
                if request.POST.get("tabPrincipal"):
                    agregar.principal = True
                else:
                    agregar.principal = False
                agregar.suma = False
                agregar.resta = True
                agregar.tabTotal = 0
                agregar.save()

            bandOriginal = False

        bandRevisar == False
        lista2 = tableOperacionCat.objects.all().filter(tabNombre=val)

        for li in lista2:

            for sum in sumas:

                if li.tabCat.nombre == sum and li.suma == True:

                    bandRevisar = True

            if bandRevisar == False:

                eliminar = tableOperacionCat.objects.filter(tabNombre=val,tabCat__nombre=li.tabCat.nombre,suma=True)
                if eliminar:
                    eliminar.delete()

            else:

                print("Posee ese type en el form")

            bandRevisar = False

            for res in restas:

                if li.tabCat.nombre == res:

                    bandRevisar = True

            if bandRevisar == False:

                eliminar = tableOperacionCat.objects.filter(tabNombre=val,tabCat__nombre=li.tabCat.nombre,resta=True)
                if eliminar:
                    eliminar.delete()

            else:

                print("Posee ese type en el form")

            bandRevisar = False

        allTables = tableOperacionCat.objects.all().order_by("tabNombre")
        allTablesNombres = tableOperacionCat.objects.all().values("tabNombre","principal").distinct().order_by("tabNombre")

        dic = {"allTables":allTables,"allTablesNombres":allTablesNombres,"allTypes":allTypes}
        return render(request,"spareapp/contListCustomTablesOp.html",dic)

    dic = {"customNombre":customNombre,"customAux":customAux,"allTypes":allTypes,"val":val}

    return render(request,"spareapp/editeCustomTableCat.html",dic)

def probarRepetido(request):

    print("Entra en probarRepetido")
    repetido = False
    contFacType = request.POST.get("contFacType")
    print(contFacType)
    contFechaCreado = request.POST.get("contFechaCreado")
    print(contFechaCreado)
    contNombre = persona.objects.get(id=request.POST.get("contNombre"))
    print(contNombre)
    contNumFac = request.POST.get("contNumFac")
    print(contNumFac)

    prueba = factura.objects.filter(num=contNumFac)
    if prueba:
        repetido = True
    else:
        repetido = False

    contCatIng = factCategory.objects.get(id=request.POST.get("contCatIng"))
    print(contCatIng)
    contTypeIng = factType.objects.get(id=request.POST.get("contTypeIng"))
    print(contTypeIng)
    print(request.POST.get("contFechaTope"))
    contMonto = request.POST.get("contMonto")
    print(contMonto)
    contItbm = request.POST.get("contItbm")
    print(contItbm)
    contTotal = request.POST.get("contTotal")
    print(contTotal)
    contNota = request.POST.get("contNota")
    print(contNota)

    pruebaRep = repetido

    print(pruebaRep)

    return JsonResponse({'pruebaRep':pruebaRep})

def combinarUsuarios(request):

    lista = persona.objects.all().order_by("nombre").order_by("id","nombre")

    dic = {"lista":lista}

    if request.method == "POST":

        seleccion = request.POST.getlist("seleccion")
        principal = request.POST.getlist("principal")

        if len(principal)!=1:

            return render(request,"spareapp/combinarUsuarios.html",dic)

        principalUser = persona.objects.get(id=principal[0])

        if seleccion:

            for p in seleccion:

                auxUser = persona.objects.get(id=p)

                facAux = factura.objects.filter(refPersona=auxUser)

                if facAux:

                    for fac in facAux:

                        fac.refPersona = principalUser

                        fac.save()

                if auxUser != principalUser:

                    auxUser.delete()

    return render(request,"spareapp/combinarUsuarios.html",dic)

from django.core import management
from django.core.management.commands import loaddata, dumpdata

import sys
 
# from django.core.management import call_command



# django.core.management.call_command(name, *args, **options)

from django.http import JsonResponse
from django.http import HttpResponse

def respaldarDb(request):

    # file = open('database.txt', 'rb')
    # output = open('database.txt', 'w')
    # call_command('dumpdata', format='json', indent=3, stdout=output)
    # output.close()

    sysout = sys.stdout
    sys.stdout = open('Spareparts/fixture/respaldo.json', 'w')
    management.call_command('dumpdata','--format=json','--indent=4')
    sys.stdout = sysout

    with open("Spareparts/fixture/respaldo.json") as f:
        long_description = f.read()
        print(long_description)

        # return HttpResponse(json.dumps(long_description), content_type = "application/json")
        # response = HttpResponse(json.dumps(long_description), content_type = "application/json")
        response = HttpResponse(long_description, content_type = "application/json")
        response['Content-Disposition'] = 'attachment; filename=respaldo.json' 
        return response
        # return JsonResponse(long_description)

    # response = HttpResponse(FileWrapper(sys.stdout.getvalue()), content_type='application/zip') 
    # response = sys.stdout
    # response['Content-Disposition'] = 'attachment; filename=Spareparts/fixture/respaldo.json' 
    # return response


    # output = open('Spareparts/fixture/respaldo.json', 'r')
    # print(output)
    # output.close()

    # -- format=json --indent=4

    # management.call_command('dumpdata','--format=json','--indent=4')

    return render(request,"spareapp/contAdmin.html")

def cargarDb(request):

    management.call_command('loaddata', 'Spareparts/fixture/respaldo.json')

    return render(request,"spareapp/contAdmin.html")

import json

def contCargarDb(request):

    facAuxAllCat = ""

    if request.method == "POST":

        facturasDelete = factura.objects.all()
        factTypeDelete = factType.objects.all()
        factCategoryDelete = factCategory.objects.all()
        spareDelete = spare.objects.all()        

        if request.FILES.get("cargar"):

            for fac in facturasDelete:

                fac.delete()

            for fac in factTypeDelete:

                fac.delete()

            for fac in factCategoryDelete:

                fac.delete()

            for fac in spareDelete:

                fac.delete()

            archivo = request.FILES['cargar'].read()
            data = json.loads(archivo)
            s = json.dumps(data, indent=4, sort_keys=True)

            output = open('Spareparts/fixture/respaldoAux.json', 'w')
            output.write(s)
            output.close()

            management.call_command('loaddata', 'Spareparts/fixture/respaldoAux.json')

            file = 'respaldoAux.json'
            location = 'Spareparts/fixture'
            path = os.path.join(location, file)
            os.remove(path)

        else:

            return render(request,"spareapp/contCargarDb.html")

        # ----------- Operacion -------------------
        toddy = datetime.now().date()
        allTypesCustom = factType.objects.all()
        tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

        if tableAuxOp:

            print("Hay tabla")

            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total
                customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0

        else:

            print("No hay tabla")
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")


                for fac in facAuxAll:
                    custAcum = custAcum + fac.total
                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = toddy
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(id=ty.id)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

        # ----------- Categoria -------------------

        tod = datetime.now().date()
        acum = 0
        cantAuxCat = tableOperacionCat.objects.filter(fecha__date=tod).values("tabNombre").distinct()
        factureAuxCat = factura.objects.filter(fechaCreado__date=tod)
        allTypesCustom = factCategory.objects.all()
        totalParcialOpCat = {}
        custAcum = 0
        
        tableAuxCat = tableOperacionCat.objects.filter(fecha__date=tod)

        if factureAuxCat:

            print("Hay facturas")

            if tableAuxCat:

                print("Hay tabla")
                toddy = datetime.now().date()
                allTypesCustom = factCategory.objects.all()
                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAllCat = factura.objects.filter(fechaCreado__date=toddy,refCategory=ty)

                    for fac in facAuxAllCat:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:

                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:

                            prob2 = tableOperacionCat.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                            if prob2:

                                costomInd = tableOperacionCat.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                                costomInd.tabTotal = custAcum
                                costomInd.save()

                            else:

                                costomInd = tableOperacionCat()
                                costomInd.fecha = toddy
                                costomInd.tabNombre = nom["tabNombre"]
                                typeAux = factCategory.objects.get(nombre=ty)
                                costomInd.tabCat = typeAux
                                costomInd.principal = principalAux[0]["principal"]
                                if sumaAux[0]["suma"]==True:
                                    costomInd.suma = True
                                    costomInd.resta = False
                                else:
                                    costomInd.suma = False
                                    costomInd.resta = True
                                costomInd.tabTotal = custAcum
                                costomInd.save()
                        
                    custAcum = 0
            else:

                print("No hay tabla")

                custAcum = 0
                for ty in allTypesCustom:
                    facAuxAll = factura.objects.filter(fechaCreado__date=tod,refCategory=ty)

                    for fac in facAuxAll:
                        custAcum = custAcum + fac.total

                    lista = tableOperacionCat.objects.all().values("tabNombre","suma","resta").distinct()
                    for nom in lista:
                        prob = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                        principalAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                        sumaAux = tableOperacionCat.objects.filter(tabNombre=nom["tabNombre"],tabCat__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                        if prob:
                            costomInd = tableOperacionCat()
                            costomInd.fecha = tod
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factCategory.objects.get(nombre=ty)
                            costomInd.tabCat = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                    custAcum = 0

    return render(request,"spareapp/contCargarDb.html")

def checkearNc(request):

    allFactures = factura.objects.none()
    allPersonas = factura.objects.none()
    allCategorys = factura.objects.none()
    allFacturesQuery = None
    allPersonasQuery = None
    allCategorysQuery = None
    filterFactures = factura.objects.none()
    filterFacturesE = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterCategorys = factura.objects.none()
    filterCategorysE = factura.objects.none()
    filterFacturesDate = factura.objects.none()
    filterFacturesDateE = factura.objects.none()
    filterFacturesDatePersonas = factura.objects.none()
    filterFacturesDateCategories = factura.objects.none()
    allFacturesPay = None
    dateFrom = None
    dateTo = None
    dayFrom = None
    dayTo = None
    dayFromQuery = None
    dayToQuery = None

    # dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

    searchMetodo = "all"
    if request.GET.get("searchMetodo") == "all":
        searchMetodo = "all"
        allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
            dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())
    else:
        searchMetodo = "range"
        dateFrom = request.GET.get("dateFrom")
        dateTo = request.GET.get("dateTo")
        dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
        dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    palabrasErase = []
    palabrasErase2 = []

    checkeado = request.GET.get("checkeado")
    filter = request.GET.get("filter")

    auxInicio = -1
    auxInicioe = -1
    auxFin = -1
    acumCom = 0
    filterAux = filter
    filterAuxErase = filter
    palabras = []
    nuevoFilter = ""
    palabraFinal = ""

    for pos,let in enumerate(filter):

        if(let == '-'):

            filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
            # auxFine = pos
            palabraFinalErase = filterErase.split(" ")
            palabrasErase.append(palabraFinalErase[0])
            auxInicioe = -1

        else:

            auxInicioe = pos

    for val in palabrasErase:

        palabrasErase2.append("-"+val)

    for pos,let in enumerate(filter):

        if(let == '"'):

            acumCom = acumCom + 1

            if acumCom == 2:

                nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                auxFin = pos
                palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                palabras.append(palabraFinal)
                acumCom = 0
                auxInicio = -1
                auxFin = -1

            else:

                auxInicio = pos

    palabrasAux = []

    for val in palabras:

        if filterAux.find(val) >= 0:
            filterAux2 = filterAux.replace(val,"")
            filterAux = filterAux2
        palabrasAux.append(val.strip('"'))

    filterAux = filterAux.split(" ")
    filterAux = [item for item in filterAux if item]
    filter = filterAux + palabrasAux
    for val in palabrasErase2:
        filter.remove(val)

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterFacturesDatePersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterFacturesDateCategories = factura.objects.values("refCategory__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    for fil in filter:

        if filterFactures:

            filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )
            filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )
            filterCategorys = filterCategorys & ( factura.objects.values("refCategory__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        else:

            filterFactures = ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )
            filterPersonas = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )
            filterCategorys = ( factura.objects.values("refCategory__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id") )

    if checkeado == "true":

        checkeado = True
    
    else:

        checkeado = False

    if checkeado == True:

        allFactures = factura.objects.filter(nc=checkeado)
        allPersonas = factura.objects.values("refPersona__nombre").filter(nc=checkeado)
        allCategorys = factura.objects.values("refCategory__nombre").filter(nc=checkeado)

    else:

        allFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
        allPersonas = factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
        allCategorys = factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")

    if filter:

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterPersonas = factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterCategorys = factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterPersonasE = filterPersonasE & ( factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterCategorysE = filterCategorysE & ( factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )

        else:

            filterFacturesE = ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterPersonasE = ( factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterCategorysE = ( factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(num__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(note__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).exclude(refCategory__nombre__icontains=fil) )

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterPersonasE = factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")
        filterCategorysE = factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaCreado","id")

    # print("filterFactures")
    # print(filterFactures)

    if searchMetodo == "range":
        allFactures = allFactures & filterFactures & filterFacturesDate & filterFacturesE
        allPersonas = allPersonas & filterPersonas & filterFacturesDatePersonas & filterPersonasE
        allCategorys = allCategorys & filterCategorys & filterFacturesDateCategories & filterCategorysE
    else:
        allFactures = allFactures & filterFactures & filterFacturesE
        allPersonas = allPersonas & filterPersonas & filterPersonasE
        allCategorys = allCategorys & filterCategorys & filterCategorysE

    # print("allFactures")
    # print(allFactures)
    
    for fac in allFactures:

        deadline = datetime.now().date() - fac.fechaCreado.date()
        deadlineDic.append(deadline.days)
        dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))
    
    allFacturesQuery = list(allFactures.values())
    allPersonasQuery = list(allPersonas)
    allCategorysQuery = list(allCategorys)

    for fac in allFactures:
        
        if fac.nc == False:

            acum = acum + fac.monto
            acum2 = acum2 + fac.total
            acumIva = acumIva + fac.iva

        else:

            acum = acum - fac.monto
            acum2 = acum2 - fac.total
            acumIva = acumIva - fac.iva   

    # for fac in allFactures:

    #     acum = acum + fac.monto
    #     acum2 = acum2 + fac.total
    #     acumIva = acumIva + fac.iva

    return JsonResponse({"dayFromQuery":dayFromQuery,"dayToQuery":dayToQuery,"dayFrom":dayFrom,"dayTo":dayTo,"searchMetodo":searchMetodo,"dateFrom":dateFrom,"dateTo":dateTo,"filter":filter,'acumIva':acumIva,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2})

def deleteDb(request):

    facturasDelete = factura.objects.all()
    tableOperacionDelete = tableOperacion.objects.all()
    carDelete = car.objects.all()
    engineDelete = engine.objects.all()
    categoryDelete = category.objects.all()
    vendorDelete = vendor.objects.all()
    spareDelete = spare.objects.all()
    dimensionDelete = dimension.objects.all()
    atributeDelete = atribute.objects.all()
    referenceDelete = reference.objects.all()
    spareCartDelete = spareCart.objects.all()
    ProfileDelete = Profile.objects.all()
    personaDelete = persona.objects.all()
    factTypeDelete = factType.objects.all()
    factCategoryDelete = factCategory.objects.all()
    mainTableDelete = mainTable.objects.all()
    mainTableAuxDelete = mainTableAux.objects.all()
    customTableDelete = customTable.objects.all()
    customAuxDelete = customAux.objects.all()
    tableOperacionAuxDelete = tableOperacionAux.objects.all()
    tableOperacionCatDelete = tableOperacionCat.objects.all()
    tableOperacionAuxCatDelete = tableOperacionAuxCat.objects.all()

    for fac in tableOperacionAuxDelete:

        fac.delete()

    for fac in tableOperacionCatDelete:

        fac.delete()

    for fac in tableOperacionAuxCatDelete:

        fac.delete()

    for fac in mainTableAuxDelete:

        fac.delete()

    for fac in customTableDelete:

        fac.delete()

    for fac in customAuxDelete:

        fac.delete()

    for fac in factTypeDelete:

        fac.delete()

    for fac in factCategoryDelete:

        fac.delete()

    for fac in mainTableDelete:

        fac.delete()

    for fac in categoryDelete:

        fac.delete()

    for fac in vendorDelete:

        fac.delete()

    for fac in spareDelete:

        fac.delete()

    for fac in dimensionDelete:

        fac.delete()

    for fac in atributeDelete:

        fac.delete()

    for fac in referenceDelete:

        fac.delete()

    for fac in spareCartDelete:

        fac.delete()

    for fac in ProfileDelete:

        fac.delete()

    for fac in personaDelete:

        fac.delete()

    for fac in engineDelete:

        fac.delete()

    for fac in carDelete:

        fac.delete()

    for fac in facturasDelete:

        fac.delete()

    for op in tableOperacionDelete:

        op.delete()

    return render(request,"spareapp/contAdmin.html")

def deleteDbSpare(request):

    spareDelete = spare.objects.all()
    carDelete = car.objects.all()
    engineDelete = engine.objects.all()
    # referenceDelete = reference.objects.all()
    # spareCartDelete = spareCart.objects.all()

    for fac in spareDelete:

        fac.delete()

    for fac in carDelete:

        fac.delete()

    for fac in engineDelete:

        fac.delete()

    # for fac in referenceDelete:

    #     fac.delete()

    # for fac in spareCartDelete:

    #     fac.delete()

    return render(request,"spareapp/contAdmin.html")

def filterToCollect(request):

    allFacturesQuery = None
    allPersonasQuery = None
    allCategorysQuery = None
    filterFactures = factura.objects.none()
    filterFacturesE = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterPersonasE = factura.objects.none()
    filterCategorysE = factura.objects.none()
    filterCategorys = factura.objects.none()
    filterFacturesDate = factura.objects.none()
    filterFacturesDatePersonas = factura.objects.none()
    filterFacturesDateCategories = factura.objects.none()
    allFacturesPay = None
    dateFrom = None
    dateTo = None
    dayFrom = None
    dayTo = None
    dayFromQuery = None
    dayToQuery = None

    searchMetodo = "all"
    if request.GET.get("searchMetodo") == "all":
        searchMetodo = "all"
        allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
            dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())
    else:
        searchMetodo = "range"
        dateFrom = request.GET.get("dateFrom")
        dateTo = request.GET.get("dateTo")
        dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
        dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")

    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []

    filter = request.GET.get("filter")
    # filter = filter.split(" ")

    # -----------------------------------------------------------------
    auxInicio = -1
    auxFin = -1
    auxInicioe = -1
    auxFine = -1
    acumCom = 0
    filterAux = filter
    filterAuxErase = filter
    palabras = []
    palabrasErase = []
    nuevoFilter = ""
    filterErase = ""
    palabraFinal = ""
    palabraFinalErase = ""
    palabrasErase2 = []

    for pos,let in enumerate(filter):

        if(let == '-'):

            filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
            # auxFine = pos
            palabraFinalErase = filterErase.split(" ")
            palabrasErase.append(palabraFinalErase[0])
            auxInicioe = -1

        else:

            auxInicioe = pos

    for val in palabrasErase:

        palabrasErase2.append("-"+val)

    # print(palabrasErase2)

    for pos,let in enumerate(filter):

        if(let == '"'):

            acumCom = acumCom + 1

            if acumCom == 2:

                nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                auxFin = pos
                palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                palabras.append(palabraFinal)
                acumCom = 0
                auxInicio = -1
                auxFin = -1

            else:

                auxInicio = pos

        else:

            auxInicio = pos

    palabrasAux = []
    # print("Salí")
    # print("filterAux")
    # print(filterAux)

    for val in palabras:

        if filterAux.find(val) or filterAux == val:

            filterAux2 = filterAux.replace(val,"")
            filterAux = filterAux2
        palabrasAux.append(val.strip('"'))

    filterAux = filterAux.split(" ")
    filterAux = [item for item in filterAux if item]
    filter = filterAux + palabrasAux
    for val in palabrasErase2:
        filter.remove(val)
    # -----------------------------------------------------------------

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterFacturesDatePersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterFacturesDateCategories = factura.objects.values("refCategory__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    # print("Antes de filtrar")
    # print(filter)

    for fil in filter:

        if filterFactures:

            filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
            filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
            filterCategorys = filterCategorys & ( factura.objects.values("refCategory__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

        else:

            filterFactures = factura.objects.filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
            filterPersonas = factura.objects.values("refPersona__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
            filterCategorys = factura.objects.values("refCategory__nombre").filter(num__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refPersona__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(note__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refType__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refCategory__nombre__icontains=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterPersonasE = filterPersonasE & ( factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterCategorysE = filterCategorysE & ( factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )

        else:

            filterFacturesE = ( factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterPersonasE = ( factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )
            filterCategorysE = ( factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(num__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refPersona__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(note__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refType__nombre__icontains=fil) & factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).exclude(refCategory__nombre__icontains=fil) )

            # filterFacturesE = factura.objects.filter(num__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
            # filterPersonasE = factura.objects.values("refPersona__nombre").filter(num__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
            # filterCategorysE = factura.objects.values("refCategory__nombre").filter(num__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refPersona__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(note__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refType__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre").filter(refCategory__nombre__icontains!=fil,pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    if filter:

        pass

    else:

        filterFactures = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterPersonas = factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterCategorys = factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterPersonasE = factura.objects.values("refPersona__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")
        filterCategorysE = factura.objects.values("refCategory__nombre").filter(pendiente=True,refCategory__limite=True,refCategory__ingreso=True).order_by("fechaCreado","id")

    if searchMetodo == "range":
        filterFactures = filterFactures & filterFacturesDate & filterFacturesE
        filterPersonas = filterPersonas & filterFacturesDatePersonas & filterPersonasE
        filterCategorys = filterCategorys & filterFacturesDateCategories & filterCategorysE
    else:
        filterFactures = filterFactures & filterFacturesE
        filterPersonas = filterPersonas & filterPersonasE
        filterCategorys = filterCategorys & filterCategorysE

    allFacturesQuery = list(filterFactures.values())
    allPersonasQuery = list(filterPersonas)
    allCategorysQuery = list(filterCategorys)

    for fac in filterFactures:

        deadline = datetime.now().date() - fac.fechaCreado.date()
        deadlineDic.append(deadline.days)
        dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

        acum = acum + fac.monto
        acum2 = acum2 + fac.total
        acumIva = acumIva + fac.iva

    return JsonResponse({'dayFromQuery':dayFromQuery,'dayToQuery':dayToQuery,'dayFrom':dayFrom,"dayTo":dayTo,'acumIva':acumIva,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2})

from django.core import serializers

def filterContType(request):

    tod = datetime.now().date()

    allFactures = factura.objects.none()
    allPersonas = factura.objects.none()
    allCategorys = factura.objects.none()
    allFacturesQuery = None
    allPersonasQuery = None
    allCategorysQuery = None
    filterFactures = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterCategorys = factura.objects.none()
    filterFacturesE = factura.objects.none()
    filterCategorysE = factura.objects.none()
    filterFacturesDate = factura.objects.none()
    filterFacturesDatePersonas = factura.objects.none()
    filterFacturesDateCategories = factura.objects.none()
    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    auxFac = None
    auxPer = None
    auxCat = None
    dayFrom = None
    dayTo = None
    dateFrom = None
    dateTo = None

    # diaAux = datetime.strptime(str(request.GET.get("val2")),"%Y-%m-%d")
    print(request.POST)
    print(request.GET)
    val1 = request.GET.get("val1")
    val2 = request.GET.get("val2")

    if request.GET.get("dateTo"):

        searchMetodo = "range"
        dateFrom = request.GET.get("dateFrom")
        dateTo = request.GET.get("dateTo")
        dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
        dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
        creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
        diaAux = creadoAuxdateFrom
        creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
        dayFrom = dateFrom
        dayTo = dateTo
        dateFrom = str(creadoAuxdateFrom.date())
        dateTo = str(creadoAuxdateTo.date())

    else:

        searchMetodo = "all"
        diaAux = datetime.strptime(str(request.GET.get("val2")),"%Y-%m-%d")
        allFacturesPay = factura.objects.filter(fechaCreado__date=diaAux.date(),refType__nombre=val1).order_by("fechaTope")
        if allFacturesPay:
            dateFrom = allFacturesPay[0].fechaCreado.date()
            creadoAuxdateFrom = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dateTo = allFacturesPay[len(allFacturesPay)-1].fechaCreado.date()
            dayFromQuery = datetime.strptime(str(dateFrom),"%Y-%m-%d")
            dayFromQuery = dayFromQuery.date().strftime("%d de %B de %Y")
            dayToQuery = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayToQuery = dayToQuery.date().strftime("%d de %B de %Y")
            creadoAuxdateTo = datetime.strptime(str(dateTo),"%Y-%m-%d")
            dayFrom = dateFrom
            dayTo = dateTo
            dateFrom = str(creadoAuxdateFrom.date())
            dateTo = str(creadoAuxdateTo.date())

    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    palabrasErase = []
    palabrasErase2 = []

    filter = request.GET.get("filter")

    auxInicio = -1
    auxInicioe = -1
    auxFin = -1
    acumCom = 0
    filterAux = filter
    filterAuxErase = filter
    palabras = []
    nuevoFilter = ""
    palabraFinal = ""

    for pos,let in enumerate(filter):

        if(let == '-'):

            filterErase = filterAuxErase.replace(filterAuxErase[:auxInicioe+2],"")
            # auxFine = pos
            palabraFinalErase = filterErase.split(" ")
            palabrasErase.append(palabraFinalErase[0])
            auxInicioe = -1

        else:

            auxInicioe = pos

    for val in palabrasErase:

        palabrasErase2.append("-"+val)

    for pos,let in enumerate(filter):

        if(let == '"'):

            acumCom = acumCom + 1

            if acumCom == 2:

                nuevoFilter = filterAux.replace(filterAux[:auxInicio],"")
                auxFin = pos
                palabraFinal = nuevoFilter[:auxFin-auxInicio+1]
                palabras.append(palabraFinal)
                acumCom = 0
                auxInicio = -1
                auxFin = -1

            else:

                auxInicio = pos

    palabrasAux = []

    for val in palabras:

        if filterAux.find(val) >= 0:

            filterAux2 = filterAux.replace(val,"")
            filterAux = filterAux2
        palabrasAux.append(val.strip('"'))

    filterAux = filterAux.split(" ")
    filterAux = [item for item in filterAux if item]
    filter = filterAux + palabrasAux
    for val in palabrasErase2:
        filter.remove(val)
    
    tod = val2
    typeAux = factType.objects.get(nombre=val1)

    if searchMetodo == "range":
        filterFacturesDate = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
        filterFacturesDatePersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
        filterFacturesDateCategories = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

    for fil in filter:

        if dateTo:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )

            if typeAux.mercPagada == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                else:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

            if typeAux.mercPagar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        else:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )

            if typeAux.mercPagada == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)",refType__nombre__icontains=fil).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id") )

                else:

                    auxFac = ( factura.objects.filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )
                    auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada",refCategory__nombre__icontains=fil).order_by("fechaCreado","id") )

            if typeAux.facCobrar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id") )

            if typeAux.mercPagar == True:

                auxFac = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxPer = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )
                auxCat = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id") )

        if filter.index(fil) == 0:

            filterFactures = auxFac
            filterPersonas = auxPer
            filterCategorys = auxCat

        if filterFactures:

            filterFactures = filterFactures & auxFac
            filterPersonas = filterPersonas & auxPer
            filterCategorys = filterCategorys & auxCat

        else:

            filterFactures = auxFac
            filterPersonas = auxPer
            filterCategorys = auxCat

        # if filter.index(fil) == 0:

        #     filterFactures = auxFac
        #     filterPersonas = auxPer
        #     filterCategorys = auxCat

        # else:

        #     filterFactures = filterFactures & auxFac
        #     filterPersonas = filterPersonas & auxPer
        #     filterCategorys = filterCategorys & auxCat

    if filter:

        pass

    else:

        if dateTo:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id")

            if typeAux.mercPagada == True:

                filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")

                else:

                    filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado__gte=dateFrom,fechaCobrado__lte=dateTo,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

            if typeAux.facCobrar == True:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

            if typeAux.mercPagar == True:

                filterFactures = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")

        else:

            if typeAux.mercPagada == False and typeAux.facCobrada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

                filterFactures = factura.objects.filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id")

            if typeAux.mercPagada == True:

                filterFactures = factura.objects.filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__egreso=True,fechaCobrado=tod,refCategory__nombre="Mercancia credito pagada").order_by("fechaCreado","id")

            if typeAux.facCobrada == True:

                if typeAux.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":

                    filterFactures = factura.objects.filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=val2,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)").order_by("fechaCreado","id")
               
                else:

                    filterFactures = factura.objects.filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterPersonas = factura.objects.values("refPersona__nombre").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__ingreso=True,fechaCobrado=tod,refCategory__nombre="Factura cobrada").order_by("fechaCreado","id")

            if typeAux.facCobrar == True:

                filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__ingreso=True).order_by("fechaCreado","id")

            if typeAux.mercPagar == True:

                filterFactures = factura.objects.filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterPersonas = factura.objects.values("refPersona__nombre").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date=tod,pendiente=True,refCategory__egreso=True).order_by("fechaCreado","id")
            
    for fil in palabrasErase:

        if filterFacturesE:

            filterFacturesE = filterFacturesE & ( factura.objects.all().exclude(num__icontains=fil))
            filterPersonasE = filterPersonasE & ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
            filterCategorysE = filterCategorysE & ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

        else:

            filterFacturesE = ( factura.objects.all().exclude(num__icontains=fil))
            filterPersonasE = ( factura.objects.values("refPersona__nombre").all().exclude(num__icontains=fil))
            filterCategorysE = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").all().exclude(num__icontains=fil))

    if palabrasErase:

        pass

    else:

        filterFacturesE = factura.objects.all().order_by("fechaCreado","id")
        filterPersonasE = factura.objects.values("refPersona__nombre").all().order_by("fechaCreado","id")
        filterCategorysE = factura.objects.values("refCategory__nombre","refCategory__ingreso").all().order_by("fechaCreado","id")

    if searchMetodo == "range":
        allFactures = filterFactures & filterFacturesDate & filterFacturesE
        allPersonas = filterPersonas & filterFacturesDatePersonas & filterPersonasE
        allCategorys = filterCategorys & filterFacturesDateCategories & filterCategorysE
    else:
        allFactures = filterFactures & filterFacturesE
        allPersonas = filterPersonas & filterPersonasE
        allCategorys = filterCategorys & filterCategorysE

    allFacturesQuery = list(allFactures.values())
    allPersonasQuery = list(allPersonas)
    allCategorysQuery = list(allCategorys)

    if typeAux.facCobrada == False and typeAux.mercPagada == False and typeAux.facCobrar == False and typeAux.mercPagar == False:

        if typeAux.ingreso:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

        else:

            for fac in allFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                if fac.nc == True:

                    acum = acum + fac.monto
                    acumIva = acumIva + fac.iva
                    acum2 = acum2 + fac.total

                else:

                    acum = acum - fac.monto
                    acumIva = acumIva - fac.iva
                    acum2 = acum2 - fac.total

    if typeAux.facCobrada == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum + fac.monto
            acumIva = acumIva + fac.iva
            acum2 = acum2 + fac.total

    if typeAux.mercPagada == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            if fac.nc == True:

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

            else:

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

    if typeAux.facCobrar == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum - fac.monto
            acumIva = acumIva - fac.iva
            acum2 = acum2 - fac.total

    if typeAux.mercPagar == True:

        for fac in allFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            if fac.nc == True:

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

            else:

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

    typeSearch = {}
    typeSearch["nombre"] = typeAux.nombre
    typeSearch["include"] = typeAux.include
    typeSearch["manual"] = typeAux.manual
    typeSearch["mercPagada"] = typeAux.mercPagada
    typeSearch["mercPagar"] = typeAux.mercPagar
    typeSearch["facCobrar"] = typeAux.facCobrar
    typeSearch["facCobrada"] = typeAux.facCobrada
    typeSearch["visa"] = typeAux.visa
    typeSearch["clave"] = typeAux.clave
    typeSearch["ingreso"] = typeAux.ingreso
    typeSearch["gasto"] = typeAux.gasto

    return JsonResponse({"filter":filter,"typeSearch":typeSearch,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2,'val':val1,'val2':val2,"acumIva":acumIva})

def filterContTypeTarjeta(request):

    allFacturesQuery = None
    allPersonasQuery = None
    allCategorysQuery = None
    filterFactures = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterCategorys = factura.objects.none()
    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []
    interes = []
    retencion = []
    neto = []
    interesTotal = 0
    retencionTotal = 0
    netoTotal = 0

    filter = request.GET.get("filter")
    filter = filter.split(" ")
    val1 = request.GET.get("val1")
    val2 = request.GET.get("val2")
    dateTo = request.GET.get("dateTo")
    dateFrom = request.GET.get("dateFrom")
    tod = val2
    typeAux = factType.objects.get(nombre=val1)

    for fil in filter:

        if filterFactures:

            if dateTo:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = filterCategorys & ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )

            else:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = filterCategorys & ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )

        else:

            if dateTo:

                filterFactures = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refType__nombre=val1).order_by("fechaCreado","id") )

            else:

                filterFactures = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = ( factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refType__nombre=val1).order_by("fechaCreado","id") )

    allFacturesQuery = list(filterFactures.values())
    allPersonasQuery = list(filterPersonas)
    allCategorysQuery = list(filterCategorys)

    # for fac in allFacturesVal:

    #     itbmMonto = float(fac.iva)

    #     if typeAux.visa == True:

    #         neto = float((fac.total)-(float(fac.total)*0.0225*1.07)-(float(itbmMonto)/2))
    #         interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
    #         retencionTotal = retencionTotal + float(itbmMonto/2)
    #         netoTotal = netoTotal + neto
    #         itbm7[fac.id] = [float(itbmMonto),float(float(fac.total)*0.0225*1.07),float(float(itbmMonto)/2),float(neto)]

    #     else:

    #         neto = float((fac.total)-(float(fac.total)*0.02*1.07)-(float(itbmMonto)/2))
    #         interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)
    #         retencionTotal = retencionTotal + float(itbmMonto/2)
    #         netoTotal = netoTotal + neto
    #         itbm7[fac.id] = [float(fac.monto)*0.07,float(float(fac.total)*0.02*1.07),float(float(itbmMonto)/2),float(neto)]

    # for fac in allFacturesVal:

    #     montoTotal = montoTotal + fac.monto
    #     itbmTotal = itbmTotal + float(fac.iva)
    #     totalTotal = totalTotal + fac.total

    for fac in filterFactures:

        deadline = datetime.now().date() - fac.fechaCreado.date()
        deadlineDic.append(deadline.days)
        dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

        acum = acum + fac.monto
        acumIva = acumIva + fac.iva
        acum2 = acum2 + fac.total

        retencion.append(float(fac.iva/2))
        retencionTotal = retencionTotal + float(fac.iva/2)

        if fac.refType.visa == True:

            interes.append(float(fac.total)*0.0225*1.07)
            neto.append(fac.total - (float(fac.total)*0.0225*1.07) - (float(fac.iva/2)))
            interesTotal = interesTotal + float(float(fac.total)*0.0225*1.07)
            netoTotal = netoTotal + float(fac.total - (float(fac.total)*0.0225*1.07) - (float(fac.iva/2)))

        else:

            interes.append(float(fac.total)*0.02*1.07)
            neto.append(fac.total - (float(fac.total)*0.02*1.07) - (float(fac.iva/2)))
            interesTotal = interesTotal + float(float(fac.total)*0.02*1.07)
            netoTotal = netoTotal + float(fac.total - (float(fac.total)*0.02*1.07) - (float(fac.iva/2)))
        
    return JsonResponse({"netoTotal":netoTotal,"retencionTotal":retencionTotal,"interesTotal":interesTotal,'neto':neto,'retencion':retencion,'interes':interes,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2,'val':val1,'val2':val2,"acumIva":acumIva})

def filterContTypeCat(request):

    allFacturesQuery = None
    allPersonasQuery = None
    allCategorysQuery = None
    filterFactures = factura.objects.none()
    filterPersonas = factura.objects.none()
    filterCategorys = factura.objects.none()
    acum = 0
    acum2 = 0
    acumIva = 0
    deadline = ""
    deadlineDic = []
    dateDic = []

    filter = request.GET.get("filter")
    filter = filter.split(" ")
    val1 = request.GET.get("val1")
    val2 = request.GET.get("val2")
    dateTo = request.GET.get("dateTo")
    dateFrom = request.GET.get("dateFrom")
    tod = datetime.now().date()
    tod = val2
    # typeAux = factType.objects.get(nombre=val1)

    # allFacturesVal = factura.objects.filter(fechaCreado__date=tod,refCategory__nombre=val).exclude(pendiente=False,refType__facCobrar=True).exclude(pendiente=False,refType__facCobrar=True)

    for fil in filter:

        if filterFactures:

            if dateTo:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = filterCategorys & ( factura.objects.values("refType__nombre","refType__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )

            else:

                filterFactures = filterFactures & ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = filterPersonas & ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = filterCategorys & ( factura.objects.values("refType__nombre","refType__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )

        else:

            if dateTo:

                filterFactures = ( factura.objects.filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = ( factura.objects.values("refType__nombre","refType__ingreso").filter(num__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(note__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo,refCategory__nombre=val1).order_by("fechaCreado","id") )

            else:

                filterFactures = ( factura.objects.filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterPersonas = ( factura.objects.values("refPersona__nombre").filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )
                filterCategorys = ( factura.objects.values("refType__nombre","refType__ingreso").filter(num__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refPersona__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(note__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refType__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") | factura.objects.values("refType__nombre","refType__ingreso").filter(refCategory__nombre__icontains=fil,fechaCreado__date=val2,refCategory__nombre=val1).order_by("fechaCreado","id") )

    allFacturesQuery = list(filterFactures.values())
    allPersonasQuery = list(filterPersonas)
    allCategorysQuery = list(filterCategorys)
    
    typeAux = factType.objects.filter(id=filterFactures[0].refType.id)

    if typeAux[0].facCobrada == False and typeAux[0].mercPagada == False and typeAux[0].facCobrar == False and typeAux[0].mercPagar == False:

        if typeAux[0].ingreso:

            for fac in filterFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum + fac.monto
                acumIva = acumIva + fac.iva
                acum2 = acum2 + fac.total

        else:

            for fac in filterFactures:

                deadline = datetime.now().date() - fac.fechaCreado.date()
                deadlineDic.append(deadline.days)
                dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

                acum = acum - fac.monto
                acumIva = acumIva - fac.iva
                acum2 = acum2 - fac.total

    if typeAux[0].facCobrada == True:

        for fac in filterFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum + fac.monto
            acumIva = acumIva + fac.iva
            acum2 = acum2 + fac.total

    if typeAux[0].mercPagada == True:

        for fac in filterFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum - fac.monto
            acumIva = acumIva - fac.iva
            acum2 = acum2 - fac.total

    if typeAux[0].facCobrar == True:

        for fac in filterFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum - fac.monto
            acumIva = acumIva - fac.iva
            acum2 = acum2 - fac.total

    if typeAux[0].mercPagar == True:

        for fac in filterFactures:

            deadline = datetime.now().date() - fac.fechaCreado.date()
            deadlineDic.append(deadline.days)
            dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

            acum = acum + fac.monto
            acumIva = acumIva + fac.iva
            acum2 = acum2 + fac.total

    typeSearch = {}
    typeSearch["nombre"] = typeAux[0].nombre
    typeSearch["include"] = typeAux[0].include
    typeSearch["manual"] = typeAux[0].manual
    typeSearch["mercPagada"] = typeAux[0].mercPagada
    typeSearch["mercPagar"] = typeAux[0].mercPagar
    typeSearch["facCobrar"] = typeAux[0].facCobrar
    typeSearch["facCobrada"] = typeAux[0].facCobrada
    typeSearch["visa"] = typeAux[0].visa
    typeSearch["clave"] = typeAux[0].clave
    typeSearch["ingreso"] = typeAux[0].ingreso
    typeSearch["gasto"] = typeAux[0].gasto

    return JsonResponse({"typeSearch":typeSearch,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2,'val':val1,'val2':val2,"acumIva":acumIva})

def totalTablasType(request):

    print("Entra")

    diccionario = {}
    # diccionarioTipo = {}
    vector = []
    dayFrom = ""
    dayTo = ""
    acum = 0
    total = 0
    totalFinal = 0
    tod = datetime.now().date()
    searchDateFrom = None
    searchDateTo = None

    # actualAux=datetime.now().date()
    # actualDay=str(actualAux.year)+"-"+str('%02d' % actualAux.month)+"-"+str('%02d' % actualAux.day)
    # deadlineDefault=(datetime.now()+timedelta(days=30)).date()
    # actual=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)
    # print((datetime.now()-timedelta(days=30)).date())

    toddy = tod

    tipoBusqueda = "today"

    if request.method == "POST":

        if request.POST.get("search") == "all":

            tipoBusqueda = "all"
            toddy = (datetime.now()-timedelta(days=30)).date()

        if request.POST.get("search") == "range":

            tipoBusqueda = "range"
            searchDateFrom = request.POST.get("searchDateFrom")
            searchDateTo = request.POST.get("searchDateTo")

    allTablesOpOnlyNames = tableOperacion.objects.values("tabNombre","principal").all().order_by("tabNombre").distinct()
    if searchDateFrom:
        fechas = factura.objects.values("fechaCreado").filter(fechaCreado__date__gte=searchDateFrom,fechaCreado__date__lte=searchDateTo).order_by("fechaCreado").distinct()
    else:
        fechas = factura.objects.values("fechaCreado").filter(fechaCreado__date__gte=toddy).order_by("fechaCreado").distinct()

    vector = []
    vectorSup = []

    if searchDateFrom:
        factureName = factura.objects.filter(fechaCreado__date__gte=searchDateFrom,fechaCreado__date__lte=searchDateTo).order_by("fechaCreado","id")
    else:
        factureName = factura.objects.filter(fechaCreado__date__gte=toddy).order_by("fechaCreado","id")
    if factureName:
        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    # ----------- Operacion -------------------
    cantAuxOp = None
    totalParcialOp = {}
    facAuxAll = None
    lista = None
    allTypesCustom = factType.objects.all()

    for fec in fechas:

        toddy = fec["fechaCreado"].date()

        acum = 0
        cantAuxOp = tableOperacion.objects.filter(fecha__date=toddy).values("tabNombre","principal").order_by("tabNombre").distinct()
        totalParcialOp = {}
        tableAuxOp = tableOperacion.objects.filter(fecha__date=toddy)

        if tableAuxOp:

            print("Hay tabla")

            allTypesCustom = factType.objects.all()
            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")

                for fac in facAuxAll:
                    custAcum = custAcum + fac.total
                # customType = tableOperacion.objects.filter(fecha__date=toddy,tabTipo=ty)

                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:

                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:

                        prob2 = tableOperacion.objects.filter(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])

                        if prob2:

                            costomInd = tableOperacion.objects.get(fecha__date=toddy,tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                            costomInd.tabTotal = custAcum
                            costomInd.save()

                        else:

                            costomInd = tableOperacion()
                            costomInd.fecha = toddy
                            costomInd.tabNombre = nom["tabNombre"]
                            typeAux = factType.objects.get(nombre=ty)
                            costomInd.tabTipo = typeAux
                            costomInd.principal = principalAux[0]["principal"]
                            if sumaAux[0]["suma"]==True:
                                costomInd.suma = True
                                costomInd.resta = False
                            else:
                                costomInd.suma = False
                                costomInd.resta = True
                            costomInd.tabTotal = custAcum
                            costomInd.save()
                    
                custAcum = 0

        else:

            print("No hay tabla")

            custAcum = 0
            for ty in allTypesCustom:
                facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty)

                if ty.facCobrar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,refType=ty,pendiente=True,refCategory__ingreso=True,refCategory__limite=True)

                if ty.mercPagar == True:
                    facAuxAll = factura.objects.filter(fechaCreado__date=toddy,pendiente=True,refType=ty,refCategory__egreso=True,refCategory__limite=True)

                if ty.mercPagada == True:
                    facAuxAll = factura.objects.filter(fechaCobrado=toddy,pendiente=False,refCategory__egreso=True,refCategory__limite=True)

                if ty.facCobrada == True:

                    if ty.nombre == "FACTURA CREDITO COBRADA (MAYORISTA)":
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")
                    else:
                        facAuxAll = factura.objects.filter(fechaCreado=toddy,pendiente=False,refCategory__ingreso=True,refType__facCobrada=True).exclude(refType__nombre = "FACTURA CREDITO COBRADA (MAYORISTA)")


                for fac in facAuxAll:
                    custAcum = custAcum + fac.total
                lista = tableOperacion.objects.all().values("tabNombre","suma","resta").distinct()
                for nom in lista:
                    prob = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"])
                    principalAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("principal").distinct()
                    sumaAux = tableOperacion.objects.filter(tabNombre=nom["tabNombre"],tabTipo__nombre=ty,suma=nom["suma"],resta=nom["resta"]).values("suma").distinct()
                    if prob:
                        costomInd = tableOperacion()
                        costomInd.fecha = toddy
                        costomInd.tabNombre = nom["tabNombre"]
                        typeAux = factType.objects.get(nombre=ty)
                        costomInd.tabTipo = typeAux
                        costomInd.principal = principalAux[0]["principal"]
                        if sumaAux[0]["suma"]==True:
                            costomInd.suma = True
                            costomInd.resta = False
                        else:
                            costomInd.suma = False
                            costomInd.resta = True
                        costomInd.tabTotal = custAcum
                        costomInd.save()
                
                custAcum = 0

    # ------------------------------------------------------------------

    for date in fechas:

        diccionario[date["fechaCreado"]] = {}

        for tables in allTablesOpOnlyNames:
            
            vector.append(tables["tabNombre"])
            acum = 0
            tableaux = tableOperacion.objects.filter(tabNombre=tables["tabNombre"],fecha__date=date["fechaCreado"].date())

            for a in tableaux:

                if a.suma == True:

                    acum = acum + a.tabTotal

                else:

                    acum = acum - a.tabTotal

            total = total + acum
            totalFinal = totalFinal + total
            vector.append(total)
            vectorSup.append(vector)
            vector = []
            total = 0

        diccionario[date["fechaCreado"]] = vectorSup
        vectorSup = []

    allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
    allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
    facturesToCollect = len(allFacturesToCollect)
    facturesToPay = len(allFacturesToPay)

    dic = {"totalFinal":totalFinal,"tipoBusqueda":tipoBusqueda,"total":total,"dayTo":dayTo,"dayFrom":dayFrom,"diccionario":diccionario,"fechas":fechas,"tod":tod,"facturesToPay":facturesToPay,"facturesToCollect":facturesToCollect}

    return render(request,"spareapp/totalTablasType.html",dic)



def filterAccountStat(request):

    filter = request.GET.get("filter")
    filter = filter.split(" ")
    dateTo = request.GET.get("dateTo")
    dateFrom = request.GET.get("dateFrom")
    filter = [item for item in filter if item]
    all = request.GET.get("all")
    month = request.GET.get("month")
    range = request.GET.get("range")
    balanceAux = request.GET.get("balance")
    soloFac = request.GET.get("soloFac")
    pagCob = None
    auxNombre = request.GET.get("nombre")
    busquedaValor = request.GET.get("busquedaValor")
    pos = -1

    if busquedaValor:
        print("Hay busqueda")
    else:
        print("No hay busqueda")

    factureName = None
    filterTypes = None
    filterCategorys = None
    acumTotal = 0

    deadline = ""
    deadlineDic = []
    dateDic = []
    cont = 0
    balanceTotal = 0
    balanceFacMerc = 0
    balance = {}

    if soloFac == "true":
        pagCob = True
    else:
        pagCob = False

    if auxNombre:

        factureNameAux = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

        for fac in factureNameAux:

            if fac.refCategory.ingreso:

                cont = cont

                if fac.refType.facCobrar==True:

                    cont = cont + fac.total
                    if fac.pendiente == True:
                        balanceFacMerc = balanceFacMerc + fac.total
                
                if fac.refCategory.nombre=="Factura cobrada":

                    cont = cont - fac.total
            
            else:

                cont = cont

                if fac.refType.mercPagar==True:

                    if fac.nc == True:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total
                    if fac.pendiente == True:
                        balanceFacMerc = balanceFacMerc - fac.total
                
                if fac.refCategory.nombre=="Mercancia credito pagada":

                    if fac.nc == False:
                        cont = cont + fac.total
                    else:
                        cont = cont - fac.total

                    # cont = cont + fac.total

            balance[fac.id] = [cont,fac.total]
            acumTotal = cont

    balanceTotal = cont

    for fil in filter:

        if all:

            if filter.index(fil) == 0:
                
                if pagCob:
                    if busquedaValor:
                        factureName = factura.objects.filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                else:
                    if busquedaValor:
                        factureName = factura.objects.filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                    else:
                        factureName = factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
            else:
                if pagCob:
                    if busquedaValor:
                        factureName = factureName & factura.objects.filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factureName & factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                else:
                    if busquedaValor:
                        factureName = factureName & factura.objects.filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                    else:
                        factureName = factureName & factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
        if month:

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()

            mes = datetime.now().date().month
            dayFrom = dateFrom
            dayTo = datetime.now().date()
            anio = datetime.now().date().year

            if filter.index(fil) == 0:

                if pagCob:
                    if busquedaValor:
                        factureName = factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    else:
                        factureName = factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                else:
                    if busquedaValor:
                        factureName = factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(num__icontains=busquedaValor),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                    else:
                        factureName = factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
            else:
                if pagCob:
                    factureName = factureName & factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                else:
                    factureName = factureName & factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                    filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                    filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")

        if range:

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:

                if filter.index(fil) == 0:

                    if pagCob:
                        factureName = factura.objects.filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    else:                
                        factureName = factura.objects.filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                else:
                    if pagCob:
                        factureName = factureName & factura.objects.filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    else:                
                        factureName = factureName & factura.objects.filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado")
        
        if balanceAux:

            for key in balance:

                if balance[key][0]==0:
                    pos = key
            
            facActAux = factura.objects.filter(id=pos)

            if facActAux:
            
                facAct = factura.objects.get(id=pos)
                factureNameSome = factura.objects.filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")

                if filter.index(fil) == 0:
                    if pagCob:
                        factureName = factura.objects.filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factura.objects.filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                else:
                    if pagCob:
                        factureName = factureName & factura.objects.filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factureName & factura.objects.filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
            
            else:

                factureName = None
                filterTypes = None
                filterCategorys = None

            if factureNameSome:
                pass
            else:
                if filter.index(fil) == 0:
                    if pagCob:
                        factureName = factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                else:
                    if pagCob:
                        factureName = factureName & factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    else:
                        factureName = factureName & factura.objects.filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterTypes = filterTypes & factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")
                        filterCategorys = filterCategorys & factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),Q(refPersona__nombre__icontains=fil) | Q(num__icontains=fil) | Q(refType__nombre__icontains=fil) | Q(refCategory__nombre__icontains=fil) | Q(note__icontains=fil)).order_by("fechaCreado","id")

    if filter:
        pass
    else:
        factureName = None
        filterTypes = None
        filterCategorys = None
        if all:
            if pagCob:
                factureName = factura.objects.filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
            else:
                factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
                filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
        if month:

            mes = datetime.now().date().month
            date_today = datetime.now()
            dateFrom = date_today.replace(month=mes,day=1, hour=0, minute=0, second=0, microsecond=0)
            dateFrom = dateFrom.date()

            mes = datetime.now().date().month
            dayFrom = dateFrom
            dayTo = datetime.now().date()
            anio = datetime.now().date().year

            if pagCob:
                factureName = factura.objects.filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__month=mes),Q(fechaCreado__year=anio),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
            else:
                factureName = factura.objects.filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")
                filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")
                filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__month=mes,fechaCreado__year=anio,refPersona__id=auxNombre).order_by("fechaCreado")

        if range:

            fecha_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            fecha_to = datetime.strptime(dateTo, '%Y-%m-%d')

            dayFrom = fecha_from.date()
            dayTo = fecha_to.date()

            if dateFrom and dateTo:

                if pagCob:
                    factureName = factura.objects.filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(fechaCreado__date__gte=dayFrom),Q(fechaCreado__date__lte=dayTo),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado")
                else:
                    factureName = factura.objects.filter(fechaCreado__date__gte=dayFrom,fechaCreado__date__lte=dayTo,refPersona__id=auxNombre).order_by("fechaCreado")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(fechaCreado__date__gte=dayFrom,fechaCreado__date__lte=dayTo,refPersona__id=auxNombre).order_by("fechaCreado")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(fechaCreado__date__gte=dayFrom,fechaCreado__date__lte=dayTo,refPersona__id=auxNombre).order_by("fechaCreado")
        
        if balanceAux:
            for key in balance:

                if balance[key][0]==0:
                    pos = key
            
            facActAux = factura.objects.filter(id=pos)

            if facActAux:

                if pagCob:
                    facAct = factura.objects.get(id=pos)
                    factureName = factura.objects.filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(id__gte=facAct.id),Q(fechaCreado__gte=facAct.fechaCreado),Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                else:            
                    facAct = factura.objects.get(id=pos)
                    factureName = factura.objects.filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(id__gte=facAct.id,fechaCreado__gte=facAct.fechaCreado,refPersona__id=auxNombre).order_by("fechaCreado","id")
            else:

                factureName = None
                filterTypes = None
                filterCategorys = None

            if factureName:
                pass
            else:
                if pagCob:
                    factureName = factura.objects.filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(Q(refPersona__id=auxNombre),(Q(refType__mercPagar=True) | Q(refType__facCobrar=True)) & Q(pendiente=True)).order_by("fechaCreado","id")
                else:
                    factureName = factura.objects.filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
                    filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(refPersona__id=auxNombre).order_by("fechaCreado","id")
                    filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__id=auxNombre).order_by("fechaCreado","id")

    # # BUSQUEDA ---------------------------------------------------------

    dayFrom = ""
    dayTo = ""

    if busquedaValor:

        tod = datetime.now().date()
        factureName = ""
        personaVarios = None
        searchMetodo = "all"
                
        factureName = factura.objects.filter(num__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.filter(refPersona__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.filter(note__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.filter(refType__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.filter(refCategory__nombre__icontains=busquedaValor).order_by("fechaCreado","id")
        filterTypes = factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(num__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(refPersona__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(note__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(refType__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refPersona__nombre","refType__mercPagada","refType__mercPagar","refType__facCobrar","refType__facCobrada","refType__nombre","refType__ingreso").filter(refCategory__nombre__icontains=busquedaValor).order_by("fechaCreado","id")
        filterCategorys = factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(num__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refPersona__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(note__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refType__nombre__icontains=busquedaValor).order_by("fechaCreado","id") | factura.objects.values("refCategory__nombre","refCategory__ingreso").filter(refCategory__nombre__icontains=busquedaValor).order_by("fechaCreado","id")


    cont = 0

    # for facT in factureName:

    #     if facT.refType.ingreso == True and facT.refType.facCobrar == False or facT.refType.mercPagar == True:

    #         acumTotal = acumTotal + abs(facT.total)
        
    #     else:

    #         acumTotal = acumTotal - abs(facT.total)

    acumTotal = 0
            
    for fac in factureName:

        deadline = datetime.now().date() - fac.fechaCreado.date()
        deadlineDic.append(deadline.days)
        dateDic.append(fac.fechaCreado.date().strftime("%b %d, %Y"))

        if fac.refCategory.ingreso:

            # print("Ingreso")

            cont = cont

            if fac.refType.facCobrar==True:

                cont = cont + fac.total
                if fac.pendiente == True:
                    balanceFacMerc = balanceFacMerc + fac.total

            if fac.refCategory.nombre=="Factura cobrada" or fac.refCategory.nombre=="Factura cobrada (Mayorista)":

                cont = cont - fac.total
        
        else:
            # print("Egreso")

            cont = cont

            if fac.refType.mercPagar==True:

                if fac.nc == True:
                    cont = cont + fac.total
                else:
                    cont = cont - fac.total

                # cont = cont - fac.total
                if fac.pendiente == True:
                    balanceFacMerc = balanceFacMerc - fac.total
            
            if fac.refCategory.nombre=="Mercancia credito pagada":

                if fac.nc == False:
                    cont = cont + fac.total
                else:
                    cont = cont - fac.total

                # cont = cont + fac.total
        if fac.refType.ingreso == True and fac.refType.facCobrar == False or fac.refType.mercPagar == True:

            acumTotal = acumTotal + abs(fac.total)
        
        else:

            acumTotal = acumTotal - abs(fac.total)

        balance[fac.id] = [cont,fac.total]
        # acumTotal = cont

    print("Valor de acumtotal: "+str(acumTotal))
    print("Valor de cont: "+str(cont))

    if factureName:

        dayFrom = factureName[0].fechaCreado.date()
        dayTo = factureName[len(factureName)-1].fechaCreado.date()

    # acumTotal = 0

    # for facT in factureName:

    #     if facT.refType.ingreso == True:
    #         if (facT.refType.facCobrar == False):

    #             acumTotal = acumTotal + abs(facT.total)

    #         else:

    #             acumTotal = acumTotal - abs(facT.total)        
        
    #     else:

    #         if facT.nc == False:

    #             if facT.refType.mercPagar == True:

    #                 acumTotal = acumTotal + abs(facT.total)

    #             else:

    #                 acumTotal = acumTotal - abs(facT.total)

    #         else:

    #             if facT.refType.mercPagar == True:

    #                 acumTotal = acumTotal - abs(facT.total)

    #             else:

    #                 acumTotal = acumTotal + abs(facT.total)

    allFacturesQuery = list(factureName.values())
    # allPersonasQuery = list(filterPersonas)
    allCategorysQuery = list(filterCategorys)
    allTypesQuery = list(filterTypes)

    # print(acumTotal)

    return JsonResponse({"acumTotal":acumTotal,"cont":cont,"dateDic":dateDic,"balance":balance,"allTypesQuery":allTypesQuery,"allCategorysQuery":allCategorysQuery,"allFacturesQuery":allFacturesQuery,"all":all,"dateTo":dateTo,"dateFrom":dateFrom})
    # return JsonResponse({"typeSearch":typeSearch,'dateDic':dateDic,'deadlineDic':deadlineDic,'allFacturesQuery':allFacturesQuery,'allPersonasQuery':allPersonasQuery,'allCategorysQuery':allCategorysQuery,"acum":acum,"acum2":acum2,'val':val1,'val2':val2,"acumIva":acumIva})


# {% for fact in factureName %}
#         {% for key,value in balance.items %}
#         {% if key == fact.id %}
#         <tr>
#             <td {% if fact.refType.ingreso == True %}  {% if fact.nc == True %} {% if fact.refType.mercPagar == False %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == False or fact.refType.mercPagar == True %} style="color: #1029b9;" {% else %} style="color: #a00a0a;" {% endif %} {% endif %} {% else %} {% if fact.nc == True %} {% if fact.refType.mercPagar == True %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == True or fact.refType.mercPagar == False %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% endif %} {% endif %} class="p-2">{{fact.fechaCreado|date:"M d, Y"}}</td>
#             <td {% if fact.refType.ingreso == True %}  {% if fact.nc == True %} {% if fact.refType.mercPagar == False %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == False or fact.refType.mercPagar == True %} style="color: #1029b9;" {% else %} style="color: #a00a0a;" {% endif %} {% endif %} {% else %} {% if fact.nc == True %} {% if fact.refType.mercPagar == True %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == True or fact.refType.mercPagar == False %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% endif %} {% endif %} class="p-2">{{fact.refPersona.nombre}}</td>
#             <td {% if fact.refType.ingreso == True %}  {% if fact.nc == True %} {% if fact.refType.mercPagar == False %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == False or fact.refType.mercPagar == True %} style="color: #1029b9;" {% else %} style="color: #a00a0a;" {% endif %} {% endif %} {% else %} {% if fact.nc == True %} {% if fact.refType.mercPagar == True %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% else %} {% if fact.refType.facCobrar == True or fact.refType.mercPagar == False %} style="color: #a00a0a;" {% else %} style="color: #1029b9;" {% endif %} {% endif %} {% endif %} id="auxMonto" class="p-2"><div {% if fact.pendiente == True %} style="font-weight: bold;" {% else %} "" {% endif %}>${{value.1|stringformat:".2f"}}</div></td>
#         </tr>
#         {% endif %}
#         {% endfor %}
#         {% endfor %}


# tod = datetime.now().date()

#     allFacturesPay = factura.objects.filter(pendiente=True,refCategory__limite=True,refCategory__egreso=True).order_by("fechaTope")
#     allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(facCobrar=True).exclude(mercPagada=True).exclude(mercPagar=True)

#     deadlineDic = {}

#     for all in allFacturesPay:

#         deadline = datetime.now().date() - all.fechaCreado.date()
#         deadline = deadline.days
#         deadlineDic[all.id] = deadline

#     acum = 0
#     acum2 = 0

#     for fac in allFacturesPay:

#         acum = acum + fac.monto
#         acum2 = acum2 + fac.total

#     allFacturesToPay = factura.objects.filter(pendiente=True,refCategory__ingreso=True,refCategory__limite=True)
#     allFacturesToCollect = factura.objects.filter(pendiente=True,refCategory__egreso=True,refCategory__limite=True)
#     facturesToCollect = len(allFacturesToPay)
#     facturesToPay = len(allFacturesToCollect)



    # cobrarPagar = None
    # cate = None

    # if request.GET.get("fecha") == "change":

    #     creado = request.GET.get("creado")
    #     creadoAux = datetime.strptime(creado,"%Y-%m-%d")
    #     deadlineDefault=(creadoAux+timedelta(days=30)).date()
    #     actualAux=str(deadlineDefault.year)+"-"+str('%02d' % deadlineDefault.month)+"-"+str('%02d' % deadlineDefault.day)

    #     actual = actualAux

    #     return JsonResponse({'actual':actual})







    # cateAux = factCategory.objects.filter(nombre=request.GET.get("cat"))

    # if cateAux:
    #     cate = factCategory.objects.get(nombre=request.GET.get("cat"))

    # if request.GET.get("val") == "entry":

    #     if cateAux and (request.GET.get("cat") == "Factura cobrada" or request.GET.get("cat") == "Factura cobrada (Mayorista)"):
    #         # print("Existe factura y es factura cobrada")
    #         allCategories = factCategory.objects.filter(ingreso=True).order_by("nombre")
    #     else:
    #         # print("No existe factura o no es factura cobrada")
    #         allCategories = factCategory.objects.filter(ingreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada").exclude(nombre="Factura cobrada (Mayorista)")
    #     if cateAux and cate.limite == True:
    #         # print("Existe factura y la categoria limite es True")
    #         allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True)
    #     else:
    #         if cateAux and cate.limite == False:
    #             # print("Existe factura y la categoría limite es False")
    #             allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
    #         else:
    #             # print("No existe factura o la categoria limite es False")
    #             allTypes = factType.objects.filter(ingreso=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
    #     cobrarPagar = factType.objects.filter(facCobrar=True)
    
    # else:

    #     if cateAux and request.GET.get("cat") == "Mercancia credito pagada":
    #         allCategories = factCategory.objects.filter(egreso=True).order_by("nombre")
    #     else:
    #         allCategories = factCategory.objects.filter(egreso=True).order_by("nombre").exclude(nombre="Factura cobrada").exclude(nombre="Mercancia credito pagada").exclude(nombre="Factura cobrada (Mayorista)")
    #     if cateAux and cate.limite == True:
    #         allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(facCobrar=True)
    #     else:
    #         allTypes = factType.objects.filter(gasto=True).order_by("nombre").exclude(facCobrada=True).exclude(mercPagada=True).exclude(mercPagar=True).exclude(facCobrar=True)
    #     cobrarPagar = factType.objects.filter(mercPagar=True)

    # allCategories = list(allCategories.values())
    # allTypes = list(allTypes.values())
    # cobrarPagar = list(cobrarPagar.values())

    # return JsonResponse({'cobrarPagar':cobrarPagar,'allTypes':allTypes,'allCategories': allCategories})


    # if request.method == "POST":

    #     nombreaux = request.POST.get("custName")
    #     identificacionaux = request.POST.get("custId")

    #     personProbar = persona.objects.filter(nombre=nombreaux,documento=identificacionaux)

    #     if personProbar:

    #         print("Ya existe")

    #     else:

    #         personAux = persona()
    #         personAux.nombre = nombreaux
    #         personAux.documento = identificacionaux
    #         personAux.save()

    #         personLast = persona.objects.filter(id=personAux.id)
    #         lastPerson = list(personLast.values())
            
    #         return JsonResponse({'lastPerson':lastPerson})

    # return render(request,"spareapp/contDay.html")


































    # return render(request,"spareapp/accountStat.html")


    # {% for tab in tableAux %}
    #             <!-- <tr data-bs-toggle="collapse" data-bs-target="#accordion{{tab.tabTipo|cut:' '}}" class="clickable"> -->
    #             <tr>
    #                 <td class="p-2"><a href="{% url 'contType' tab.tabTipo tod %}">{{tab.tabTipo}}</a></td>
    #                 <td class="p-2">${{tab.tabTotal|floatformat:2}}</td>
    #             </tr>
    #             <!-- {% for fact in allFactures %}
    #             {% if tab.tabTipo == fact.refType %}
    #             <tr style="background-color: rgb(207, 207, 207);">
    #                 <td id="accordion{{tab.tabTipo|cut:' '}}" class="collapse p-2"><a href="">Facture number: {{fact.num}}</a></td>
    #                 <td id="accordion{{tab.tabTipo|cut:' '}}" class="collapse p-2">${{fact.total}}</td>
    #             </tr>
    #             {% endif %}
    #             {% endfor %} -->
    #             {% endfor %}
    
    # <label class="form-check mb-2">Facture type</label>
    # <input onclick="factTypeFun('entry');" class="form-check-input" id="contFacType1" type="radio" name="contFacType" checked>
    # <label class="form-check-label" for="contFacType1">Entry</label>
    # <input onclick="factTypeFun('spending');" class="form-check-input" id="contFacType2" type="radio" name="contFacType">
    # <label class="form-check-label" for="contFacType2">Spending</label>




# CONTDAY ANTES DE SEPARAR EN DOS TABLAS

# {% extends "spareapp/contBase.html" %}

# {% block title %} Day {% endblock %}

# {% block content %}

# {% include "spareapp/contSuperior.html" %}

# <form method="POST" action="">
# {% csrf_token %}
# <div class="container mt-2">
#     <table style="font-size: small;" class="table-bordered invoice">
#         <thead>
#             <tr class="color text-white">
#                 <th class="p-2">Type</th>
#                 <th class="p-2">Total Remanining</th>
#             </tr>
#         </thead>

#         <tbody>

#             {% if tableAux and editPrueba == False %}

#                 {% for tab in tableAux %}
#                 <tr>
#                     {% if tab.tabTipo.include == True %}
#                     <td class="p-2"><a {% if tab.tabTipo.nombre == "TARJETA VISA" or tab.tabTipo.nombre == "TARJETA CLAVE" %} href="{% url 'contTypeTarjeta' tab.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' tab.tabTipo|cut:'/' tod %}" {% endif %}>{{tab.tabTipo}}</a></td>
#                     {% else %}
#                     <td class="p-2"><a {% if tab.tabTipo.nombre == "TARJETA VISA" or tab.tabTipo.nombre == "TARJETA CLAVE" %} href="{% url 'contTypeTarjeta' tab.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' tab.tabTipo|cut:'/' tod %}" {% endif %}>( {{tab.tabTipo}} )</a></td>
#                     {% endif %}
#                     <td class="p-2">${{tab.tabTotal|floatformat:2}}</td>
#                 </tr>
#                 {% endfor %}
                
#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">${{contTotal|floatformat:2}}</td>
#                 </tr>

#                 <!-- <tr>
#                     <td class="p-2 color text-white">Spending total</td>
#                     <td class="p-2">${{spendTotal|floatformat:2}}</td>
#                 </tr> -->

#             {% else %}

#                 {% if editPrueba == True %}

#                     {% for tab in tableAux %}
#                     <tr>
#                         {% if tod %}
#                         <td class="p-2"><a href="{% url 'contType' tab.tabTipo tod %}">{{tab.tabTipo}}</a></td>
#                         {% else %}
#                         <td class="p-2"><a href="{% url 'contType' tab.tabTipo 'today' %}">{{tab.tabTipo}}</a></td>
#                         {% endif %}
#                         {% if tab.tabTipo.manual == True %}
#                         <td class="p-2">$<input value="{{tab.tabTotal}}" id="{{tab.tabTipo.nombre|cut:' '}}Total" name="{{tab.tabTipo.nombre|cut:' '}}Total" type="number"></td>
#                         {% else %}
#                         <td class="p-2">$ {{tab.tabTotal|floatformat:2}}</td>
#                         {% endif %}
#                     </tr>
#                     {% endfor %}
#                     <tr>
#                         <td class="p-2 color text-white">Total</td>
#                         <td class="p-2">${{contTotal|floatformat:2}}</td>
#                     </tr>

#                 {% else %}

#                     {% for tab in allTypes %}
#                     <tr>
#                         {% if tab.include == True %}
#                         <td class="p-2"><a href="">{{tab.nombre}}</a></td>
#                         {% else %}
#                         <td class="p-2"><a href="">(  {{tab.nombre}}  )</a></td>
#                         {% endif %}
#                         {% if tab.manual == True %}
#                         <td class="p-2">$<input id="{{tab.nombre|cut:' '}}Total" name="{{tab.nombre|cut:' '}}Total" type="number"></td>
#                         {% else %}
#                         <td class="p-2">$ 0.00</td>
#                         {% endif %}
#                     </tr>
#                     {% endfor %}
#                     <tr>
#                         <td class="p-2 color text-white">Total</td>
#                         <td class="p-2">$ 0.00</td>
#                     </tr>

#                 {% endif %}

#             {% endif %}

#         </tbody>
#     </table>
# </div>



# <div class="container mt-2">
#     <table style="font-size: small;" class="table-bordered invoice">
        

#         <tbody>

#             {% if tableAux and editPrueba == False %}

#                 {% for tab in tableAux %}
#                 <tr>
#                     {% if tab.tabTipo.include == False %}
#                     <td class="p-2"><a {% if tab.tabTipo.nombre == "TARJETA VISA" or tab.tabTipo.nombre == "TARJETA CLAVE" %} href="{% url 'contTypeTarjeta' tab.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' tab.tabTipo|cut:'/' tod %}" {% endif %}>( {{tab.tabTipo}} )</a></td>
#                     <td class="p-2">${{tab.tabTotal|floatformat:2}}</td>
#                     {% endif %}
#                 </tr>
#                 {% endfor %}
                
#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">${{contTotal|floatformat:2}}</td>
#                 </tr>

#                 <!-- <tr>
#                     <td class="p-2 color text-white">Spending total</td>
#                     <td class="p-2">${{spendTotal|floatformat:2}}</td>
#                 </tr> -->

#             {% else %}

#                 {% if editPrueba == True %}

#                     {% for tab in tableAux %}
#                     <tr>
#                         {% if tod %}
#                         <td class="p-2"><a href="{% url 'contType' tab.tabTipo tod %}">{{tab.tabTipo}}</a></td>
#                         {% else %}
#                         <td class="p-2"><a href="{% url 'contType' tab.tabTipo 'today' %}">{{tab.tabTipo}}</a></td>
#                         {% endif %}
#                         {% if tab.tabTipo.manual == True %}
#                         <td class="p-2">$<input value="{{tab.tabTotal}}" id="{{tab.tabTipo.nombre|cut:' '}}Total" name="{{tab.tabTipo.nombre|cut:' '}}Total" type="number"></td>
#                         {% else %}
#                         <td class="p-2">$ {{tab.tabTotal|floatformat:2}}</td>
#                         {% endif %}
#                     </tr>
#                     {% endfor %}
#                     <tr>
#                         <td class="p-2 color text-white">Total</td>
#                         <td class="p-2">${{contTotal|floatformat:2}}</td>
#                     </tr>

#                 {% else %}

#                     {% for tab in allTypes %}
#                     <tr>
#                         {% if tab.include == False %}
#                         <td class="p-2"><a href="">(  {{tab.nombre}}  )</a></td>
                        
#                         {% if tab.manual == True %}
#                         <td class="p-2">$<input id="{{tab.nombre|cut:' '}}Total" name="{{tab.nombre|cut:' '}}Total" type="number"></td>
#                         {% else %}
#                         <td class="p-2">$ 0.00</td>
#                         {% endif %}
#                         {% endif %}
#                     </tr>
#                     {% endfor %}
#                     <tr>
#                         <td class="p-2 color text-white">Total</td>
#                         <td class="p-2">$ 0.00</td>
#                     </tr>

#                 {% endif %}

#             {% endif %}

#         </tbody>
#     </table>
# </div>
# </form>

# {% endblock %}








# <div class="container mt-2">
#     <table style="font-size: small;" class="table-bordered invoice">
#         <thead>
#             <tr class="color text-white">
#                 <th class="p-2">Tipo de pago</th>
#                 <th style="min-width: 100px;" class="p-2">Total</th>
#             </tr>
#         </thead>

#         <tbody>

#             {% if tableAux %}

#                 {% for tab in tableAux %}
#                 <tr>
#                     {% if tab.tabTipo.include == True %}
#                     <td class="p-2"><a {% if tab.tabTipo.visa == True or tab.tabTipo.clave == True %} href="{% url 'contTypeTarjeta' tab.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' tab.tabTipo|cut:'/' tod %}" {% endif %}>{{tab.tabTipo}}</a></td>
#                     <td class="p-2">${{tab.tabTotal|floatformat:2}}</td>
#                     {% endif %}
#                 </tr>
#                 {% endfor %}
                
#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">${{contTotal|floatformat:2}}</td>
#                 </tr>

#             {% else %}

#                 {% for tab in allTypes %}
#                 <tr>
#                     {% if tab.include == True %}
#                     <td class="p-2"><a href="">{{tab.nombre}}</a></td>
#                     {% if tab.manual == True %}
#                     <td class="p-2">$<input id="{{tab.nombre|cut:' '}}Total" name="{{tab.nombre|cut:' '}}Total" type="number"></td>
#                     {% else %}
#                     <td class="p-2">$ 0.00</td>
#                     {% endif %}
#                     {% endif %}
#                 </tr>
#                 {% endfor %}
#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">$ 0.00</td>
#                 </tr>
#             {% endif %}
#         </tbody>
#     </table>
# </div>


# <div class="container mt-2">
#     <table style="font-size: small;" class="table-bordered invoice">
#         <thead>
#             <tr class="color text-white">
#                 <th class="p-2">Tipo de pago</th>
#                 <th style="min-width: 100px;" class="p-2">Total</th>
#             </tr>
#         </thead>

#         <tbody>

#             {% if tableAux %}

#                 {% for tab in tableAux %}
#                 <tr>
#                     {% if tab.tabTipo.include == False %}
#                     <td class="p-2"><a {% if tab.tabTipo.visa == True or tab.tabTipo.clave == True %} href="{% url 'contTypeTarjeta' tab.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' tab.tabTipo|cut:'/' tod %}" {% endif %}>( {{tab.tabTipo}} )</a></td>
#                     <td class="p-2">${{tab.tabTotal|floatformat:2}}</td>
#                     {% endif %}
#                 </tr>
#                 {% endfor %}

#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">${{noIncludeTotal|floatformat:2}}</td>
#                 </tr>

#             {% else %}
                
#                 {% for tab in allTypes %}
#                 <tr>
#                     {% if tab.include == False %}
#                     <td class="p-2"><a href="">(  {{tab.nombre}}  )</a></td>
#                     {% if tab.manual == True %}
#                     <td class="p-2">$<input id="{{tab.nombre|cut:' '}}Total" name="{{tab.nombre|cut:' '}}Total" type="number"></td>
#                     {% else %}
#                     <td class="p-2">$ 0.00</td>
#                     {% endif %}
#                     {% endif %}
#                 </tr>
#                 {% endfor %}
#                 <tr>
#                     <td class="p-2 color text-white">Total</td>
#                     <td class="p-2">$ 0.00</td>
#                 </tr>

#             {% endif %}

#         </tbody>
#     </table>
# </div>



# tfoot para footer de tabla
# caption para header superior



# // function generatePDF2() {

# //   var tableHeaderText = [...document.querySelectorAll('#invoice thead tr th')].map(thElement => ({ text: thElement.textContent, style: 'tableHeader' }));
# //   var tableRowCells = [...document.querySelectorAll('#invoice tbody tr td')].map(tdElement => ({ text: tdElement.textContent, style: 'tableData' }));
# //   var tableDataAsRows = tableRowCells.reduce((rows, cellData, index) => {
# //     if (index % 4 === 0) {
# //       rows.push([]);
# //     }
# //     rows[rows.length - 1].push(cellData);
# //     return rows;
# //   }, []);
    

# //   var docDefinition = {
# //     // header: { text: 'MLB World Series Winners', alignment: 'center' },
# //     // footer: function(currentPage, pageCount) { return ({ text: `Page ${currentPage} of ${pageCount}`, alignment: 'center' }); },
# //     content: [
# //       {
# //         style: 'tableExample',
# //         table: {
# //           headerRows: 1,
# //           body: [
# //             tableHeaderText,
# //             ...tableDataAsRows,
# //           ]
# //         },
# //         layout: {
# //           fillColor: function(rowIndex) {
# //             if (rowIndex === 0) {
# //               return '#0f4871';
# //             }
# //             return (rowIndex % 2 === 0) ? '#f2f2f2' : null;
# //           }
# //         },
# //       },
# //     ],
# //     styles: {
# //       tableExample: {
# //         margin: [0, 20, 0, 80],
# //       },
# //       tableHeader: {
# //         margin: 12,
# //         color: 'white',
# //       },
# //       tableData: {
# //         margin: 12,
# //       },
# //     },
# //   };
# // //   pdfMake.createPdf(docDefinition).download('Table.pdf');
# //   pdfMake.createPdf(docDefinition).download('MLB World Series Winners');
# // }

# // const area = $("#invoice")[0].getBoundingClientRect()
# // function generatePDF2(){
# //     html2canvas($("#invoice")[0],{
# //         onrendered:function(canvas){
# //             var data=canvas.toDataURL();
# //             var docDefinition={
# //                 content:[{
# //                     image:data,
# //                     // scrollX: 0,
# //                     // scrollY: 0,
# //                     // width: area.width
# //                     // height: area.height
# //                     width:500
# //                 }]
# //             };
# //             pdfMake.createPdf(docDefinition).download("Table.pdf");
# //         }
# //     })
# // }


















# <!-- {% for tab in cantAuxOp %}
# {% if tab.principal == True %}
# <div class="container">
#     <table style="font-size: small;" class="table-bordered invoice mb-2">
#         <thead>
#             <tr style="background-color: #F5F087" class="">
#                 <th colspan="2" class="p-2">{{tab.tabNombre}}  |  Día: {{tod}}</th>
#             </tr>
#             <tr style="background-color: #7C8FEC;">
#                 <th class="p-2">Tipo de pago</th>
#                 <th style="width: 350px;" class="p-2">Total</th>
#             </tr>
#         </thead>
#         <tbody>
#             {% for table in tableAuxOp %}
#             {% if table.tabNombre == tab.tabNombre %}
#             <tr>
#                 <td class="p-2"><a {% if table.tabTipo.visa == True or table.tabTipo.clave == True %} href="{% url 'contTypeTarjeta' table.tabTipo|cut:'/' tod %}" {% else %} href="{% url 'contType' table.tabTipo|cut:'/' tod %}" {% endif %}>{{table.tabTipo}}</a></td>
#                 <td class="p-2">${{table.tabTotal|floatformat:2}}</td>
#             </tr>
#             {% endif %}
#             {% endfor %}
#             {% for key,value in totalParcialOp.items %}
#                 {% if key == tab.tabNombre %}
#                 <tr style="background-color: #E77C7C;" class="">
#                     <td colspan="2" class="p-2">TOTAL: ${{value|floatformat:2}}</td>
#                 </tr>
#                 {% endif %}
#             {% endfor %}
#         </tbody>
#     </table>
# </div>
# {% endif %}
# {% endfor %} -->






# {% for n in cantAuxEmpty %}
# {% if n.principal == True %}

# <div class="mb-3" id="">
# <div class="container">
#     <table style="font-size: small;" class="table-bordered invoice">
#         <thead>
#             <tr style="background-color: #F5F087" class="">
#                 <th colspan="2" class="p-2">{{n.tabNombre}}  |  Día: {{tod}}</th>
#             </tr>
#             <tr style="background-color: #7C8FEC;">
#                 <th class="p-2">Tipo de pago</th>
#                 <th style="width: 350px;" class="p-2">Total</th>
#             </tr>
#         </thead>

#         <tbody>

#             {% if tableAux2Empty %}

#                 {% for tab in tableAux2Empty %}
#                 {% if tab.tabNombre == n.tabNombre %}
#                 <tr>
#                     <td class="p-2"><a>{{tab.tabTipo__nombre}}</a></td>
#                     <td class="p-2">$0.00</td>
#                 </tr>
#                 {% endif %}
#                 {% endfor %}
#                 <tr style="background-color: #E77C7C;" class="">
#                     <td colspan="2" class="p-2">TOTAL: $0.00</td>
#                 </tr>
#             {% endif %}

#         </tbody>
#     </table>
# </div>
# </div>

# {% endif %}
# {% endfor %}

def totalAccountStat(request):

    acumTotal = 0
    
    print(request.GET)
    col = request.GET.get("columna")
    filtro = request.GET.get("filtro")
    dateFrom=request.GET.get("dateFrom")
    dateTo=request.GET.get("dateTo")
    checkAll=request.GET.get("all")
    month=request.GET.get("month")
    soloFac=request.GET.get("soloFac")
    range=request.GET.get("range")
    # Es el id
    nombre=request.GET.get("nombre")
    nombre=persona.objects.get(id=request.GET.get("nombre"))

    if checkAll:
        factureName = factura.objects.filter(refPersona__nombre__icontains=nombre).order_by("fechaCreado","id")
    if month:
        factureName = factura.objects.filter(refPersona__nombre__icontains=nombre).order_by("fechaCreado","id")
    if range:
        factureName = factura.objects.filter(refPersona__nombre__icontains=nombre,fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).order_by("fechaCreado","id")

    # if col == "Fecha":
    #     factureName = factura.objects.filter(fechaCreado__date__gte=dateFrom,fechaCreado__date__lte=dateTo).order_by("fechaCreado","id")
    if col == "Cliente":
        factureName = factureName.filter(refPersona__nombre__icontains=filtro).order_by("fechaCreado","id")
    if col == "Fact":
        factureName = factureName.filter(num__icontains=filtro).order_by("fechaCreado","id")
    if col == "Categoría":
        factureName = factureName.filter(refCategory__nombre__icontains=filtro).order_by("fechaCreado","id")
    if col == "Tipo":
        factureName = factureName.filter(refType__nombre__icontains=filtro).order_by("fechaCreado","id")
    if col == "Monto":
        factureName = factureName.filter(monto__icontains=filtro).order_by("fechaCreado","id")
    if col == "Balance":
        factureName = factureName.filter(monto__icontains=filtro).order_by("fechaCreado","id")
    if col == "Nota":
        factureName = factureName.filter(note__icontains=filtro).order_by("fechaCreado","id")

    acumTotal = 0

    if factureName:

        for fac in factureName:

            if fac.refType.ingreso == True and fac.refType.facCobrar == False or fac.refType.mercPagar == True:

                acumTotal = acumTotal + abs(fac.total)
            
            else:

                acumTotal = acumTotal - abs(fac.total)

    return JsonResponse({"acumTotal":acumTotal})


